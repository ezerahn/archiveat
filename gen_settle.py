# -*- coding: utf-8 -*-
"""
정산 리포트 생성기 — 여러 마스터를 읽어 정산 현황으로
입력: class-master-all-v3.xlsx + 갤러리아_강좌이력.xlsx 등 제휴사 마스터
출력: class-settle-report.html
사용법: python gen_settle.py   ※ gen_class.py 를 먼저 돌린 뒤 실행

════════════════════════════════════════════════════════════════
★ 수정 전 반드시 class-guide.html 을 읽으세요. 확정 규칙이 거기 있습니다.
════════════════════════════════════════════════════════════════

■ 이 화면의 역할
    클래스 목록 = 얼마 벌었나 (완료된 것 중심)
    정산 리포트 = 돈이 언제 오가나 (아직 안 받은 돈 · 안 준 돈이 위에 크게)

■ 정산 상태 판정
    클래스 마스터 : 입금상태 · 지급상태 칸을 그대로 씀
    제휴사 마스터 : 입금상태·지급상태 칸(W·X열)을 그대로 씀
                    (대응 회차 없는 이체만 「확인 대기」로 따로 표시)

■ 상태 두 갈래 — 섞지 말 것
    수업 상태 : 기획중 → 제안중/검토중 → 확정 → 모집중 → 수업완료 / 폐강
    정산 상태 : 입금대기 ↔ 입금완료 · 지급대기 ↔ 지급완료
    「정산완료」는 수업 상태가 아니다. 정산 여부는 입금·지급 배지로 본다

■ 지급일 규칙 (앞으로 기준)
    수업일 1~15일  → 당월 25일
    수업일 16~말일 → 익월 10일
    ※ 갤러리아 오프라인은 발주처입금연동이라 익월 23~25일이었음 (종료된 규칙)
"""
import os, json, datetime, re
from openpyxl import load_workbook

KST = datetime.timezone(datetime.timedelta(hours=9))
UPDATED = datetime.datetime.now(KST).strftime("%Y.%m.%d %H:%M")
TODAY = datetime.datetime.now(KST).date()
VERSION = "v56"
CHAT = "https://claude.ai/chat/909e7f28-5718-4bde-8997-e37348632306"
CO = dict(name="(주) 공공의주방", ceo="안상미", tel="010.2770.5538", biz="368-81-00400")

def D(v):
    if isinstance(v,datetime.datetime): return v.date()
    if isinstance(v,datetime.date): return v
    if isinstance(v,str) and len(v)>=10:
        try: return datetime.date(int(v[:4]),int(v[5:7]),int(v[8:10]))
        except: return None
    return None
def due_partner(d):
    """제휴사 — 수업 익월 25일 (익월 10일 정산금 입금 확인 후 지급)"""
    if not d: return None
    ny,nm=(d.year+1,1) if d.month==12 else (d.year,d.month+1)
    return datetime.date(ny,nm,25)
def due(d):
    if not d: return None
    if d.day<=15: return datetime.date(d.year,d.month,25)
    ny,nm=(d.year+1,1) if d.month==12 else (d.year,d.month+1)
    return datetime.date(ny,nm,10)

ROWS=[]; PEND=[]

# ══ ① 클래스 마스터 ════════════════════════════
CM='class-master-all-v3.xlsx'
if os.path.exists(CM):
    wb=load_workbook(CM,data_only=True); ws=wb['회차']
    h=[c.value for c in ws[1]]
    cnt={}
    if '수강생' in wb.sheetnames:
        for r in wb['수강생'].iter_rows(min_row=2,values_only=True):
            if r and r[0]: cnt[str(r[0]).strip()]=cnt.get(str(r[0]).strip(),0)+1
    for r in ws.iter_rows(min_row=2,values_only=True):
        if not r[0]: continue
        d=dict(zip(h,r)); cid=str(d['회차ID']).strip()
        dt=D(d.get('수업일'))
        pax=cnt.get(cid,0) or (d.get('신청인원') or 0) or (d.get('최소인원') or 0)
        price=d.get('판매가') or 0
        rev=price*pax
        sup=rev/1.1 if d.get('부가세구분')=='포함' else rev
        cost=((d.get('재료비(1인)') or 0)+(d.get('패킹비') or 0)+(d.get('배송비') or 0))*pax
        pay=(d.get('강사비') or 0)*pax
        state=str(d.get('수업상태') or '')
        ROWS.append(dict(
            id=cid, partner=str(d.get('발주처') or '—'), dt=dt, name=str(d.get('수업명') or ''),
            kind=str(d.get('구분') or ''), teacher=str(d.get('선생님') or ''),
            held='폐강' if state in ('폐강','취소','무산') else '개최',
            state=state, pax=pax, rev=round(rev), base=round(sup),
            fee=0, recv=round(rev), recvx=round(sup),
            pay=round(pay), payout=round(pay), cost=round(cost),
            ours=round(sup-cost-pay),
            paid_in=str(d.get('입금상태') or '입금대기'),
            paid_out=str(d.get('지급상태') or '지급대기'),
            due=D(d.get('정산예정일')) or due(dt),
            paykind=str(d.get('지급방식') or '미확인'), src='클래스 마스터',
            price=round(price), vatin=str(d.get('부가세구분') or '포함'),
            rate=0, teachbase=round((d.get('강사비') or 0)*pax), teachrate='인당 정액',
            wt=0, addvat=0,
            costdetail=[['재료비',round((d.get('재료비(1인)') or 0)*pax),f"인당 {round(d.get('재료비(1인)') or 0):,}원 × {pax}명"],
                        ['패킹비',round((d.get('패킹비') or 0)*pax),f"인당 {round(d.get('패킹비') or 0):,}원 × {pax}명"],
                        ['배송비',round((d.get('배송비') or 0)*pax),f"인당 {round(d.get('배송비') or 0):,}원 × {pax}명"]]))

# ══ ② 제휴사 마스터 ════════════════════════════
PARTNERS=[('갤러리아_강좌이력.xlsx','갤러리아광교','오프라인')]
UNMATCHED=[]
for f,pname,kind in PARTNERS:
    if not os.path.exists(f): continue
    wb=load_workbook(f,data_only=True); ws=wb['회차이력']
    H=[str(ws.cell(row=5,column=c).value).replace('\n','') for c in range(1,25)]
    for r in range(6,ws.max_row+1):
        a=ws.cell(row=r,column=1).value
        if not a or str(a)=='합계': continue
        d={H[c-1]:ws.cell(row=r,column=c).value for c in range(1,25)}
        dt=D(d['날짜'])
        ROWS.append(dict(
            id=str(d['회차ID']), partner=pname, dt=dt, name=str(d['클래스명'] or ''),
            kind=kind, teacher=str(d['선생님'] or ''), held=str(d['개최'] or ''),
            state='수업완료' if d['개최']=='개최' else '폐강',
            pax=d['인원'] or 0, rev=round(d['클래스매출'] or 0),
            base=round((d['클래스매출'] or 0)/1.1),
            fee=round(d['제휴사 수수료'] or 0), recv=round(d['제휴사 입금'] or 0),
            recvx=round((d['제휴사 입금'] or 0)/1.1),
            pay=round((d['제휴사 입금'] or 0)/1.1-(d['PK 수익'] or 0)),
            payout=round(d['선생님 정산'] or 0), cost=0,
            ours=round(d['PK 수익'] or 0),
            paid_in=str(d.get('입금상태') or '입금완료'), paid_out=str(d.get('지급상태') or '지급완료'),
            due=due_partner(dt),
            paykind=str(d.get('지급방식') or '—'), src=pname,
            price=0, vatin='포함', rate=0, teachbase=0, teachrate='75%',
            wt=0, addvat=0, costdetail=[]))
    # 명세에서 정산서용 상세 채우기
    det={}
    for sh,cid,cp,cr,cw,cf in (('명세원본-2025',17,7,10,15,16),('명세원본-2024',1,9,22,32,34)):
        if sh not in wb.sheetnames: continue
        w=wb[sh]
        for r in range(5,w.max_row+1):
            g=w.cell(row=r,column=cid).value
            if not g: continue
            e=det.setdefault(str(g),dict(price=0,rate=0,wt=0,fin=0))
            e['price']=max(e['price'],w.cell(row=r,column=cp).value or 0)
            e['rate']=max(e['rate'],w.cell(row=r,column=cr).value or 0)
            e['wt']+=w.cell(row=r,column=cw).value or 0
            e['fin']+=w.cell(row=r,column=cf).value or 0
    for x in ROWS:
        if x['src']!=pname: continue
        e=det.get(x['id'])
        if not e: continue
        x['price']=round(e['price']); x['rate']=round(e['rate']*1000)/10; x['wt']=round(e['wt'])
        x['teachbase']=round(x['rev']/1.1)
        x['addvat']=round(e['fin']-(x['rev']/1.1*0.75-e['wt'])) if e['wt']==0 and e['fin'] else 0

    # 대응 없는 이체
    if '지급내역' in wb.sheetnames:
        w2=wb['지급내역']
        for r in range(6,w2.max_row+1):
            a=w2.cell(row=r,column=1).value
            if not a or not str(a).startswith('U-'): continue
            UNMATCHED.append(dict(id=str(a), partner=pname,
                who=str(w2.cell(row=r,column=3).value or ''),
                dt=D(w2.cell(row=r,column=11).value),
                amt=round(w2.cell(row=r,column=12).value or 0),
                note=str(w2.cell(row=r,column=15).value or '')))

ROWS.sort(key=lambda x:(x['dt'] or datetime.date(1900,1,1), x['partner']))
for x in ROWS: x['dt']=str(x['dt']) if x['dt'] else ''; x['due']=str(x['due']) if x['due'] else ''
for x in UNMATCHED: x['dt']=str(x['dt']) if x['dt'] else ''
DATA=json.dumps({'rows':ROWS,'un':UNMATCHED,'today':str(TODAY)},ensure_ascii=False)

CSS = """
:root{--bg:#FAF5EC;--ink:#241F1B;--sub:#6B6256;--line:#E7DCC8;--card:#FFFDF9;
--coral:#FF5019;--point:#B4A032;--wait:#9A8F7C;--mute:#ADA294;--ok:#5E7360;--blue:#3D6B8E}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--ink);font-family:Pretendard,-apple-system,"Apple SD Gothic Neo",sans-serif;
line-height:1.55;-webkit-font-smoothing:antialiased;padding-bottom:70px;font-size:15px;
-webkit-text-size-adjust:100%;text-size-adjust:100%}
.wrap{max-width:1400px;margin:0 auto;padding:0 20px}
.topbar{display:flex;justify-content:space-between;align-items:flex-start;padding:18px 0;gap:12px;flex-wrap:wrap}
.pill{display:inline-flex;align-items:center;gap:4px;font-size:13px;text-decoration:none;color:var(--ink);
background:var(--card);border:1px solid var(--line);border-radius:20px;padding:5px 13px;margin-right:8px}
.pill:hover{border-color:var(--point);color:var(--point)}
.crumb{font-size:12px;color:var(--mute);text-align:right;line-height:1.7}
.crumb .meta{font-family:"DM Mono",monospace;font-size:11px}
.head{padding:14px 0 20px;border-bottom:1px solid var(--line);margin-bottom:18px}
h1{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-weight:800;font-size:40px;letter-spacing:-.02em}
.lead{font-size:15px;color:var(--sub);margin-top:12px}
.now{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:16px 20px;margin-bottom:12px}
.now-h{display:flex;align-items:baseline;gap:10px;margin-bottom:12px;flex-wrap:wrap}
.now-h b{font-family:Hahmlet,serif;font-size:19px}
.now-h span{font-size:12px;color:var(--wait)}
.ncards{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px}
.nc{background:#FCFAF4;border:1px solid var(--line);border-radius:12px;padding:12px 15px}
.nc.hot{border-color:var(--coral);background:#FFF1EB}
.nc .k{font-size:12px;color:var(--sub)}
.nc .v{font-family:"DM Mono",monospace;font-size:21px;font-weight:700;margin-top:3px;letter-spacing:-.02em}
.nc .u{font-size:11px;color:var(--wait);margin-top:2px}
.nc.hot .v{color:var(--coral)}
.sec{margin-top:22px}
h2{font-family:Hahmlet,serif;font-size:20px;margin-bottom:4px}
.h2d{font-size:12.5px;color:var(--wait);margin-bottom:10px}
.bar{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:8px}
.bar b{font-size:12px;color:var(--mute);font-weight:500;min-width:44px}
.chip{font-size:13px;padding:4px 12px;border-radius:20px;border:1px solid var(--line);
background:#fff;color:var(--sub);cursor:pointer;font-family:inherit}
.chip.on{background:var(--coral);border-color:var(--coral);color:#fff}
.srch{font-family:inherit;font-size:13.5px;padding:7px 12px;border:1px solid var(--line);
border-radius:9px;background:#fff;color:var(--ink);width:280px}
.cnt{font-family:"DM Mono",monospace;font-size:12px;color:var(--mute);margin:10px 0}
.tw{overflow-x:auto;background:var(--card);border:1px solid var(--line);border-radius:12px}
table{border-collapse:collapse;width:100%;font-size:12.5px;min-width:1080px}
th{background:#F1EFE9;font-size:11px;font-weight:700;color:var(--sub);padding:7px 6px;
text-align:center;white-space:nowrap;line-height:1.3;cursor:pointer;user-select:none}
th:hover{color:var(--coral)}
tr.grp th{background:#E4DFD0;font-size:12px;color:#3A322A;border-bottom:1px solid var(--line);
border-right:1.5px solid #CFC6B0;cursor:default}
tr.grp th:last-child{border-right:none}
td{padding:6px 7px;border-top:1px solid #F0EADC;vertical-align:middle;line-height:1.4}
td.n{font-family:"DM Mono",monospace;text-align:right;white-space:nowrap;font-size:12px;padding-right:12px}
td.c{text-align:center;white-space:nowrap}
td.nm{max-width:200px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
tr:hover td{background:#F7F3E9}
tr.tot td{background:#F1EFE9;font-weight:700;border-top:1.5px solid var(--line)}
tr.late td{background:#FFF6F2}
.bd{display:inline-block;font-size:10.5px;padding:2px 7px;border-radius:20px;white-space:nowrap}
.b-cor{background:#FFE7DF;color:var(--coral);font-weight:500}
.b-mut{background:#F1EFE9;color:#6B6256}
.b-yel{background:#F5B21F;color:#4A3405;font-weight:500}
.b-grn{background:#E4EAE2;color:var(--ok)}
.b-blu{background:#E6EEF4;color:var(--blue)}
.b-off{background:#EFEDEA;color:var(--mute)}
.note{font-size:12.5px;color:var(--wait);margin-top:11px;line-height:1.75}
.files{padding:24px 0 10px;border-top:1px solid var(--line);margin-top:28px}
.files-l{font-family:"DM Mono",monospace;font-size:11px;letter-spacing:.1em;color:var(--mute);margin-bottom:10px}
.fbtn{display:inline-flex;align-items:center;gap:7px;font-size:13px;padding:10px 16px;border-radius:10px;
border:1px solid var(--line);background:#fff;color:var(--sub);text-decoration:none;margin-right:8px}
.fbtn:hover{border-color:var(--coral);color:var(--coral)}
.foot{margin-top:22px;font-family:"DM Mono",monospace;font-size:11px;color:var(--wait);text-align:center}
@media(max-width:820px){h1{font-size:29px}.crumb{text-align:left}.srch{width:100%}}
.docbtn{font-family:inherit;font-size:11.5px;padding:4px 10px;border-radius:7px;border:1px solid var(--line);
background:#fff;color:var(--sub);cursor:pointer;white-space:nowrap}
code{font-family:"DM Mono",monospace;font-size:11.5px;background:#F1EFE9;padding:1px 5px;border-radius:4px}
.docbtn:hover{border-color:var(--coral);color:var(--coral)}
#modal{display:none;position:fixed;inset:0;z-index:90}
.mbg{position:absolute;inset:0;background:rgba(36,31,27,.42)}
.mbox{position:relative;width:calc(100% - 32px);max-width:720px;margin:20px auto;background:#fff;
border-radius:14px;max-height:calc(100vh - 40px);display:flex;flex-direction:column;overflow:hidden;
box-shadow:0 18px 50px rgba(0,0,0,.22)}
.mtop{display:flex;justify-content:space-between;align-items:center;gap:10px;
padding:11px 14px;border-bottom:1px solid var(--line);background:var(--bg);flex-wrap:wrap}
.mtabs{display:flex;gap:5px}
.mtab{font-family:inherit;font-size:13px;font-weight:600;padding:7px 15px;border-radius:9px;
border:1px solid var(--line);background:#fff;color:var(--wait);cursor:pointer}
.mtab.on{background:#5C1A12;border-color:#5C1A12;color:#fff}
.macts{display:flex;gap:5px}
.mact{font-family:inherit;font-size:12.5px;padding:7px 12px;border-radius:9px;
border:1px solid var(--line);background:#fff;color:var(--sub);cursor:pointer;white-space:nowrap}
.mact.hot{background:#8C2A1E;border-color:#8C2A1E;color:#fff}
.mact:hover{border-color:var(--coral)}
.mbody{overflow-y:auto;padding:24px 26px 20px;position:relative}
.stamp{position:absolute;top:22px;right:26px;border:2px solid #8C2A1E;color:#8C2A1E;
font-size:11.5px;font-weight:700;padding:4px 10px;border-radius:4px;letter-spacing:.06em}
.dt{font-family:Hahmlet,serif;font-size:27px;letter-spacing:.13em;margin-bottom:4px}
.dsub{font-size:14px;color:var(--sub);margin-bottom:13px}
.dmeta{width:100%;border-collapse:collapse;font-size:12.5px;margin-bottom:12px;table-layout:fixed}
.dmeta td{padding:5px 6px;border-bottom:1px solid #EFEAE0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.dmeta td:nth-child(odd){color:var(--wait);width:62px;font-size:11.5px}
.dnote{font-size:12px;color:var(--wait);margin-bottom:10px}
.dtbl{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:13px;table-layout:fixed}
.dtbl th{background:transparent;border-top:1.5px solid var(--ink);border-bottom:1px solid var(--ink);
padding:7px 6px;font-size:11.5px;color:var(--sub);text-align:left;cursor:default}
.dtbl th:not(:first-child){text-align:right}
.dtbl td{padding:8px 6px;border-bottom:1px solid #EFEAE0;overflow:hidden}
.dtbl td:first-child{white-space:nowrap;text-overflow:ellipsis}
.dtbl th:nth-child(2),.dtbl td:nth-child(2){width:20%}
.dtbl th:nth-child(3),.dtbl td:nth-child(3){width:14%}
.dtbl th:nth-child(4),.dtbl td:nth-child(4){width:24%}
.dtbl.two th:nth-child(2),.dtbl.two td:nth-child(2){width:34%}
.dtbl.three th:nth-child(2),.dtbl.three td:nth-child(2){width:38%}
.dtbl.three th:nth-child(3),.dtbl.three td:nth-child(3){width:26%}
.dtbl.three td:nth-child(2){white-space:normal;word-break:keep-all}
.dtbl td.n{font-family:"DM Mono",monospace;text-align:right}
.dtbl td.c{text-align:center}
.dtbl tr.sum td{border-top:1px solid var(--ink);font-weight:600}
.dtbl tr.tot td{background:#EDE9DE;font-weight:700;font-size:14px;border-bottom:1.5px solid var(--ink)}
.dh3{font-family:Hahmlet,serif;font-size:15px;margin:14px 0 6px}
.dpay{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:14px}
.dpay td{padding:6px 6px;border-bottom:1px solid #EFEAE0}
.dpay td:first-child{color:var(--wait);width:88px;font-size:12px}
.dpay td.n{font-family:"DM Mono",monospace}
.dpay td.b{font-weight:700}
.dfoot{display:flex;justify-content:space-between;gap:18px;margin-top:16px;
padding-top:13px;border-top:1px solid var(--line);flex-wrap:wrap}
.dfl{font-size:11px;color:var(--wait);line-height:1.8;flex:1;min-width:180px}
.dco{border-collapse:collapse;font-size:12px;min-width:210px}
.dco td{padding:5px 8px;border-bottom:1px solid #EFEAE0}
.dco td:first-child{color:var(--wait);width:60px;font-size:11px}
.dco .con{font-weight:700;font-size:13.5px;border-bottom:1px solid var(--ink);padding-bottom:7px}
@media(max-width:560px){
  .mbox{margin:0;max-width:100%;height:100%;max-height:100%;border-radius:0}
  .mtop{padding:9px 11px;gap:7px}
  .mtabs{flex:1}
  .mtab{font-size:12.5px;padding:7px 11px;flex:1}
  .mact{font-size:12px;padding:7px 10px}
  .macts .mact.wide{display:none}
  .mbody{padding:20px 18px 18px}
  .stamp{top:17px;right:18px;font-size:11px;padding:4px 8px}
  .dt{font-size:24px;letter-spacing:.1em}
  .dsub{font-size:13.5px;margin-bottom:13px}
  .dmeta{font-size:12.5px;margin-bottom:12px;table-layout:fixed}
  .dmeta td{padding:5px 4px}
  .dmeta td:nth-child(odd){width:52px;font-size:11px}
  .dmeta td:nth-child(2){width:36%}
  .dmeta td:nth-child(4){width:auto}
  .dnote{font-size:11.5px;margin-bottom:10px}
  .dtbl{font-size:12px;margin-bottom:13px;table-layout:fixed}
  .dtbl th{padding:6px 4px;font-size:11px}
  .dtbl td{padding:7px 4px;word-break:keep-all}
  .dtbl th:first-child,.dtbl td:first-child{width:auto}
  .dtbl th:nth-child(2),.dtbl td:nth-child(2){width:22%}
  .dtbl th:nth-child(3),.dtbl td:nth-child(3){width:15%}
  .dtbl th:nth-child(4),.dtbl td:nth-child(4){width:27%}
  .dtbl.two th:nth-child(2),.dtbl.two td:nth-child(2){width:40%}
  .dtbl.three th:nth-child(2),.dtbl.three td:nth-child(2){width:40%}
  .dtbl.three th:nth-child(3),.dtbl.three td:nth-child(3){width:30%}
  .dtbl td.n{font-size:11.5px;letter-spacing:-.03em}
  .dtbl tr.tot td{font-size:13.5px}
  .dh3{font-size:14.5px;margin:14px 0 6px}
  .dpay{font-size:12.5px;margin-bottom:14px}
  .dpay td{padding:6px 5px}
  .dpay td:first-child{width:80px;font-size:11.5px}
  .dfoot{gap:14px;margin-top:16px;padding-top:13px;flex-direction:column}
  .dfl{font-size:11px;line-height:1.8;min-width:0}
  .dco{font-size:11.5px;min-width:0;width:100%}
  .dco td{padding:5px 7px}
  .dco td:first-child{width:58px;font-size:11px}
  .dco .con{font-size:13px;padding-bottom:7px}
}
@media print{
  body{background:#fff;padding:0}
  .wrap{display:none !important}
  #modal{display:block !important;position:static}
  .mbg,.mtop{display:none !important}
  .mbox{max-width:100%;margin:0;box-shadow:none;border-radius:0;max-height:none;display:block}
  .mbody{overflow:visible;padding:0}
  .stamp{top:0;right:0}
}
"""

JS = r"""
var RAW=__DATA__, D=RAW.rows, UN=RAW.un, TODAY=RAW.today;
var n=function(v){var x=Math.round(Number(v)||0);return x?x.toLocaleString('ko-KR'):'—';};
var e=function(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){
 return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});};
var F={p:'전체',st:'전체',q:''};
var SORT={key:'due',asc:true};
var OPEN_IN=function(x){return x.paid_in!=='입금완료';};
var OPEN_OUT=function(x){return x.paid_out!=='지급완료';};
var LIVE=function(x){return x.held==='개최'&&(OPEN_IN(x)||OPEN_OUT(x));};
var PRE={'확정':1,'모집중':1,'개강':1,'신청중':1};
var STALE=function(x){return x.held==='개최'&&PRE[x.state]&&x.dt&&x.dt<TODAY;};
function stB(s){
  var m={'수업완료':'b-grn','정산완료':'b-grn','확정':'b-cor','모집중':'b-cor','개강':'b-cor',
         '신청중':'b-yel','기획중':'b-mut','제안중':'b-mut','검토중':'b-mut','폐강':'b-off'};
  return '<span class="bd '+(m[s]||'b-mut')+'">'+e(s)+'</span>';}
function payB(v,ok){return '<span class="bd '+(ok?'b-grn':'b-yel')+'">'+e(v)+'</span>';}

function now(){
  var live=D.filter(LIVE);
  var inW=live.filter(OPEN_IN), outW=live.filter(OPEN_OUT);
  var late=live.filter(function(x){return x.due&&x.due<TODAY;});
  var stale=D.filter(STALE);
  var s={inA:0,outA:0};
  inW.forEach(function(x){s.inA+=+x.recv||0;});
  outW.forEach(function(x){s.outA+=+x.payout||0;});
  var unA=0; UN.forEach(function(x){unA+=+x.amt||0;});
  document.getElementById('ncards').innerHTML=
   '<div class="nc hot"><div class="k">받을 돈</div><div class="v">'+n(s.inA)+'</div><div class="u">미입금 '+inW.length+'건</div></div>'+
   '<div class="nc hot"><div class="k">줄 돈</div><div class="v">'+n(s.outA)+'</div><div class="u">미지급 '+outW.length+'건</div></div>'+
   '<div class="nc'+(late.length?' hot':'')+'"><div class="k">지급일 경과</div><div class="v">'+late.length+'</div><div class="u">건</div></div>'+
   '<div class="nc"><div class="k">진행 중 회차</div><div class="v">'+live.length+'</div><div class="u">정산 미완료</div></div>'+
   '<div class="nc'+(UN.length?' hot':'')+'"><div class="k">확인 대기 이체</div><div class="v">'+UN.length+'</div><div class="u">'+n(unA)+'원</div></div>'+
   (stale.length?'<div class="nc hot"><div class="k">상태 갱신 필요</div><div class="v">'+stale.length+'</div><div class="u">수업일 지남</div></div>':'');
  // 진행 중 표
  if(!live.length){document.getElementById('livewrap').innerHTML=
    '<div class="tw" style="padding:20px;text-align:center;color:var(--mute);font-size:13px">정산이 필요한 회차가 없습니다</div>';}
  else{
    live.sort(function(a,b){return String(a.due).localeCompare(String(b.due));});
    var h='<div class="tw"><table style="min-width:900px"><thead><tr>'+
      '<th>회차ID</th><th>제휴사</th><th>수업일</th><th>클래스명</th><th>선생님</th><th>상태</th>'+
      '<th>지급 예정일</th><th>받을 돈</th><th>입금</th><th>줄 돈</th><th>지급</th><th></th></tr></thead><tbody>';
    live.forEach(function(x){
      var lt=x.due&&x.due<TODAY;
      h+='<tr'+(lt?' class="late"':'')+'>'+
        '<td class="c" style="font-family:\'DM Mono\',monospace;font-size:11px;color:var(--mute)">'+e(x.id)+'</td>'+
        '<td class="c">'+e(x.partner)+'</td><td class="c">'+e(x.dt)+'</td>'+
        '<td class="nm" title="'+e(x.name)+'">'+e(x.name)+'</td>'+
        '<td class="c">'+e(x.teacher)+'</td>'+
        '<td class="c">'+stB(x.state)+(STALE(x)?'<br><span class="bd b-yel" style="margin-top:3px">수업일 지남</span>':'')+'</td>'+
        '<td class="c">'+e(x.due)+(lt?' <span class="bd b-cor">경과</span>':'')+'</td>'+
        '<td class="n">'+(OPEN_IN(x)?n(x.recv):'—')+'</td>'+
        '<td class="c">'+payB(x.paid_in,!OPEN_IN(x))+'</td>'+
        '<td class="n">'+(OPEN_OUT(x)?n(x.payout):'—')+'</td>'+
        '<td class="c">'+payB(x.paid_out,!OPEN_OUT(x))+'</td>'+
        '<td class="c"><button class="docbtn" onclick="doc(\''+x.id+'\')">📄 정산서</button></td></tr>';});
    document.getElementById('livewrap').innerHTML=h+'</tbody></table></div>';
  }
  // 확인 대기 이체
  if(!UN.length){document.getElementById('unwrap').innerHTML='';}
  else{
    var u='<div class="tw"><table style="min-width:760px"><thead><tr>'+
      '<th>ID</th><th>제휴사</th><th>이체 명의</th><th>이체일</th><th>금액</th><th>비고</th></tr></thead><tbody>';
    UN.forEach(function(x){
      u+='<tr><td class="c" style="font-family:\'DM Mono\',monospace;font-size:11px">'+e(x.id)+'</td>'+
        '<td class="c">'+e(x.partner)+'</td><td class="c">'+e(x.who)+'</td>'+
        '<td class="c">'+e(x.dt)+'</td><td class="n">'+n(x.amt)+'</td>'+
        '<td style="font-size:11.5px;color:var(--wait)">'+e(x.note)+'</td></tr>';});
    document.getElementById('unwrap').innerHTML=u+'</tbody></table></div>';
  }
}
function filt(){
  return D.filter(function(x){
    if(x.held!=='개최')return false;
    if(F.p!=='전체'&&x.partner!==F.p)return false;
    if(F.st==='정산완료'&&(OPEN_IN(x)||OPEN_OUT(x)))return false;
    if(F.st==='정산 대기'&&!(OPEN_IN(x)||OPEN_OUT(x)))return false;
    if(F.q){var q=F.q.toLowerCase();
      if((x.name+' '+x.dt+' '+x.teacher+' '+x.id+' '+x.partner).toLowerCase().indexOf(q)<0)return false;}
    return true;});
}
function render(){
  var L=filt().slice();
  var k=SORT.key,a=SORT.asc?1:-1;
  L.sort(function(x,y){
    var vx=x[k],vy=y[k];
    if(typeof vx==='number'||typeof vy==='number')return ((vx||0)-(vy||0))*a;
    return String(vx).localeCompare(String(vy),'ko')*a;});
  var t={rev:0,fee:0,recv:0,pay:0,payout:0,ours:0};
  var h='';
  L.forEach(function(x){
    for(var key in t)t[key]+=+x[key]||0;
    h+='<tr>'+
      '<td class="c" style="font-family:\'DM Mono\',monospace;font-size:11px;color:var(--mute)">'+e(x.id)+'</td>'+
      '<td class="c">'+e(x.partner)+'</td><td class="c">'+e(x.dt)+'</td>'+
      '<td class="nm" title="'+e(x.name)+'">'+e(x.name)+'</td>'+
      '<td class="c">'+e(x.teacher)+'</td>'+
      '<td class="c">'+e(x.due)+'</td>'+
      '<td class="n">'+n(x.rev)+'</td><td class="n">'+n(x.fee)+'</td><td class="n">'+n(x.recv)+'</td>'+
      '<td class="c">'+payB(x.paid_in,!OPEN_IN(x))+'</td>'+
      '<td class="n">'+n(x.payout)+'</td>'+
      '<td class="c">'+payB(x.paid_out,!OPEN_OUT(x))+'</td>'+
      '<td class="n" style="font-weight:600">'+n(x.ours)+'</td></tr>';});
  h+='<tr class="tot"><td colspan="6" class="c">합계 '+L.length+'건</td>'+
     '<td class="n">'+n(t.rev)+'</td><td class="n">'+n(t.fee)+'</td><td class="n">'+n(t.recv)+'</td><td></td>'+
     '<td class="n">'+n(t.payout)+'</td><td></td><td class="n">'+n(t.ours)+'</td></tr>';
  document.getElementById('tb').innerHTML=h;
  document.getElementById('cnt').textContent=L.length+'건 / 전체 '+D.filter(function(x){return x.held==='개최';}).length+'건';
}
function mkChips(id,arr,key,sel){
  var el=document.getElementById(id);
  el.innerHTML=el.innerHTML+arr.map(function(v){
    return '<button class="chip'+(v===(sel||arr[0])?' on':'')+'" data-v="'+e(v)+'">'+e(v)+'</button>';}).join('');
  el.addEventListener('click',function(ev){var b=ev.target.closest('.chip');if(!b)return;
    this.querySelectorAll('.chip').forEach(function(x){x.classList.remove('on');});
    b.classList.add('on');F[key]=b.dataset.v;render();});
}
mkChips('fP',['전체'].concat(Array.from(new Set(D.map(function(x){return x.partner;}))).sort()),'p');
mkChips('fS',['전체','정산 대기','정산완료'],'st');
document.getElementById('q').addEventListener('input',function(){F.q=this.value;render();});
document.querySelectorAll('th[data-k]').forEach(function(th){
  th.addEventListener('click',function(){
    var k=th.dataset.k;
    if(SORT.key===k)SORT.asc=!SORT.asc;else{SORT.key=k;SORT.asc=true;}
    render();});
});
document.getElementById('dlx').addEventListener('click',function(ev){ev.preventDefault();
  var C=[['회차ID','id'],['제휴사','partner'],['수업일','dt'],['클래스명','name'],['선생님','teacher'],
    ['지급 예정일','due'],['클래스매출','rev'],['제휴사 수수료','fee'],['제휴사 입금','recv'],
    ['입금상태','paid_in'],['선생님 실지급','payout'],['지급상태','paid_out'],['PK 수익','ours']];
  var q=function(v){v=String(v==null?'':v);return /[",\n]/.test(v)?'"'+v.replace(/"/g,'""')+'"':v;};
  var rows=[C.map(function(c){return c[0];}).join(',')];
  filt().forEach(function(x){rows.push(C.map(function(c){return q(x[c[1]]);}).join(','));});
  var b=new Blob(['\ufeff'+rows.join('\r\n')],{type:'text/csv;charset=utf-8'});
  var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='정산리포트.csv';a.click();});
// ══ 정산서 ══════════════════════════════
var DOC=null, DTAB='partner';
function doc(id){DOC=D.filter(function(x){return x.id===id;})[0];DTAB='partner';drawDoc();}
function dtab(t){DTAB=t;drawDoc();}
function closeDoc(){DOC=null;document.getElementById('modal').style.display='none';document.body.style.overflow='';}
function won(v){return n(v)+'원';}
function drawDoc(){
  if(!DOC)return;
  var x=DOC, el=document.getElementById('modal');
  var isOn=x.kind==='온라인';
  var fee=x.rev-x.recv, sup=Math.round(x.recv/1.1), vat=x.recv-sup;
  var h='<div class="mbg" onclick="closeDoc()"></div><div class="mbox">'+
    '<div class="mtop">'+
      '<div class="mtabs">'+
        '<button class="mtab'+(DTAB==='partner'?' on':'')+'" onclick="dtab(\'partner\')">제휴사 정산서</button>'+
        '<button class="mtab'+(DTAB==='teacher'?' on':'')+'" onclick="dtab(\'teacher\')">선생님 정산서</button>'+
      '</div>'+
      '<div class="macts">'+
        '<button class="mact" onclick="saveImg()">🖼 이미지 저장</button>'+
        '<button class="mact wide" onclick="copyDoc()">📋 복사</button>'+
        '<button class="mact hot" onclick="window.print()">🖨 인쇄 / PDF</button>'+
        '<button class="mact" onclick="closeDoc()">✕</button>'+
      '</div>'+
    '</div><div class="mbody" id="mbody">';
  if(DTAB==='partner') h+=docPartner(x,fee,sup,vat,isOn);
  else h+=docTeacher(x,isOn);
  el.innerHTML=h+'</div></div>';
  el.style.display='block'; document.body.style.overflow='hidden';
}
function docHead(t,x,sub){
  return '<div class="stamp">공공의주방</div>'+
    '<h1 class="dt">'+e(t)+'</h1>'+
    '<div class="dsub">&lt;'+e(x.name)+'&gt;</div>'+
    '<table class="dmeta"><tr><td>수업일</td><td>'+e(x.dt)+'</td><td>회차ID</td><td>'+e(x.id)+'</td></tr>'+
    '<tr><td>제휴사</td><td>'+e(x.partner)+'</td><td>선생님</td><td>'+e(x.teacher)+'</td></tr>'+
    '<tr><td>인원</td><td>'+e(x.pax)+'명</td><td>구분</td><td>'+e(x.kind)+' · '+e(x.state)+'</td></tr></table>'+
    (sub?'<p class="dnote">'+sub+'</p>':'');
}
function docPartner(x,fee,sup,vat,isOn){
  var h=docHead('정 산 서',x,'')+
    '<table class="dtbl"><tr><th>항목</th><th>단가</th><th>인원·수량</th><th>금액</th></tr>'+
    '<tr><td>'+e(x.name)+'</td><td class="n">'+n(x.price)+'</td><td class="c">'+e(x.pax)+'</td><td class="n">'+n(x.rev)+'</td></tr>'+
    '<tr class="sum"><td colspan="3">클래스매출 (VAT 포함)</td><td class="n">'+n(x.rev)+'</td></tr>';
  if(fee) h+='<tr><td colspan="3">− 제휴사 수수료 '+(x.rate||10)+'%</td><td class="n">−'+n(fee)+'</td></tr>';
  h+='<tr class="tot"><td colspan="3">정산 금액 (VAT 포함)</td><td class="n">'+n(x.recv)+'</td></tr>'+
    '</table>'+
    '<h3 class="dh3">세금계산서 발행 기준</h3>'+
    '<table class="dtbl two"><tr><th>구분</th><th>금액</th></tr>'+
    '<tr><td>공급가액</td><td class="n">'+n(sup)+'</td></tr>'+
    '<tr><td>부가세 (10%)</td><td class="n">'+n(vat)+'</td></tr>'+
    '<tr class="tot"><td>합계</td><td class="n">'+n(x.recv)+'</td></tr></table>'+
    '<table class="dpay"><tr><td>입금 예정일</td><td class="n b">'+e(x.due)+'</td></tr></table>'+
    docFoot();
  return h;
}
function docTeacher(x,isOn){
  var base=x.teachbase, pay=isOn?x.payout:Math.round(base*0.75);
  var biz=x.paykind==='사업자';
  var h=docHead('정 산 내 역 서',x,'선생님께 지급되는 금액 내역입니다. 상호 확인 후 지급됩니다.')+
    '<table class="dtbl three"><tr><th>구분</th><th>내용</th><th>금액</th></tr>'+
    '<tr><td>정산 기준</td><td>'+(isOn?'인당 정액 × '+e(x.pax)+'명':'클래스매출 '+n(x.rev)+' ÷ 1.1')+'</td><td class="n">'+n(base)+'</td></tr>'+
    (isOn?'':'<tr><td>강사료 비율</td><td>정산 기준의 75%</td><td class="n">'+n(pay)+'</td></tr>')+
    '<tr class="sum"><td colspan="2">정산 금액</td><td class="n">'+n(pay)+'</td></tr>';
  if(biz&&x.addvat) h+='<tr><td>부가세</td><td>세금계산서 발행분 10%</td><td class="n">+'+n(x.addvat)+'</td></tr>';
  if(x.wt){
    var it=Math.floor(pay*0.03/10)*10, lt=Math.floor(it*0.1/10)*10;
    if(it+lt!==x.wt){it=Math.round(x.wt/1.1);lt=x.wt-it;}
    h+='<tr><td>소득세</td><td>3%</td><td class="n">−'+n(it)+'</td></tr>'+
       '<tr><td>지방소득세</td><td>소득세의 10%</td><td class="n">−'+n(lt)+'</td></tr>';
  }
  h+='<tr class="tot"><td colspan="2">실지급액</td><td class="n">'+n(x.payout)+'</td></tr></table>'+
    '<table class="dpay"><tr><td>지급 방식</td><td class="b">'+e(x.paykind)+
      (biz?' · 세금계산서/현금영수증 수취':(x.wt?' · 원천세 3.3% 공제 후 지급':''))+'</td></tr>'+
    '<tr><td>지급 예정일</td><td class="n b">'+e(x.due)+'</td></tr></table>'+
    docFoot();
  return h;
}
function docFoot(){
  return '<div class="dfoot"><div class="dfl">'+
    '<p>※ 위 금액은 회차 정산 기준이며, 지급 예정일에 입금됩니다.</p>'+
    '<p>※ 금액에 이견이 있으시면 지급 전에 알려주시기 바랍니다.</p></div>'+
    '<table class="dco"><tr><td colspan="2" class="con">__CO_NAME__</td></tr>'+
    '<tr><td>대표자</td><td>__CO_CEO__</td></tr>'+
    '<tr><td>연락처</td><td>__CO_TEL__</td></tr>'+
    '<tr><td>등록번호</td><td>__CO_BIZ__</td></tr></table></div>';
}
function saveImg(){
  var el=document.querySelector('.mbody');
  var go=function(){
    html2canvas(el,{scale:2,backgroundColor:'#ffffff',useCORS:true}).then(function(cv){
      cv.toBlob(function(b){
        var a=document.createElement('a');
        a.href=URL.createObjectURL(b);
        a.download='정산서_'+(DOC?DOC.id:'')+'_'+(DTAB==='partner'?'제휴사':'선생님')+'.png';
        a.click();});});
  };
  if(window.html2canvas){go();return;}
  var sc=document.createElement('script');
  sc.src='https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js';
  sc.onload=go;
  sc.onerror=function(){alert('이미지 저장을 불러오지 못했습니다. 인쇄/PDF를 이용해 주세요.');};
  document.head.appendChild(sc);
}
function copyDoc(){
  var t=document.getElementById('mbody').innerText.replace(/\n{3,}/g,'\n\n');
  navigator.clipboard.writeText(t).then(function(){alert('복사했습니다');});
}
document.addEventListener('keydown',function(ev){if(ev.key==='Escape')closeDoc();});
now();render();
"""

doc = f"""<!DOCTYPE html><html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex, nofollow">
<meta name="version" content="{VERSION}"><meta name="updated" content="{UPDATED}">
<title>정산 리포트</title>
<link rel="apple-touch-icon" href="app-icon.png"><link rel="icon" href="app-icon.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Hahmlet:wght@400;700;800&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css" rel="stylesheet">
<style>{CSS}</style></head><body><div class="wrap">

<div class="topbar">
<div><a class="pill" href="class.html">← 클래스 홈</a><a class="pill" href="class-list.html">📋 클래스 목록</a><a class="pill" href="{CHAT}" target="_blank" rel="noopener">💬 작업 대화</a></div>
<div class="crumb">AI 워크스페이스 › 클래스 › 정산 리포트<br><span class="meta">업데이트 {UPDATED} · {VERSION}</span></div>
</div>

<div class="head">
<h1>정산 리포트</h1>
<p class="lead"><b>아직 받지 못한 돈과 나가야 할 돈</b>을 봅니다. 지급일은 자체 클래스가 수업일 1~15일 → 당월 25일 · 16~말일 → 익월 10일, <b>제휴사는 수업 익월 25일</b>입니다.</p>
</div>

<div class="now">
<div class="now-h"><b>지금 챙길 것</b><span>정산이 끝나지 않은 회차만</span></div>
<div class="ncards" id="ncards"></div>
</div>

<div class="sec">
<h2>진행 중 · 정산 대기</h2>
<p class="h2d">지급 예정일 순 · 날짜가 지난 건은 배경으로 표시됩니다<br>
<b>「수업일 지남」</b> 배지가 붙은 건은 수업이 끝났는데 상태가 <code>확정</code>·<code>모집중</code> 그대로입니다 — 마스터에서 <code>수업완료</code>로 바꿔주세요</p>
<div id="livewrap"></div>
</div>

<div class="sec">
<h2>확인 대기 이체</h2>
<p class="h2d">통장에 이체 기록은 있으나 대응되는 회차를 찾지 못한 건</p>
<div id="unwrap"></div>
</div>

<div class="sec">
<h2>전체 정산 내역</h2>
<p class="h2d">개최된 모든 회차의 입금·지급 상태</p>
<div class="bar" id="fP"><b>제휴사</b></div>
<div class="bar" id="fS"><b>상태</b></div>
<div class="bar" style="margin-top:4px"><input class="srch" id="q" placeholder="클래스명 · 날짜 · 선생님 · 회차ID 검색"></div>
<div class="cnt" id="cnt"></div>
<div class="tw"><table>
<thead>
<tr class="grp"><th colspan="6">클래스</th><th colspan="4">받을 돈</th><th colspan="2">줄 돈</th><th>수익</th></tr>
<tr>
<th data-k="id">회차ID</th><th data-k="partner">제휴사</th><th data-k="dt">수업일</th>
<th data-k="name">클래스명</th><th data-k="teacher">선생님</th><th data-k="due">지급 예정일</th>
<th data-k="rev">클래스매출</th><th data-k="fee">제휴사 수수료</th><th data-k="recv">제휴사 입금</th><th data-k="paid_in">입금</th>
<th data-k="payout">선생님 실지급</th><th data-k="paid_out">지급</th>
<th data-k="ours">PK 수익</th>
</tr></thead><tbody id="tb"></tbody></table></div>
<p class="note">※ 제휴사 마스터는 <b>입금상태·지급상태 칸</b>을 그대로 읽습니다. 대응 회차를 못 찾은 이체는 위 「확인 대기 이체」에 모았습니다.<br>
※ <b>받을 돈</b>은 제휴사가 우리에게 줄 돈(VAT 포함), <b>줄 돈</b>은 선생님께 실제로 보낼 금액입니다.<br>
※ 회차별 수익 구조는 <a href="class-list.html" style="color:var(--coral)">클래스 목록</a>에서 봅니다.</p>
</div>

<div class="files">
<div class="files-l">📦 원본 파일</div>
<a class="fbtn" href="#" id="dlh">📄 이 화면 HTML 받기</a>
<a class="fbtn" href="#" id="dlx">📊 정산 내역 CSV 받기</a>
<a class="fbtn" href="{CHAT}" target="_blank" rel="noopener">💬 대화창에서 전체 받기</a>
</div>
<div class="foot">정산 리포트 · {UPDATED} · {VERSION}</div>
</div>
<div id="modal"></div>
<script>
document.getElementById('dlh').addEventListener('click',function(ev){{ev.preventDefault();
var b=new Blob(['<!DOCTYPE html>'+document.documentElement.outerHTML],{{type:'text/html'}});
var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='class-settle-report.html';a.click();}});
{JS.replace('__DATA__',DATA).replace('__CO_NAME__',CO['name']).replace('__CO_CEO__',CO['ceo']).replace('__CO_TEL__',CO['tel']).replace('__CO_BIZ__',CO['biz'])}
</script></body></html>"""

open("class-settle-report.html","w",encoding="utf-8").write(doc)
print(f"class-settle-report.html {VERSION} — 회차 {len(ROWS)}건 · 확인 대기 이체 {len(UNMATCHED)}건")
live=[x for x in ROWS if x['held']=='개최' and (x['paid_in']!='입금완료' or x['paid_out']!='지급완료')]
print(f"  정산 대기 {len(live)}건")
print(f"  크기 {os.path.getsize('class-settle-report.html')/1024:.0f}KB")
