# -*- coding: utf-8 -*-
"""
클래스 메뉴 지식 사전 생성기 (박제형 · 유형 B)
입력: class-menu-master.xlsx (시트 '메뉴사전' · '분류요약')
출력: class-menu.html
사용법: python gen_menu.py class-menu-master.xlsx
"""
import sys, html, json, datetime, os, base64, mimetypes
from openpyxl import load_workbook

KST=datetime.timezone(datetime.timedelta(hours=9))
UPDATED=datetime.datetime.now(KST).strftime("%Y.%m.%d %H:%M")
VERSION="v41"
CHAT_URL="https://claude.ai/chat/909e7f28-5718-4bde-8997-e37348632306"
ICON={"튀김":"🔥","도구":"🔧","조림":"♨️","굳힘":"⏱️","건조":"☀️","조형":"🌸","없음":"○"}

IMGDIR="menu-images"
_miss, _bytes = [], [0]
def embed(fn):
    """menu-images/의 사진을 data URI로 읽어 HTML에 박제한다. 없으면 건너뜀."""
    p=os.path.join(IMGDIR,fn)
    if not os.path.exists(p):
        _miss.append(fn); return ""
    mt=mimetypes.guess_type(p)[0] or "image/webp"
    raw=open(p,"rb").read(); _bytes[0]+=len(raw)
    return f"data:{mt};base64," + base64.b64encode(raw).decode()

src=sys.argv[1] if len(sys.argv)>1 else "class-menu-master.xlsx"
wb=load_workbook(src,data_only=True)
ws=wb["메뉴사전"]; hdr=[c.value for c in ws[1]]
rows=[]
for r in ws.iter_rows(min_row=2,values_only=True):
    if not r[0]: continue
    d={hdr[i]:("" if r[i] is None else str(r[i]).strip()) for i in range(len(hdr))}
    d["테마목록"]=[t for t in d["테마"].replace("·","|").split("|") if t]
    d["제약목록"]=[t for t in d["제약"].replace("·","|").split("|") if t]
    d["아이콘"]="".join(ICON.get(t,"") for t in d["제약목록"])
    d["이미지목록"]=[u for u in (embed(x.strip()) for x in d["이미지"].split(",") if x.strip()) if u]
    rows.append(d)
grp=[[("" if c is None else str(c).strip()) for c in r]
     for r in wb["분류요약"].iter_rows(min_row=2,values_only=True) if r[0]]

cats=sorted({r["소분류"] for r in rows}, key=lambda x:[g[0] for g in grp].index(x) if x in [g[0] for g in grp] else 99)
themes=["추석","설날","봄","여름","가을","겨울","상시"]
themes=[t for t in themes if any(t in r["테마목록"] for r in rows)]
cons=["튀김","도구","조림","굳힘","건조","조형","없음"]
cons=[c for c in cons if any(c in r["제약목록"] for r in rows)]
DATA=json.dumps(rows,ensure_ascii=False)

CSS="""
:root{--bg:#FAF5EC;--ink:#241F1B;--sub:#6B6256;--line:#E7DCC8;--card:#FFFDF9;
--coral:#FF5019;--blue:#2C7BB6;--mute:#ADA294}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--ink);
font-family:Pretendard,-apple-system,"Apple SD Gothic Neo","Malgun Gothic",sans-serif;
font-size:15px;line-height:1.62;-webkit-text-size-adjust:100%}
.wrap{max-width:1080px;margin:0 auto;padding:20px 18px 70px}
.topbar{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;flex-wrap:wrap}
.pill{display:inline-block;font-size:13px;padding:5px 13px;border-radius:20px;margin-right:8px;
background:#F1EFE9;border:1px solid var(--line);color:var(--ink);text-decoration:none}
.tb-right{font-family:"DM Mono",ui-monospace,monospace;font-size:12px;color:var(--mute);text-align:right}
h1{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-weight:700;
font-size:40px;letter-spacing:-.02em;margin:20px 0 6px}
.lead{font-size:15px;color:var(--sub);margin:0 0 24px}
h2{font-family:Hahmlet,"Gowun Batang",serif;font-size:19px;margin:30px 0 10px}
.gtab{width:100%;border-collapse:collapse;font-size:13.5px;background:var(--card);
border:1px solid var(--line);border-radius:12px;overflow:hidden}
.gtab th{background:#F1EFE9;font-size:12px;color:var(--sub);font-weight:500;
padding:8px 11px;text-align:left;border-bottom:1px solid var(--line)}
.gtab td{padding:10px 11px;border-bottom:1px solid #F0EADC;vertical-align:top}
.gtab tr:last-child td{border-bottom:none}
.gtab .g1{font-weight:600;white-space:nowrap}
.gtab .g2{white-space:nowrap;font-size:12.5px}
.gtab .g4{color:var(--sub);font-size:12.5px}
.bar{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:7px;align-items:center}
.bar b{font-size:12px;color:var(--mute);margin-right:2px;font-weight:500;min-width:44px}
.chip{font-size:13px;padding:4px 12px;border-radius:20px;border:1px solid var(--line);
background:#fff;color:var(--sub);cursor:pointer;font-family:inherit}
.chip.on{background:var(--blue);border-color:var(--blue);color:#fff}
.cnt{font-family:"DM Mono",monospace;font-size:12px;color:var(--mute);margin:14px 0 10px}
.card{background:var(--card);border:1px solid var(--line);border-radius:13px;
padding:15px 17px;margin-bottom:11px}
.card.dim{opacity:.62}
.imgs{display:flex;gap:8px;flex-wrap:wrap;margin:2px 0 11px}
.thumb{width:112px;height:112px;border-radius:11px;object-fit:cover;
display:block;background:#F1EFE9;border:1px solid var(--line)}
.thumb.ph{display:flex;align-items:center;justify-content:center;font-size:30px;color:#D8D2C4}
.ct{display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;margin-bottom:6px}
.nm{font-family:Hahmlet,"Gowun Batang",serif;font-size:20px;font-weight:700}
.bd{font-size:12px;padding:2px 9px;border-radius:20px;white-space:nowrap}
.b-info{background:#F1EFE9;color:#6B6256}
.b-con{background:#F1EFE9;color:#4A4038;font-weight:600}
.b-best{background:#FFE7DF;color:var(--coral);font-weight:500}
.b-off{background:#EFEDEA;color:#ADA294}
.need{display:flex;gap:8px;align-items:flex-start;margin:9px 0 11px;padding:9px 12px;
border-radius:9px;background:#F4F1E8;font-size:13.5px}
.need i{font-style:normal;color:var(--mute);font-size:12px;white-space:nowrap;padding-top:1px}
.need b{font-weight:600}
.desc{font-size:14px;color:var(--sub);margin:0 0 10px}
.gr{display:grid;grid-template-columns:74px 1fr;gap:4px 12px;font-size:13.5px}
.gr dt{color:var(--mute);font-size:12.5px}
.gr dd{margin:0}
.why{margin-top:10px;font-size:13px;padding:8px 11px;border-radius:9px;background:#F7F3E9;color:#6B6256}
.why.no{background:#FBF3EF;color:#B44A1E}
.num{font-family:"DM Mono",monospace}
.note{margin-top:34px;padding-top:16px;border-top:1px solid var(--line);
font-size:13px;color:var(--mute);line-height:1.8}
.dl{margin-top:20px;display:flex;gap:8px;flex-wrap:wrap}
.dl a{font-size:13px;padding:7px 14px;border-radius:9px;background:#F1EFE9;
border:1px solid var(--line);color:var(--ink);text-decoration:none}
.empty{padding:40px 0;text-align:center;color:var(--mute);font-size:14px}
@media(max-width:640px){h1{font-size:30px}.wrap{padding:16px 14px 60px}
.thumb{width:84px;height:84px}
.gr{grid-template-columns:64px 1fr}.gtab .g3{display:none}.gtab{font-size:12.5px}}
"""

JS="""
var D=__DATA__, fS='전체', fT='전체', fO='전체', fC='전체';
function esc(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){
 return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});}
function okBadge(v){
 if(v==='◎')return '<span class="bd b-best">온라인 최적</span>';
 if(v==='△')return '<span class="bd b-info">온라인 조건부</span>';
 return '<span class="bd b-off">온라인 불가</span>';}
function card(r){
 var h='<div class="card'+(r['온라인적합']==='✕'?' dim':'')+'">';
 h+='<div class="ct"><span class="nm">'+esc(r['메뉴명'])+'</span>';
 h+='<span class="bd b-info">'+esc(r['소분류'])+'</span>';
 h+='<span class="bd b-con">'+r['아이콘']+' '+esc(r['제약'])+'</span>';
 h+=okBadge(r['온라인적합'])+'</div>';
 h+='<div class="imgs">'+(r['이미지목록'].length
   ? r['이미지목록'].map(function(f){return '<img class="thumb" src="'+f+'" alt="'+esc(r['메뉴명'])+'" loading="lazy">';}).join('')
   : '<div class="thumb ph">🍡</div>')+'</div>';
 h+='<p class="desc">'+esc(r['설명'])+'</p>';
 h+='<div class="need"><i>필요 도구</i><span><b>'+esc(r['필요도구·설비'])+'</b></span></div>';
 h+='<dl class="gr">';
 h+='<dt>주요리법</dt><dd>'+esc(r['주요리법'])+'</dd>';
 h+='<dt>주재료</dt><dd>'+esc(r['주재료'])+'</dd>';
 h+='<dt>소요시간</dt><dd class="num">'+esc(r['소요시간'])+' · 난이도 '+esc(r['난이도'])+'</dd>';
 h+='<dt>굳힘·보존</dt><dd>'+esc(r['굳힘/보존'])+'</dd>';
 h+='<dt>테마</dt><dd>'+esc(r['테마'])+'</dd>';
 if(r['선생님후보'])h+='<dt>선생님</dt><dd>'+esc(r['선생님후보'])+'</dd>';
 h+='</dl>';
 var w=r['온라인 불가/조건 사유']||r['비고'];
 if(w)h+='<div class="why'+(r['온라인적합']==='✕'?' no':'')+'">'+esc(w)+'</div>';
 return h+'</div>';}
function render(){
 var L=D.filter(function(r){
  if(fS!=='전체'&&r['소분류']!==fS)return false;
  if(fT!=='전체'&&r['테마목록'].indexOf(fT)<0)return false;
  if(fC!=='전체'&&r['제약목록'].indexOf(fC)<0)return false;
  if(fO!=='전체'&&r['온라인적합']!==fO)return false;
  return true;});
 var rk={'◎':0,'△':1,'✕':2};
 L.sort(function(a,b){return (rk[a['온라인적합']]-rk[b['온라인적합']])
  ||a['소분류'].localeCompare(b['소분류'],'ko');});
 document.getElementById('cnt').textContent=L.length+'건 / 전체 '+D.length+'건';
 document.getElementById('list').innerHTML=
  L.length?L.map(card).join(''):'<div class="empty">조건에 맞는 메뉴가 없습니다.</div>';}
function bind(id,set){var el=document.getElementById(id); if(!el)return;
 el.addEventListener('click',function(e){var b=e.target.closest('.chip'); if(!b)return;
  this.querySelectorAll('.chip').forEach(function(x){x.classList.remove('on');});
  b.classList.add('on'); set(b.dataset.v); render();});}
bind('fS',function(v){fS=v;}); bind('fT',function(v){fT=v;});
bind('fC',function(v){fC=v;}); bind('fO',function(v){fO=v;});
render();
"""

def chips(items,gid,lab,pre=None):
    h=f'<div class="bar" id="{gid}"><b>{lab}</b><button class="chip on" data-v="전체">전체</button>'
    for i in items:
        t=(ICON.get(i,"")+" "+i) if pre else i
        h+=f'<button class="chip" data-v="{html.escape(i)}">{html.escape(t)}</button>'
    return h+"</div>"

gtab='<table class="gtab"><tr><th>분류</th><th>제약</th><th class="g3">무엇이 걸리나</th><th>온라인</th></tr>'
for g in grp:
    gtab+=(f'<tr><td class="g1">{html.escape(g[0])}</td>'
           f'<td class="g2">{ICON.get(g[1].split("·")[0],"")} {html.escape(g[1])}</td>'
           f'<td class="g3">{html.escape(g[2])}</td>'
           f'<td class="g4">{html.escape(g[4])}</td></tr>')
gtab+="</table>"

doc=f"""<!DOCTYPE html><html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="version" content="{VERSION}"><meta name="updated" content="{UPDATED}">
<title>클래스 메뉴 지식 사전</title>
<link rel="apple-touch-icon" href="app-icon.png"><link rel="icon" href="app-icon.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Hahmlet:wght@400;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css" rel="stylesheet">
<style>{CSS}</style></head><body><div class="wrap">

<div class="topbar">
<div><a class="pill" href="class.html">← 클래스 홈</a><a class="pill" href="{CHAT_URL}" target="_blank" rel="noopener">💬 작업 대화</a></div>
<div class="tb-right">AI 워크스페이스 › 클래스 › 메뉴 사전<br>업데이트 {UPDATED} · {VERSION}</div>
</div>

<h1>메뉴 지식 사전 · 한식디저트</h1>
<p class="lead">수업으로 굴릴 수 있는 메뉴인지 판단하기 위한 자료입니다. 한과·양갱·떡을 함께 담습니다. 레시피는 담지 않습니다.</p>

<h2>분류별 제약 — 먼저 이걸 보세요</h2>
{gtab}

<h2>메뉴</h2>
{chips(cats,'fS','분류')}
{chips(cons,'fC','제약',pre=True)}
{chips(themes,'fT','테마')}
<div class="bar" id="fO"><b>온라인</b>
<button class="chip on" data-v="전체">전체</button>
<button class="chip" data-v="◎">최적</button>
<button class="chip" data-v="△">조건부</button>
<button class="chip" data-v="✕">불가</button></div>

<div class="cnt" id="cnt"></div>
<div id="list"></div>

<div class="note">
※ 이 화면은 보기 전용입니다. 고칠 땐 마스터(class-menu-master.xlsx) 또는 작업 대화창에서.<br>
※ 제약 · 필요도구 · 소요시간 · 온라인적합은 수업 운영 관점의 판단이며, 실제 진행 후 보정이 필요합니다.<br>
※ 레시피 원본은 쿠킹박스 레시피마스터 소관입니다. 메뉴명으로 연결됩니다.<br>
※ 사진은 이 HTML 안에 들어 있습니다. 서버에 이미지 파일을 따로 올릴 필요가 없습니다.<br>\n※ 사진 추가·교체는 원본을 prep_image.py로 규격화한 뒤 마스터 '이미지' 칸에 파일명을 적고 gen을 다시 돌립니다.
</div>

<div class="dl">
<a href="#" id="dlh">📄 이 화면 HTML 받기</a>
<a href="class-menu-master.xlsx" download>📊 엑셀 마스터 받기</a>
<a href="{CHAT_URL}" target="_blank" rel="noopener">💬 대화창에서 전체 받기</a>
</div>

</div><script>
document.getElementById('dlh').addEventListener('click',function(e){{e.preventDefault();
var b=new Blob(['<!DOCTYPE html>'+document.documentElement.outerHTML],{{type:'text/html'}});
var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='class-menu.html';a.click();}});
{JS.replace('__DATA__',DATA)}
</script></body></html>"""

open("class-menu.html","w",encoding="utf-8").write(doc)
import os as _os
_kb=_os.path.getsize("class-menu.html")/1024
print(f"class-menu.html {VERSION} — 메뉴 {len(rows)}건 · 분류 {len(cats)} · 제약 {len(cons)} · 테마 {len(themes)}")
print(f"  사진 내장 {_bytes[0]/1024:.0f}KB → HTML {_kb:.0f}KB")
if _miss: print("  ⚠️ 파일 없음:", ", ".join(_miss))
