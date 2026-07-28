# -*- coding: utf-8 -*-
"""
클래스 목록 통합 생성기 — 여러 마스터를 읽어 하나의 목록으로
입력: class-master-all-v3.xlsx (필수) + 갤러리아_강좌이력.xlsx 등 제휴사 마스터
출력: class-list.html
사용법: python gen_list.py

■ 공통 12항목 — 마스터가 늘어도 이 형태로만 맞추면 붙는다
  회차ID · 제휴사 · 수업일 · 수업명 · 구분 · 유형 · 선생님 · 개최 · 인원
  클래스매출(부가세 포함) · 제휴 수수료 · 우리 몫(부가세 제외)
■ 부가 항목
  우리 수령 · 선생님 지급 · 우리 부담 원가 · 수업상태 · 입금상태 · 지급상태 · 출처
"""
import os, json, datetime, re
from openpyxl import load_workbook

KST = datetime.timezone(datetime.timedelta(hours=9))
UPDATED = datetime.datetime.now(KST).strftime("%Y.%m.%d %H:%M")
VERSION = "v60"
CHAT_URL = "https://claude.ai/chat/909e7f28-5718-4bde-8997-e37348632306"

def D(v):
    if isinstance(v,(datetime.datetime,)): return v.date()
    if isinstance(v,datetime.date): return v
    if isinstance(v,str) and len(v)>=10:
        try: return datetime.date(int(v[:4]),int(v[5:7]),int(v[8:10]))
        except: return None
    return None

ROWS=[]

# ══ ① 클래스 마스터 ════════════════════════════
CM='class-master-all-v3.xlsx'
if os.path.exists(CM):
    wb=load_workbook(CM,data_only=True)
    ws=wb['회차']; h=[c.value for c in ws[1]]
    # 수강생 시트로 신청인원 집계
    cnt={}
    if '수강생' in wb.sheetnames:
        w2=wb['수강생']
        for r in w2.iter_rows(min_row=2,values_only=True):
            if r and r[0]: cnt[str(r[0]).strip()]=cnt.get(str(r[0]).strip(),0)+1
    for r in ws.iter_rows(min_row=2,values_only=True):
        if not r[0]: continue
        d=dict(zip(h,r)); cid=str(d['회차ID']).strip()
        dt=D(d.get('수업일'))
        pax=cnt.get(cid,0) or (d.get('신청인원') or 0) or (d.get('최소인원') or 0)
        price=d.get('판매가') or 0
        rev=price*pax
        sup=rev/1.1 if d.get('부가세구분')=='포함' else rev
        cost=((d.get('재료비(1인)') or 0)+(d.get('패킹비') or 0)+(d.get('배송비') or 0))*pax
        teach=(d.get('강사비') or 0)*pax
        fee=0                       # 제휴 수수료 — 현재 회차는 없음
        state=str(d.get('수업상태') or '')
        ROWS.append(dict(
            id=cid, partner=str(d.get('발주처') or '—'), dt=dt, name=str(d.get('수업명') or ''),
            kind=str(d.get('구분') or ''), ty='단일', teacher=str(d.get('선생님') or ''),
            held='폐강' if state in ('폐강','취소','무산') else '개최',
            pax=pax, rev=round(rev), fee=fee, recv=round(rev-fee),
            pay=round(teach), payout=round(teach), cost=round(cost),
            recvx=round(rev/1.1), base=round(sup), rate=0,
            ours=round(sup-fee/1.1-cost-teach),
            state=state, paid_in=str(d.get('입금상태') or ''), paid_out=str(d.get('지급상태') or ''),
            paykind='미확인', st='', src='클래스 마스터'))
    print(f"  클래스 마스터 {sum(1 for x in ROWS)}건")

# ══ ② 제휴사 마스터 ════════════════════════════
PARTNERS=[('갤러리아_강좌이력.xlsx','갤러리아광교','오프라인')]
for f,pname,kind in PARTNERS:
    if not os.path.exists(f): print(f"  {f} 없음 — 건너뜀"); continue
    wb=load_workbook(f,data_only=True); ws=wb['회차이력']
    H=[str(ws.cell(row=5,column=c).value).replace('\n','') for c in range(1,22)]
    n=0
    for r in range(6,ws.max_row+1):
        a=ws.cell(row=r,column=1).value
        if not a or str(a)=='합계': continue
        d={H[c-1]:ws.cell(row=r,column=c).value for c in range(1,22)}
        ROWS.append(dict(
            id=str(d['회차ID']), partner=pname, dt=D(d['날짜']), name=str(d['클래스명'] or ''),
            kind=kind, ty=str(d['유형'] or ''), teacher=str(d['선생님'] or ''),
            held=str(d['개최'] or ''), pax=d['인원'] or 0,
            rev=round(d['클래스매출'] or 0), fee=round(d['갤 수수료'] or 0),
            recv=round(d['갤러리아 정산'] or 0), recvx=round((d['갤러리아 정산'] or 0)/1.1),
            base=round((d['클래스매출'] or 0)/1.1),
            rate=(round((d['갤 수수료'] or 0)/(d['클래스매출'] or 1)*1000)/10 if d['클래스매출'] else 0),
            pay=round((d['갤러리아 정산'] or 0)/1.1-(d['PK 수익'] or 0)),
            payout=round(d['선생님 지급'] or 0), cost=0,
            ours=round(d['PK 수익'] or 0),
            state='수업완료' if d['개최']=='개최' else '폐강', paid_in='', paid_out='',
            paykind=str(d.get('지급방식') or '—'), st=str(d.get('수수료 구조') or ''), src=pname)); n+=1
    print(f"  {pname} {n}건")

# 명세 항목 (펼침용)
ITEMS={}
for f,pname,kind in PARTNERS:
    if not os.path.exists(f): continue
    wb=load_workbook(f,data_only=True)
    if '명세원본-2025' in wb.sheetnames:
        ws=wb['명세원본-2025']
        for r in range(5,ws.max_row+1):
            gid=ws.cell(row=r,column=17).value
            if not gid: continue
            rev=ws.cell(row=r,column=9).value or 0
            gal=ws.cell(row=r,column=11).value or 0
            pre=ws.cell(row=r,column=13).value or 0
            ITEMS.setdefault(gid,[]).append(dict(
                it=ws.cell(row=r,column=18).value or '단일',
                sup=round(ws.cell(row=r,column=6).value or 0),
                price=round(ws.cell(row=r,column=7).value or 0),
                pax=ws.cell(row=r,column=8).value or 0,
                rev=round(rev), rate=ws.cell(row=r,column=10).value or 0,
                gal=round(gal), galx=round(gal/1.1),
                base=round(rev/1.1), pay=round(pre),
                fin=round(ws.cell(row=r,column=16).value or 0),
                wh=round(ws.cell(row=r,column=15).value or 0),
                pk=round(gal/1.1-pre)))
    if '명세원본-2024' in wb.sheetnames:
        ws=wb['명세원본-2024']
        for r in range(5,ws.max_row+1):
            gid=ws.cell(row=r,column=1).value
            if not gid: continue
            rev=ws.cell(row=r,column=21).value or 0
            gal=ws.cell(row=r,column=23).value or 0
            pre=ws.cell(row=r,column=28).value or 0
            ITEMS.setdefault(gid,[]).append(dict(
                it='단일', sup=round(ws.cell(row=r,column=8).value or 0),
                price=round(ws.cell(row=r,column=9).value or 0),
                pax=ws.cell(row=r,column=15).value or 0,
                rev=round(rev), rate=ws.cell(row=r,column=22).value or 0,
                gal=round(gal), galx=round(gal/1.1),
                base=round(rev/1.1), pay=round(pre),
                fin=round(ws.cell(row=r,column=34).value or 0),
                wh=round(ws.cell(row=r,column=32).value or 0),
                pk=round(gal/1.1-pre)))

ROWS.sort(key=lambda x:(x['dt'] or datetime.date(1900,1,1), x['partner'], x['name']))
for x in ROWS: x['dt']=str(x['dt']) if x['dt'] else ''
DATA=json.dumps({'rows':ROWS,'items':ITEMS},ensure_ascii=False)

CSS = """
:root{--bg:#FAF5EC;--ink:#241F1B;--sub:#6B6256;--line:#E7DCC8;--card:#FFFDF9;
--coral:#FF5019;--point:#B4A032;--wait:#9A8F7C;--mute:#ADA294;--ok:#5E7360}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--ink);font-family:Pretendard,-apple-system,"Apple SD Gothic Neo",sans-serif;
line-height:1.55;-webkit-font-smoothing:antialiased;padding-bottom:70px;font-size:15px}
.wrap{max-width:1520px;margin:0 auto;padding:0 20px}
.topbar{display:flex;justify-content:space-between;align-items:flex-start;padding:18px 0;gap:12px;flex-wrap:wrap}
.pill{display:inline-flex;align-items:center;gap:4px;font-size:13px;text-decoration:none;color:var(--ink);
background:var(--card);border:1px solid var(--line);border-radius:20px;padding:5px 13px;margin-right:8px}
.pill:hover{border-color:var(--point);color:var(--point)}
.crumb{font-size:12px;color:var(--mute);text-align:right;line-height:1.7}
.crumb .meta{font-family:"DM Mono",monospace;font-size:11px}
.head{padding:14px 0 22px;border-bottom:1px solid var(--line);margin-bottom:20px}
h1{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-weight:800;font-size:40px;letter-spacing:-.02em}
.lead{font-size:15px;color:var(--sub);margin-top:12px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:11px;margin-bottom:20px}
.c{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:14px 16px}
.c .k{font-size:12.5px;color:var(--sub)}
.c .v{font-family:"DM Mono",monospace;font-size:22px;font-weight:700;margin-top:4px;letter-spacing:-.02em}
.c .u{font-size:11.5px;color:var(--wait);margin-top:2px}
.bar{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:8px}
.bar b{font-size:12px;color:var(--mute);font-weight:500;min-width:44px}
.chip{font-size:13px;padding:4px 12px;border-radius:20px;border:1px solid var(--line);
background:#fff;color:var(--sub);cursor:pointer;font-family:inherit}
.chip.on{background:var(--coral);border-color:var(--coral);color:#fff}
.srch{font-family:inherit;font-size:13.5px;padding:7px 12px;border:1px solid var(--line);
border-radius:9px;background:#fff;color:var(--ink);width:280px}
.cnt{font-family:"DM Mono",monospace;font-size:12px;color:var(--mute);margin:10px 0}
.tw{overflow-x:auto;background:var(--card);border:1px solid var(--line);border-radius:12px}
tr.grp th{background:#EDE9DE;font-size:11.5px;color:#4A4038;border-bottom:1px solid var(--line);
border-right:1.5px solid #D6CFBC}
tr.grp th:last-child{border-right:none}
thead tr:nth-child(2) th{border-right:1px solid var(--line)}
thead tr:nth-child(2) th:last-child{border-right:none}
tbody td{border-right:1px solid #F4EFE3}
tbody td:last-child{border-right:none}
th.g,td.g{border-right:1.5px solid #D6CFBC !important}
tr.sub td{background:#F7F3E9;padding:0 0 0 26px;border-top:none}
.drill{border-left:3px solid var(--coral);background:#fff;border-radius:0 10px 10px 0;
padding:11px 14px 12px;margin:0 14px 10px 0}
.drill-h{font-size:12.5px;font-weight:600;color:var(--coral);margin-bottom:8px}
.drill table{border:1px solid var(--line);border-radius:8px;overflow:hidden;min-width:0}
.drill th{position:static;white-space:nowrap;font-size:10.5px;padding:6px 9px;
border-right:1px solid var(--line)}
.drill th:last-child{border-right:none}
.drill td{border-right:1px solid #F4EFE3}
.drill td:last-child{border-right:none}
.drill th.g,.drill td.g{border-right:1.5px solid #D6CFBC}
.drill tr.tot td{background:#F1EFE9;font-weight:700;border-top:1.5px solid var(--line)}
tr.clk{cursor:pointer}
tr.clk:hover td{background:#F7F3E9}
tr.clk.opened td{background:#FFF1EB}
tr.clk.opened .cx{color:var(--coral)}
.cx{color:var(--mute);font-size:10px}
table{border-collapse:collapse;width:100%;font-size:12.5px;min-width:1180px}
th .sub2{display:block;font-size:9.5px;font-weight:400;color:var(--mute);margin-top:1px;letter-spacing:-.02em}
th{background:#F1EFE9;font-size:11px;font-weight:700;color:var(--sub);padding:7px 6px;
text-align:center;white-space:nowrap;line-height:1.3;position:sticky;top:0;z-index:2;cursor:pointer;user-select:none}
th:hover{color:var(--coral)}
td{padding:6px 7px;border-top:1px solid #F0EADC;vertical-align:middle;line-height:1.4}
td.n{font-family:"DM Mono",monospace;text-align:right;white-space:nowrap;font-size:12px}
td.c{text-align:center;white-space:nowrap}
td.nm{max-width:168px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
td.pt{max-width:74px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:11.5px}
.bd{font-size:10.5px;padding:2px 6px}
th{vertical-align:middle}
table{table-layout:auto}
tr.off td{color:var(--mute)}
tr:hover td{background:#F7F3E9}
tr.tot td{background:#F1EFE9;font-weight:700;border-top:1.5px solid var(--line)}
th.pk,td.pk{position:sticky;right:0;z-index:3;background:var(--card);
box-shadow:-6px 0 8px -6px rgba(0,0,0,.10)}
th.pk{background:#F1EFE9;z-index:4}
tr.tot td.pk{background:#F1EFE9}
tr:hover td.pk{background:#F7F3E9}
tr.off td.pk{background:var(--card)}
tr.clk.opened td.pk{background:#FFF1EB}
.drill td.pk,.drill th.pk{position:static;box-shadow:none}
.bd{display:inline-block;font-size:11px;padding:2px 8px;border-radius:20px;white-space:nowrap}
.b-mut{background:#F1EFE9;color:#6B6256}
.b-cor{background:#FFE7DF;color:var(--coral);font-weight:500}
.b-lit{background:#FFF1EB;color:#C2570E}
.b-yel{background:#F5B21F;color:#4A3405;font-weight:500}
.b-grn{background:#E4EAE2;color:var(--ok)}
.b-off{background:#EFEDEA;color:var(--mute)}
.note{font-size:12.5px;color:var(--wait);margin-top:11px;line-height:1.75}
.files{padding:24px 0 10px;border-top:1px solid var(--line);margin-top:28px}
.files-l{font-family:"DM Mono",monospace;font-size:11px;letter-spacing:.1em;color:var(--mute);margin-bottom:10px}
.fbtn{display:inline-flex;align-items:center;gap:7px;font-size:13px;padding:10px 16px;border-radius:10px;
border:1px solid var(--line);background:#fff;color:var(--sub);text-decoration:none;margin-right:8px}
.fbtn:hover{border-color:var(--coral);color:var(--coral)}
.foot{margin-top:22px;font-family:"DM Mono",monospace;font-size:11px;color:var(--wait);text-align:center}
@media(max-width:820px){h1{font-size:29px}.crumb{text-align:left}.srch{width:100%}}
"""

JS = r"""
var RAW=__DATA__, D=RAW.rows, IT=RAW.items;
var n=function(v){var x=Math.round(Number(v)||0);return x?x.toLocaleString('ko-KR'):'—';};
var sn=function(v){var x=Math.round(Number(v)||0);return x?(x<0?'−'+Math.abs(x).toLocaleString('ko-KR'):x.toLocaleString('ko-KR')):'—';};
var e=function(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){
 return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];});};
var F={p:'전체',y:'전체',k:'전체',t:'전체',w:'전체',h:'개최',q:''};
var SORT={key:'dt',asc:true}, OPEN={};
function pkB(k){
  if(!k||k==='—'||k==='미확인')return '<span class="bd b-yel">미확인</span>';
  var c=k==='사업자'?'b-cor':'b-mut';
  var lab=k==='사업자'?'사업자':'3.3%';
  return '<span class="bd '+c+'">'+lab+'</span>';}
function tyB(t){if(!t||t==='단일')return '';
  var c=t==='정규'?'b-cor':(t==='키즈'?'b-lit':'b-mut');
  return '<span class="bd '+c+'">'+e(t)+'</span>';}
function stB(s){
  if(!s)return '';
  var m={'수업완료':'b-grn','정산완료':'b-grn','확정':'b-cor','모집중':'b-cor','개강':'b-cor',
         '신청중':'b-lit','기획중':'b-mut','제안중':'b-mut','검토중':'b-mut','미달':'b-off','폐강':'b-off','보류':'b-off'};
  return '<span class="bd '+(m[s]||'b-mut')+'">'+e(s)+'</span>';}
function filt(){
  return D.filter(function(x){
    if(F.p!=='전체'&&x.partner!==F.p)return false;
    if(F.y!=='전체'&&String(x.dt).slice(0,4)!==F.y)return false;
    if(F.k!=='전체'&&x.kind!==F.k)return false;
    if(F.t!=='전체'&&x.teacher!==F.t)return false;
    if(F.w!=='전체'&&x.paykind!==F.w)return false;
    if(F.h!=='전체'&&x.held!==F.h)return false;
    if(F.q){var q=F.q.toLowerCase();
      if((x.name+' '+x.dt+' '+x.teacher+' '+x.id+' '+x.partner).toLowerCase().indexOf(q)<0)return false;}
    return true;});
}
function detail(x){
  var L=IT[x.id]||[];
  if(L.length<2) return '';
  var h='<tr class="sub"><td colspan="16"><div class="drill">'+
    '<div class="drill-h">'+e(x.id)+' &nbsp;·&nbsp; 과금 항목 '+L.length+'건</div>'+
    '<table><tr><th>항목</th><th>기준(VAT제외)</th><th class="g">기준(VAT포함)</th><th class="g">인원</th><th>클래스매출</th>'+
    '<th class="g">수수료</th><th>갤러리아정산(VAT포함)</th><th class="g">갤러리아정산(VAT제외)</th>'+
    '<th>선생님 정산기준</th><th class="g">선생님정산 계</th><th>PK 수익</th></tr>';
  var T={rev:0,gal:0,galx:0,base:0,pay:0,pk:0};
  L.forEach(function(i){
    for(var k in T) T[k]+=+i[k]||0;
    h+='<tr><td>'+e(i.it)+'</td><td class="n">'+n(i.sup)+'</td><td class="n g">'+n(i.price)+'</td>'+
      '<td class="c g">'+(i.pax||'—')+'</td><td class="n">'+n(i.rev)+'</td>'+
      '<td class="c g">'+(i.rate?Math.round(i.rate*100)+'%':'<span style="color:var(--mute)">0%</span>')+'</td>'+
      '<td class="n">'+n(i.gal)+'</td><td class="n g">'+n(i.galx)+'</td>'+
      '<td class="n">'+n(i.base)+'</td><td class="n g">'+n(i.pay)+'</td>'+
      '<td class="n" style="font-weight:600'+(i.pk<0?';color:#B44A1E':'')+'">'+sn(i.pk)+'</td></tr>';});
  h+='<tr class="tot"><td colspan="3" class="c g">회차 합계</td><td class="c g">'+(x.pax||'—')+'</td>'+
     '<td class="n">'+n(T.rev)+'</td><td class="c g">'+(T.rev?(T.rev-T.gal)/T.rev*100:0).toFixed(1)+'%</td>'+
     '<td class="n">'+n(T.gal)+'</td><td class="n g">'+n(T.galx)+'</td>'+
     '<td class="n">'+n(T.base)+'</td><td class="n g">'+n(T.pay)+'</td>'+
     '<td class="n" style="font-weight:700;color:var(--coral)">'+sn(T.pk)+'</td></tr>';
  return h+'</table></div></td></tr>';
}
function render(){
  var L=filt().slice();
  var k=SORT.key, a=SORT.asc?1:-1;
  L.sort(function(x,y){
    var vx=x[k],vy=y[k];
    if(typeof vx==='number'||typeof vy==='number'){return ((vx||0)-(vy||0))*a;}
    return String(vx).localeCompare(String(vy),'ko')*a;});
  var t={rev:0,fee:0,recv:0,recvx:0,base:0,pay:0,payout:0,cost:0,ours:0,pax:0};
  var h='';
  L.forEach(function(x){
    for(var key in t) t[key]+=+x[key]||0;
    var many=(IT[x.id]||[]).length>1, op=OPEN[x.id];
    h+='<tr class="'+(x.held==='폐강'?'off ':'')+(many?'clk ':'')+(op?'opened':'')+'"'+
      (many?' onclick="tog(\''+x.id+'\')"':'')+'>'+
      '<td class="c" style="font-family:\'DM Mono\',monospace;font-size:11px;color:var(--mute)">'+
      (many?'<span class="cx">'+(op?'▾':'▸')+'</span> ':'')+e(x.id)+'</td>'+
      '<td class="c">'+e(x.partner)+'</td><td class="c">'+e(x.dt)+'</td>'+
      '<td class="c">'+tyB(x.ty)+'</td><td class="nm" title="'+e(x.name)+'">'+e(x.name)+'</td>'+
      '<td class="c">'+e(x.teacher)+'</td><td class="c">'+pkB(x.paykind)+'</td>'+
      '<td class="c">'+stB(x.state)+'</td><td class="c g">'+(x.pax||'—')+'</td>'+
      '<td class="n g">'+n(x.rev)+'</td>'+
      '<td class="c" style="color:var(--sub)"'+(many?' title="'+e(x.st||'')+' — 클릭하면 항목별로 펼쳐집니다"':'')+'>'+
      (x.rate?x.rate+'%'+(many?' <span class="cx">▾</span>':''):'—')+'</td>'+
      '<td class="n">'+n(x.recv)+'</td><td class="n g">'+n(x.recvx)+'</td>'+
      '<td class="n">'+n(x.base)+'</td><td class="n g">'+n(x.pay)+'</td>'+
      '<td class="n pk" style="font-weight:600">'+sn(x.ours)+'</td></tr>';
    if(op) h+=detail(x);});
  h+='<tr class="tot"><td colspan="9" class="c g">합계 '+L.length+'건</td>'+
     '<td class="n g">'+n(t.rev)+'</td><td></td><td class="n">'+n(t.recv)+'</td><td class="n g">'+n(t.recvx)+
     '</td><td class="n">'+n(t.base)+'</td><td class="n g">'+n(t.pay)+'</td><td class="n pk">'+n(t.ours)+'</td></tr>';
  document.getElementById('tb').innerHTML=h;
  document.getElementById('cnt').textContent=L.length+'건 / 전체 '+D.length+'건';
  cards(L);
}
function tog(id){OPEN[id]=!OPEN[id];render();}
function cards(L){
  var o=L.filter(function(x){return x.held==='개최';});
  var s={rev:0,fee:0,recv:0,recvx:0,pay:0,payout:0,ours:0};
  o.forEach(function(x){for(var k in s)s[k]+=+x[k]||0;});
  document.getElementById('cards').innerHTML=
   '<div class="c"><div class="k">개최 회차</div><div class="v">'+o.length+'</div><div class="u">전체 '+L.length+'건 중</div></div>'+
   '<div class="c"><div class="k">클래스매출</div><div class="v" style="font-size:19px">'+n(s.rev)+'</div><div class="u">VAT 포함</div></div>'+
   '<div class="c"><div class="k">갤러리아 정산</div><div class="v" style="font-size:19px">'+n(s.recvx)+'</div><div class="u">VAT 제외 · 수수료 '+n(s.fee)+' 차감</div></div>'+
   '<div class="c"><div class="k">선생님 정산</div><div class="v" style="font-size:19px">'+n(s.pay)+'</div><div class="u">계 · 실지급 '+n(s.payout)+'</div></div>'+
   '<div class="c" style="border-color:var(--coral)"><div class="k">PK 수익</div>'+
   '<div class="v" style="font-size:19px;color:var(--coral)">'+n(s.ours)+'</div>'+
   '<div class="u">클래스매출(VAT제외) 대비 '+(s.rev?(s.ours/(s.rev/1.1)*100).toFixed(1):'0')+'%</div></div>';
}
function mkChipsSel(id,arr,key,sel){
  var el=document.getElementById(id);
  el.innerHTML=el.innerHTML+arr.map(function(v){
    return '<button class="chip'+(v===sel?' on':'')+'" data-v="'+e(v)+'">'+e(v)+'</button>';}).join('');
  el.addEventListener('click',function(ev){var b=ev.target.closest('.chip');if(!b)return;
    this.querySelectorAll('.chip').forEach(function(x){x.classList.remove('on');});
    b.classList.add('on');F[key]=b.dataset.v;OPEN={};render();});
}
function mkChips(id,arr,key){
  var el=document.getElementById(id);
  el.innerHTML=el.innerHTML+arr.map(function(v,i){
    return '<button class="chip'+(i===0?' on':'')+'" data-v="'+e(v)+'">'+e(v)+'</button>';}).join('');
  el.addEventListener('click',function(ev){var b=ev.target.closest('.chip');if(!b)return;
    this.querySelectorAll('.chip').forEach(function(x){x.classList.remove('on');});
    b.classList.add('on');F[key]=b.dataset.v;OPEN={};render();});
}
var uniq=function(f){return ['전체'].concat(Array.from(new Set(D.map(f).filter(Boolean))).sort());};
mkChips('fP',uniq(function(x){return x.partner;}),'p');
mkChips('fY',uniq(function(x){return String(x.dt).slice(0,4);}),'y');
mkChips('fK',uniq(function(x){return x.kind;}),'k');
mkChips('fT',uniq(function(x){return x.teacher;}),'t');
mkChips('fW',uniq(function(x){return x.paykind;}),'w');
mkChipsSel('fH',['전체','개최','폐강'],'h','개최');
document.getElementById('q').addEventListener('input',function(){F.q=this.value;OPEN={};render();});
document.querySelectorAll('th[data-k]').forEach(function(th){
  th.addEventListener('click',function(){
    var k=th.dataset.k;
    if(SORT.key===k)SORT.asc=!SORT.asc; else {SORT.key=k;SORT.asc=true;}
    render();});
});
document.getElementById('dlx').addEventListener('click',function(ev){ev.preventDefault();
  var C=[['회차ID','id'],['제휴사','partner'],['수업일','dt'],['유형','ty'],['클래스명','name'],
    ['선생님','teacher'],['지급방식','paykind'],['상태','state'],['개최','held'],['인원','pax'],
    ['클래스매출','rev'],['갤러리아정산 VAT포함','recv'],['갤러리아정산 VAT제외','recvx'],
    ['선생님 정산기준','base'],['선생님정산 계','pay'],['최종정산금액','payout'],['PK 수익','ours'],['출처','src']];
  var q=function(v){v=String(v==null?'':v);return /[",\n]/.test(v)?'"'+v.replace(/"/g,'""')+'"':v;};
  var rows=[C.map(function(c){return c[0];}).join(',')];
  filt().forEach(function(x){rows.push(C.map(function(c){return q(x[c[1]]);}).join(','));});
  var b=new Blob(['\ufeff'+rows.join('\r\n')],{type:'text/csv;charset=utf-8'});
  var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='클래스목록.csv';a.click();});
render();
"""

doc = f"""<!DOCTYPE html><html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex, nofollow">
<meta name="version" content="{VERSION}"><meta name="updated" content="{UPDATED}">
<title>클래스 목록</title>
<link rel="apple-touch-icon" href="app-icon.png"><link rel="icon" href="app-icon.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Hahmlet:wght@400;700;800&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css" rel="stylesheet">
<style>{CSS}</style></head><body><div class="wrap">

<div class="topbar">
<div><a class="pill" href="class.html">← 클래스 홈</a><a class="pill" href="{CHAT_URL}" target="_blank" rel="noopener">💬 작업 대화</a></div>
<div class="crumb">AI 워크스페이스 › 클래스 › 클래스 목록<br><span class="meta">업데이트 {UPDATED} · {VERSION}</span></div>
</div>

<div class="head">
<h1>클래스 목록</h1>
<p class="lead">우리가 해온 <b>모든 클래스</b>입니다. 클래스 마스터와 제휴사 마스터를 함께 읽습니다. 제휴사별 상세는 <a href="class-partner.html" style="color:var(--coral)">제휴사별 실적</a>에서 봅니다.</p>
</div>

<div class="cards" id="cards"></div>

<div class="bar" id="fP"><b>제휴사</b></div>
<div class="bar" id="fY"><b>연도</b></div>
<div class="bar" id="fK"><b>구분</b></div>
<div class="bar" id="fT"><b>선생님</b></div>
<div class="bar" id="fW"><b>지급방식</b></div>
<div class="bar" id="fH"><b>개최</b></div>
<div class="bar" style="margin-top:4px"><input class="srch" id="q" placeholder="수업명 · 날짜 · 선생님 · 회차ID 검색"></div>
<div class="cnt" id="cnt"></div>

<div class="tw"><table>
<thead>
<tr class="grp">
<th colspan="9">클래스</th>
<th>클래스매출</th>
<th colspan="3">갤러리아 정산</th>
<th colspan="2">선생님 정산</th>
<th class="pk">PK 수익</th>
</tr>
<tr>
<th data-k="id">회차ID</th><th data-k="partner">제휴사</th><th data-k="dt">수업일</th>
<th data-k="ty">유형</th><th data-k="name">클래스명</th><th data-k="teacher">선생님</th>
<th data-k="paykind">지급방식</th><th data-k="state">상태</th><th data-k="pax" class="g">인원</th>
<th data-k="rev" class="g">VAT 포함</th>
<th data-k="rate">수수료</th><th data-k="recv">VAT 포함</th><th data-k="recvx" class="g">VAT 제외</th>
<th data-k="base">정산기준<span class="sub2">매출(VAT제외)</span></th><th data-k="pay" class="g">계</th>
<th data-k="ours" class="pk">차액</th>
</tr></thead><tbody id="tb"></tbody></table></div>

<p class="note">※ <b>▸ 표시가 있는 행을 클릭하면 과금 항목별로 펼쳐집니다.</b> 정규는 강사료·재료비로 나뉩니다.<br>
※ <b>열 제목을 클릭하면 정렬</b>됩니다. 한 번 더 누르면 역순.<br>
※ <b>회사 이익</b> = 공급가 매출 − 제휴 수수료 − 재료·배송비 − 선생님 지급(계약 기준). <b>모든 금액은 부가세를 뺀 기준</b>이며, 인건비·임대료 같은 고정비는 아직 빼지 않은 금액입니다.<br>
※ <b>선생님 지급</b>이 두 열인 이유 — <b>계약 기준</b>은 드리기로 한 금액(공급가의 일정 비율), <b>실지급</b>은 통장에서 실제로 나간 금액입니다.
사업자 선생님은 부가세가 얹혀 더 나가고(나중에 매입세액공제로 환급), 프리랜서 선생님은 원천세를 떼고 나갑니다(우리가 대신 납부).
<b>손익 계산에는 계약 기준을 씁니다.</b> <b>지급방식</b> 열로 어느 쪽인지 구분됩니다 — <span class="bd b-cor">사업자</span> 부가세 가산 · <span class="bd b-mut">3.3%</span> 원천세 공제.<br>
※ 온라인은 재료·패킹·배송을 우리가 부담해 <b>재료·배송비</b>가 있고, 오프라인 제휴는 선생님이 재료를 준비해 <b>제휴 수수료</b>가 있습니다.</p>

<div class="files">
<div class="files-l">📦 원본 파일</div>
<a class="fbtn" href="#" id="dlh">📄 이 화면 HTML 받기</a>
<a class="fbtn" href="#" id="dlx">📊 목록 CSV 받기</a>
<a class="fbtn" href="{CHAT_URL}" target="_blank" rel="noopener">💬 대화창에서 전체 받기</a>
</div>
<div class="foot">클래스 목록 · {UPDATED} · {VERSION}</div>
</div>
<script>
document.getElementById('dlh').addEventListener('click',function(ev){{ev.preventDefault();
var b=new Blob(['<!DOCTYPE html>'+document.documentElement.outerHTML],{{type:'text/html'}});
var a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='class-list.html';a.click();}});
{JS.replace('__DATA__',DATA)}
</script></body></html>"""

open("class-list.html","w",encoding="utf-8").write(doc)
print(f"class-list.html {VERSION} — 총 {len(ROWS)}건")
from collections import Counter
for k,v in Counter(x['partner'] for x in ROWS).most_common(): print(f"    {k} {v}건")
print(f"  크기 {os.path.getsize('class-list.html')/1024:.0f}KB")
