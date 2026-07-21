# -*- coding: utf-8 -*-
"""
클래스 홈 생성기 (박제형 · 유형 B) — v2 마스터(회차·콘텐츠·수강생) 대응
입력: class-master-all-v*.xlsx
출력: class.html (3층 구조: 진행중 / 주요기능 / 도구)
사용법: python gen_class.py class-master-all-v2.xlsx
개인정보: 수강생 연락처·주소·이름은 굽지 않음. 인원 집계만.
"""
import sys, html, json, datetime
from collections import Counter
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "class-master-all-v2.xlsx"
TODAY = datetime.date.today()
KST = datetime.timezone(datetime.timedelta(hours=9))
UPDATED = datetime.datetime.now(KST).strftime("%Y.%m.%d %H:%M")
CHAT_URL = "https://claude.ai/chat/dce33b2f-5756-4005-9648-0eb426113e53"
VERSION = "v19"

ACTIVE = ["검토중","개요서","확정","모집중","개강"]   # 1층 진행 중
def to_date(v):
    if isinstance(v,datetime.datetime): return v.date()
    if isinstance(v,datetime.date): return v
    if isinstance(v,str):
        s=v.strip().replace(".","-").replace("/","-")
        for f in ("%Y-%m-%d","%y-%m-%d","%m-%d"):
            try:
                d=datetime.datetime.strptime(s,f).date()
                if f=="%m-%d": d=d.replace(year=TODAY.year)
                return d
            except: continue
    return None
def fmt_date(d): return "-" if not d else f"{d.month}/{d.day}({'월화수목금토일'[d.weekday()]})"
def ymd_full(d): return "" if not d else f"{d.year}.{d.month}.{d.day} ({'월화수목금토일'[d.weekday()]})"
def md_short(d): return "" if not d else f"{d.month}/{d.day}"
def esc(x): return html.escape(str(x)) if x is not None else ""

wb=load_workbook(SRC,data_only=True)
ws=wb["회차"]; wstu=wb["수강생"]
wcon=wb["콘텐츠"] if "콘텐츠" in wb.sheetnames else None

cnt=Counter()
for r in range(2,wstu.max_row+1):
    cid=wstu.cell(row=r,column=1).value
    if cid: cnt[str(cid).strip()]+=1

content_map={}
if wcon:
    for r in range(2,wcon.max_row+1):
        ccid=wcon.cell(row=r,column=1).value
        if not ccid: continue
        cg=lambda c: wcon.cell(row=r,column=c).value
        content_map[str(ccid).strip()]=dict(강의명=cg(2),소개문=cg(3),강의내용=cg(4),대상=cg(5),
            선발방식=cg(6),소요시간=cg(7),키트구성품=cg(8),개별준비물=cg(9),강사이력=cg(10),
            해시태그=cg(11),배송안내=cg(12),참여안내=cg(13),장소상세=cg(14),동반인규정=cg(15),
            재료구성=cg(16),줌링크=cg(17),견적코드=cg(18),레시피=cg(19))

# 발주 시트 → 회차ID별 재료별 발주 정보
order_map={}
if "발주" in wb.sheetnames:
    word=wb["발주"]
    for r in range(2,word.max_row+1):
        ocid=word.cell(row=r,column=1).value
        onm=word.cell(row=r,column=2).value
        if not ocid or not onm: continue
        og=lambda c: word.cell(row=r,column=c).value
        def _s(v): return "" if v is None else str(v).strip()
        rec=dict(keep=_s(og(3)),store=_s(og(4)),odate=_s(og(5)),idate=_s(og(6)),
                 ordered=(_s(og(7)) in("완료","O","o","✓","true","True","1")),
                 arrived=(_s(og(8)) in("완료","O","o","✓","true","True","1")),
                 qty=_s(og(9)),realamt=_s(og(10)),stock=(_s(og(11)) in("재고","O","o","✓","true","True","1")))
        order_map.setdefault(str(ocid).strip(),{})[_s(onm)]=rec

sessions=[]
for r in range(2,ws.max_row+1):
    cid=ws.cell(row=r,column=1).value
    if not cid: continue
    cid=str(cid).strip(); g=lambda c: ws.cell(row=r,column=c).value
    kind=(g(4) or "").strip(); mtype=(g(8) or "").strip()
    d=to_date(g(5))
    try: mn=int(g(9) or 0)
    except: mn=0
    try: mx=int(g(10) or 0)
    except: mx=0
    applied=cnt.get(cid,0)
    state=(g(34) or "").strip()
    # 개폐강 판정 (A형=사전확정 / B형=수업일 D-4 자정 마감 기준)
    if mtype=="A":
        openst="사전확정"
    elif d is None:
        openst="신청중"                       # 수업일 미정
    else:
        deadline=d-datetime.timedelta(days=4)  # 신청 마감 = 수업일 D-4
        if TODAY<=deadline:
            openst="신청중"                    # 마감 전
        elif mn>0 and applied>=mn:
            openst="확정"                      # 마감 후 최소 달성
        elif mn>0:
            openst="미달"                      # 마감 후 최소 미달
        else:
            openst="신청중"
    signup_deadline=(d-datetime.timedelta(days=4)) if d else None   # 개최 문자 발송일(D-4 09:00)
    def N(c):
        try: return float(g(c))
        except: return None
    sessions.append(dict(cid=cid,name=g(2) or "(수업명 없음)",orderer=g(3) or "-",kind=kind,mtype=mtype,
        date=d,time=g(6),place=g(7),mn=mn,mx=mx,applied=applied,openst=openst,signup_deadline=signup_deadline,
        price=N(16),vat=(g(17) or "").strip(),struct=(g(18) or "").strip(),
        mat=N(19),pack=N(20),ship=N(21),teach=N(22),profit=N(23),rate=N(24),
        teacher=g(25) or "-",ttype=(g(26) or "").strip(),feeway=(g(27) or "").strip(),feerate=N(28),settle=N(29),
        taxto=g(30),payer=g(31),paywhen=g(32),paydate=g(33),
        confirmday=to_date(g(13)),refundday=to_date(g(14)),refundsrc=g(15),
        state=state or "기획중",
        paidin=(g(38) or "입금대기").strip(),paidout=(g(39) or "지급대기").strip(),
        packfile=(g(40) or "").strip(),
        memo=g(35),content=content_map.get(cid,{})))
def sk(s):
    if s["date"] is None: return (1,datetime.date.max)
    if s["date"]>=TODAY: return (0,s["date"])
    return (2,s["date"])
sessions.sort(key=sk)

# 1층: 상태 기준 그룹 (온/오프 구분 없음)
def is_settled(s): return s["state"]=="수업완료" and s["paidin"]=="입금완료" and s["paidout"]=="지급완료"
plan=[s for s in sessions if s["state"]=="기획중"]
recruit=[s for s in sessions if s["state"]=="모집중"]
confirmed=[s for s in sessions if s["state"]=="확정"]
settle_wait=[s for s in sessions if s["state"]=="수업완료" and not is_settled(s)]  # 수업완료+정산 미완료
upcoming=[s for s in sessions if s["date"] and s["date"]>=TODAY]
n_open=sum(1 for s in sessions if s["state"]=="확정")
n_recruit=len(recruit)
n_stu=sum(s["applied"] for s in sessions)

def rep_badge(s):
    """색 체계: 기획중=회색 / 진행중(모집중·개강확정)=코랄 / 정산남음=주황 / 정산완료=초록 / 폐강=묻힘"""
    st=s["state"]
    if st=="기획중": return ("기획중","st-plan")
    if st=="모집중": return ("모집중","st-live")
    if st=="확정":   return ("개강확정","st-live")
    if st=="미달":   return ("폐강","st-off")
    if st=="보류":   return ("보류","st-off")
    # 수업완료 이후 → 정산 진행 (주황), 다 끝나면 초록
    if s["paidin"]!="입금완료": return ("입금대기","st-wait")
    if s["paidout"]!="지급완료": return ("지급대기","st-wait")
    return ("정산완료","st-done")

def card(s):
    kc="k-online" if s["kind"]=="온라인" else ("k-offline" if s["kind"]=="오프라인" else "")
    # 상단 배지: 주최·형태 (오프라인은 위치)
    if s["kind"]=="오프라인":
        head_txt=(s["place"] or s["orderer"] or "-")+" · 오프라인"
    elif s["kind"]=="온라인":
        head_txt=(s["orderer"] or "-")+" · 온라인"
    else:
        head_txt=s["orderer"] or "-"
    badge_txt,st_cls=rep_badge(s)
    # 정보: 선생님 · 수업일
    teacher=esc(s["teacher"]) if (s["teacher"] and s["teacher"]!="-") else "선생님 미정"
    if s["date"]:
        datestr="수업일 "+ymd_full(s["date"])+((" "+esc(s["time"])) if s["time"] else "")
    else:
        datestr="수업일 미정"
    info=(f'<div class="c-info"><div class="c-tea">👤 {teacher}</div>'
          f'<div class="c-day">📅 {datestr}</div></div>')
    # 신청 현황 (A형=사전확정 / B형=신청·최소·최대 + 개폐강)
    if s["mtype"]=="A":
        pax=f'<div class="c-pax">👥 {s["applied"] or s["mn"]}명 <span class="pax-tag pt-fix">사전확정</span></div>'
    else:
        opcls={"신청중":"pt-ing","확정":"pt-open","미달":"pt-short"}.get(s["openst"],"pt-ing")
        pax=(f'<div class="c-pax">👥 신청 {s["applied"]}명 '
             f'<span class="pax-lim">(최소 {s["mn"]} · 최대 {s["mx"]})</span> '
             f'<span class="pax-tag {opcls}">{s["openst"]}</span></div>')
    # 맨 아래: 총매출 · 원가 · 순이익 (전체 기준, 공급가)
    fin=""
    supply=(s["price"]/1.1 if s["vat"]=="포함" else s["price"]) if s["price"] else None
    n=s["applied"] or s["mn"] or 0
    if supply and n:
        rev=supply*n
        cost=((s["mat"] or 0)+(s["pack"] or 0)+(s["ship"] or 0))*n
        net=rev-cost-(s["settle"] or 0)
        rt=f" ({round(net/rev*100,1)}%)" if rev else ""
        fin=(f'<div class="c-fin"><span class="fi">총매출 {int(round(rev)):,}원</span>'
             f'<span class="fsep">|</span><span class="fi">원가 {int(round(cost)):,}원</span>'
             f'<span class="fsep">|</span><span class="fi net">순이익 {int(round(net)):,}원{rt}</span></div>')
    cid=esc(s["cid"])
    thumb=(f'<img class="c-thumb" src="class-images/{cid}.jpg" alt="" loading="lazy" '
           f'onerror="this.style.display=&quot;none&quot;;this.nextElementSibling.style.display=&quot;flex&quot;">'
           f'<div class="c-thumb c-thumb-ph" style="display:none">🍽</div>')
    return (f'<article class="card {kc}" onclick="openDetail(\'{cid}\')">'
        f'<header class="c-head"><span class="c-src">{esc(head_txt)}</span>'
        f'<span class="badge {st_cls}">{esc(badge_txt)}</span></header>'
        f'<div class="c-top"><div class="c-thumb-wrap">{thumb}</div>'
        f'<div class="c-main"><div class="c-name">{esc(s["name"])}</div>{info}</div></div>'
        f'<div class="c-div"></div>{pax}{fin}</article>')

def group(title,sub,items):
    if not items: return ""
    body='<div class="grid">'+"".join(card(s) for s in items)+'</div>'
    return f'<div class="grp"><div class="grp-h"><h3>{title}</h3><span>{sub}</span></div>{body}</div>'
layer1=group("기획중","개설·준비 중",plan)+group("모집 중","신청 받는 중",recruit)+group("곧 수업","확정 · 준비",confirmed)+group("정산 예정","수업 완료 · 정산 전",settle_wait)
if not layer1: layer1='<div class="empty">진행 중인 클래스가 없습니다. 계산기로 등록하세요.</div>'

# 2층 주요 기능
feat_cards=(
 '<a class="feat" href="class-list.html"><div class="ft-t">전체 클래스 목록</div><div class="ft-d">'+str(len(sessions))+'건 · 필터·검색</div><div class="ft-go">열기 →</div></a>'
 '<a class="feat soon" href="#"><div class="ft-t">정산 리포트</div><div class="ft-d">받을 돈·줄 돈·지급 예정</div><span class="tag">준비 중</span></a>'
)
# 3층 도구
tools=[("재료비·수익률 계산기","등록·견적·수정 겸용","class-quote.html",False),
 ("설계 가이드","클래스 구조·정산·콘텐츠","class-guide.html",False),
 ("컬리 송장 변환","경기도지식 명단 → 컬리","#",True),
 ("메뉴 레퍼런스","수업별 메뉴·레시피","#",True),
 ("사진 관리","대표·후기 사진","#",True)]
tool_cards="".join(
 (f'<a class="tool{" soon" if soon else ""}" href="{href}"'+("" if soon else' ')+'>'
  +('<span class="tag">준비 중</span>' if soon else '')
  +f'<span class="t-name">{esc(n)}</span><span class="t-desc">{esc(d)}</span></a>')
 for n,d,href,soon in tools)

# 전체 목록(2층 앵커) — 모든 회차 요약 행
def listrow(s):
    return (f'<tr><td>{fmt_date(s["date"])}</td><td>{esc(s["orderer"])}</td><td class="ln">{esc(s["name"])}</td>'
     f'<td>{esc(s["kind"])}</td><td><span class="badge2 b-{ "open" if s["openst"] in("개강","확정") else ("short" if s["openst"]=="미달" else "wait")}">{s["openst"]}</span></td>'
     f'<td class="rt">{s["applied"]}</td><td class="st">{esc(s["state"])}</td></tr>')
list_html="".join(listrow(s) for s in sessions) or '<tr><td colspan="9" class="empty">등록된 클래스가 없습니다.</td></tr>'

CSS = """
:root{--bg:#FAF5EC;--ink:#241F1B;--sub:#6B6256;--line:#E7DCC8;--card:#FFFDF9;
--point:#B4A032;--online:#2C7BB6;--offline:#C2570E;--open:#2E7D32;--short:#E2611B;--wait:#9A8F7C;--coral:#FF5019;}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--ink);font-family:Pretendard,-apple-system,BlinkMacSystemFont,"Apple SD Gothic Neo","Malgun Gothic",sans-serif;line-height:1.55;-webkit-font-smoothing:antialiased;padding-bottom:60px}
.wrap{max-width:1080px;margin:0 auto;padding:0 20px}
.topbar{display:flex;justify-content:space-between;align-items:center;padding:18px 0;border-bottom:1px solid var(--line);gap:12px;flex-wrap:wrap}
.tb-left{display:flex;align-items:center;gap:8px}
.pill{display:inline-flex;align-items:center;gap:4px;font-size:13px;text-decoration:none;color:var(--ink);background:var(--card);border:1px solid var(--line);border-radius:20px;padding:5px 13px;transition:.12s}
.pill:hover{border-color:var(--point);color:var(--point)}
.tb-right{display:flex;align-items:center;gap:7px;font-size:12px;color:var(--wait);text-align:right}
.crumb{color:var(--wait)}.tb-sep{color:#CFC5B2}.updated{font-family:"DM Mono",monospace;font-size:12px;color:var(--wait)}
.hero{padding:30px 0 8px}
.hero h1{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-weight:700;font-size:40px;letter-spacing:-.02em}
.hero .lead{margin-top:8px;color:var(--sub);font-size:15px}
.summary{display:flex;flex-wrap:wrap;margin:22px 0 6px;border:1px solid var(--line);border-radius:14px;overflow:hidden;background:var(--card)}
.s-item{flex:1;min-width:120px;padding:15px 18px;border-right:1px solid var(--line)}
.s-item:last-child{border-right:0}
.s-num{font-family:"DM Mono",monospace;font-size:25px;font-weight:500;line-height:1}
.s-lab{font-size:12px;color:var(--wait);margin-top:6px}
.s-open .s-num{color:var(--open)}.s-short .s-num{color:var(--short)}.s-recruit .s-num{color:var(--short)}
.layer{margin-top:34px}
.layer-h{display:flex;align-items:baseline;gap:10px;margin-bottom:4px}
.layer-h h2{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-size:22px;font-weight:700}
.layer-h .ls{font-size:12px;color:var(--wait)}
.layer-h .bar{flex:1;height:2px;background:var(--ink);align-self:center;margin-left:6px;opacity:.85}
.grp{margin-top:16px}
.grp-h{display:flex;align-items:baseline;gap:8px;margin-bottom:10px}
.grp-h h3{font-size:15px;font-weight:700}.grp-h span{font-size:12px;color:var(--wait)}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px}
.card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:15px;border-top:3px solid var(--line);transition:.12s;cursor:pointer}
.card:hover{transform:translateY(-2px);box-shadow:0 6px 18px rgba(36,31,27,.08)}
/* 형태별 카드 테두리 색 제거 — 색은 상태 배지에만 */
.c-head{display:flex;justify-content:space-between;align-items:center;gap:8px;margin-bottom:9px}
.c-top{display:flex;gap:12px;align-items:flex-start}
.c-thumb-wrap{flex-shrink:0}
.c-thumb{width:66px;height:66px;border-radius:11px;object-fit:cover;display:block;background:#EEE9DD}
.c-thumb-ph{align-items:center;justify-content:center;font-size:26px;color:#B8AE98}
.c-main{flex:1;min-width:0}
.c-src{font-size:12px;color:var(--sub);font-weight:600;background:#F4EFE3;padding:4px 10px;border-radius:20px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
/* c-src 온/오프 색 제거 — 기본 회색 유지 */
.badge{font-size:11px;font-weight:700;padding:3px 10px;border-radius:20px;font-family:"DM Mono",monospace;white-space:nowrap;flex-shrink:0}
.st-plan{background:#F3ECD7;color:#A08A4E}
.st-live{background:#FFE7DF;color:var(--coral)}
.st-wait{background:#E4EEF6;color:var(--online)}
.st-done{background:#EAE6DE;color:#B0A794}
.st-off{background:#EAE6DE;color:#BFB6A6;text-decoration:line-through}
.c-name{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-size:19px;font-weight:700;margin:2px 0 8px}
.c-info{font-size:12.5px;color:var(--sub);line-height:1.7}
.c-info .c-day{font-family:"DM Mono",monospace;font-size:12px;color:var(--ink)}
.c-div{border-top:1px dashed var(--line);margin:11px 0 10px}
.c-pax{font-size:13px;color:var(--ink);display:flex;align-items:center;flex-wrap:wrap;gap:6px}
.pax-lim{font-size:11.5px;color:var(--wait);font-family:"DM Mono",monospace}
.pax-tag{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:10px;font-family:"DM Mono",monospace}
.pt-fix{background:#EEE9DD;color:var(--sub)}.pt-ing{background:#E8F1F8;color:var(--online)}
.pt-open{background:#E6F4E6;color:var(--open)}.pt-short{background:#F6DDD2;color:#B44A1E}
.c-fin{display:flex;flex-wrap:wrap;align-items:center;gap:7px;margin-top:11px;padding-top:10px;border-top:1px solid var(--line);font-family:"DM Mono",monospace;font-size:12px;color:var(--sub)}
.c-fin .fsep{color:#CFC5B2}.c-fin .net{color:var(--open);font-weight:700}
.empty{padding:26px;text-align:center;color:var(--wait);font-size:13.5px;background:var(--card);border:1px dashed var(--line);border-radius:12px}
.feats{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));gap:12px;margin-top:14px}
.feat{position:relative;display:block;padding:18px;background:var(--card);border:1px solid var(--line);border-radius:14px;text-decoration:none;color:var(--ink);transition:.12s}
.feat:hover{border-color:var(--point);transform:translateY(-2px)}
.feat.soon{opacity:.55;pointer-events:none}
.ft-t{font-weight:700;font-size:16px}.ft-d{font-size:12.5px;color:var(--wait);margin-top:3px}.ft-go{font-size:12px;color:var(--point);margin-top:8px}
.tools{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-top:14px}
.tool{position:relative;display:flex;flex-direction:column;gap:4px;padding:17px;background:var(--card);border:1px solid var(--line);border-radius:14px;text-decoration:none;color:var(--ink);transition:.12s}
.tool:hover{border-color:var(--point);transform:translateY(-2px)}
.tool.soon{opacity:.55;pointer-events:none}
.t-name{font-weight:700;font-size:15px}.t-desc{font-size:12px;color:var(--wait)}
.tag{position:absolute;top:12px;right:12px;font-size:10px;background:#EEE9DD;color:var(--wait);padding:2px 7px;border-radius:10px;font-family:"DM Mono",monospace}
.listwrap{margin-top:14px;overflow-x:auto;border:1px solid var(--line);border-radius:12px;background:var(--card)}
table.list{width:100%;border-collapse:collapse;font-size:13px;min-width:600px}
table.list th,table.list td{padding:9px 12px;border-bottom:1px solid var(--line);text-align:left}
table.list th{background:#F4EFE3;font-size:12px;color:var(--sub);font-weight:700}
table.list tr:last-child td{border-bottom:none}
table.list .ln{font-weight:600}.table .rt,.list .rt{text-align:center;font-family:"DM Mono",monospace}
table.list .st{color:var(--online);font-size:12px}
.badge2{font-size:10.5px;font-weight:700;padding:2px 7px;border-radius:10px;font-family:"DM Mono",monospace}
.badge2.b-open{background:#E6F4E6;color:var(--open)}.badge2.b-short{background:#FCE8DD;color:var(--short)}.badge2.b-wait{background:#EEE9DD;color:var(--wait)}
.backup-row{display:flex;flex-wrap:wrap;gap:10px;margin-top:14px}
.bk-btn{display:inline-flex;align-items:center;gap:6px;font-size:13.5px;font-family:inherit;color:var(--ink);background:var(--card);border:1px solid var(--line);border-radius:12px;padding:12px 16px;cursor:pointer;transition:.12s;text-decoration:none}
.bk-btn:hover{border-color:var(--point);color:var(--point)}
.bk-tag{font-family:"DM Mono",monospace;font-size:10px;background:#EEE9DD;color:var(--wait);padding:2px 7px;border-radius:10px;margin-left:2px}
.backup-note{margin-top:11px;font-size:12.5px;color:var(--sub);line-height:1.9}
.backup-note code{font-family:"DM Mono",monospace;font-size:11.5px;background:#F4EFE3;padding:2px 7px;border-radius:5px}
.backup-note a{color:var(--point);text-decoration:none;border-bottom:1px solid var(--point)}
.ops-note{margin-top:30px;padding:18px 20px;background:var(--card);border:1px solid var(--line);border-radius:12px;font-size:12.5px;color:var(--sub);line-height:1.9}
.ops-note b{color:var(--ink)}
.foot{margin-top:26px;padding-top:16px;border-top:1px solid var(--line);font-family:"DM Mono",monospace;font-size:11px;color:var(--wait);text-align:center}
.packbar{position:relative;overflow:hidden;background:linear-gradient(135deg,#FFF6E0 0%,#FDEBD2 100%);border:1px solid #EBD8AE;border-radius:16px;padding:18px 22px;margin-top:18px;display:flex;gap:18px;align-items:center;flex-wrap:wrap;box-shadow:0 1px 3px rgba(180,160,50,.10)}
.packbar::before{content:"";position:absolute;left:0;top:0;bottom:0;width:5px;background:var(--point)}
.packbar .pb-ic{font-size:26px;line-height:1;flex-shrink:0;margin-left:2px}
.packbar .pb-l{flex:1;min-width:230px}
.packbar .pb-k{display:inline-flex;align-items:center;gap:6px;font-family:"DM Mono",monospace;font-size:11px;font-weight:500;color:#fff;background:var(--point);border-radius:20px;padding:3px 11px;letter-spacing:.06em}
.packbar .pb-n{font-size:17px;font-weight:700;margin-top:7px;letter-spacing:-.01em;line-height:1.35}
.packbar .pb-d{font-family:"DM Mono",monospace;font-size:12.5px;color:var(--sub);margin-top:5px}
.packbar a.pb-b{font-size:14px;font-weight:600;text-decoration:none;color:#fff;background:var(--point);border-radius:11px;padding:12px 22px;white-space:nowrap;transition:.14s;box-shadow:0 2px 6px rgba(180,160,50,.28)}
.packbar a.pb-b:hover{filter:brightness(1.08);transform:translateY(-1px)}
@media(max-width:560px){.packbar{padding:16px 18px}.packbar a.pb-b{width:100%;text-align:center}}
@media(max-width:520px){.hero h1{font-size:30px}.s-item{min-width:50%;flex:none}.topbar{align-items:flex-start}.tb-right{flex-direction:column;align-items:flex-end;gap:2px}.tb-sep{display:none}}
/* 자세히보기 모달 */
.modal{position:fixed;inset:0;z-index:50;display:none}
.modal.on{display:block}
.modal-bg{position:absolute;inset:0;background:rgba(36,31,27,.42);backdrop-filter:blur(2px)}
.modal-panel{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);width:min(960px,95vw);height:auto;min-height:min(60vh,520px);max-height:92vh;background:var(--bg);border:1px solid var(--line);border-radius:16px;box-shadow:0 20px 60px rgba(36,31,27,.28);display:flex;flex-direction:column;overflow:hidden}
.modal-head{display:flex;justify-content:space-between;align-items:flex-start;padding:20px 22px 14px;border-bottom:1px solid var(--line)}
.m-orderer{font-size:12px;color:var(--wait);font-weight:500}
.modal-head h2{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-size:23px;font-weight:700;margin-top:3px}
.m-actions{display:flex;gap:8px;align-items:center;flex-shrink:0}
.m-edit{font-size:13px;text-decoration:none;color:#fff;background:var(--point);border-radius:8px;padding:7px 13px;white-space:nowrap}
.m-edit:hover{filter:brightness(1.06)}
.m-pack{font-size:13px;text-decoration:none;color:var(--point);background:#FFFDF9;border:1px solid var(--point);border-radius:8px;padding:6px 12px;white-space:nowrap;margin-right:6px}
.m-pack:hover{background:var(--point);color:#fff}
.m-close{border:none;background:var(--card);border:1px solid var(--line);border-radius:8px;width:32px;height:32px;cursor:pointer;font-size:14px;color:var(--sub)}
.m-close:hover{border-color:var(--point);color:var(--point)}
.modal-tabs{display:flex;gap:2px;padding:0 22px;border-bottom:1px solid var(--line);overflow-x:auto}
.mtab{font-family:inherit;font-size:13px;color:var(--sub);background:none;border:none;border-bottom:2px solid transparent;padding:11px 12px;cursor:pointer;white-space:nowrap;margin-bottom:-1px}
.mtab:hover{color:var(--ink)}
.mtab.on{color:var(--point);border-bottom-color:var(--point);font-weight:700}
.modal-body{padding:18px 22px 24px;overflow-y:auto;flex:1;scrollbar-gutter:stable}
.mpanel{display:none}.mpanel.on{display:block}
.drow{display:flex;gap:12px;padding:8px 0;border-bottom:1px dashed var(--line);font-size:13.5px}
.drow:last-child{border-bottom:none}
.drow .k{color:var(--sub);min-width:96px;flex-shrink:0}
.drow .v{color:var(--ink);font-weight:500;white-space:pre-wrap}
.drow .v.mono{font-family:"DM Mono",monospace}
.dsub{font-size:12px;color:var(--point);font-weight:700;margin:14px 0 4px}
.dsub:first-child{margin-top:0}
.dempty{color:var(--wait);font-size:13px;padding:10px 0}
.dnote{margin-top:14px;font-size:12px;color:var(--wait);background:var(--card);border:1px solid var(--line);border-radius:8px;padding:10px 12px}
.dtext{white-space:pre-wrap;font-size:13.5px;line-height:1.65;color:var(--ink);margin-top:2px}
.dcopy{display:inline-flex;align-items:center;gap:5px;font-family:inherit;font-size:12.5px;color:#fff;background:var(--point);border:none;border-radius:8px;padding:8px 14px;cursor:pointer;margin-bottom:12px}
.dcopy:hover{filter:brightness(1.06)}
.dbig{display:flex;justify-content:space-between;align-items:center;background:#EEF6EE;border:1px solid #CBE5CB;border-radius:10px;padding:13px 16px;margin-bottom:12px}
.dbig .bk{font-size:13px;color:var(--open);font-weight:600}
.dbig .bv{font-family:"DM Mono",monospace;font-size:20px;font-weight:700;color:var(--open)}
table.mat{border-collapse:collapse;width:100%;min-width:820px;font-size:12px;margin-top:4px}
table.mat th,table.mat td{border:1px solid var(--line);padding:5px 7px;text-align:center;white-space:nowrap}
table.mat th{background:#F4EFE3;font-size:11px;color:var(--sub);font-weight:700}
table.mat th.grp{background:#EFE7D4}
table.mat td.tl{text-align:left;font-weight:600}
table.mat td.rt{text-align:right;font-family:"DM Mono",monospace}
table.mat td.a{background:#F3F7FA;color:var(--online)}
table.mat td.p{background:#EAF3FA;font-weight:700}
table.mat tr.subtot td{background:#F0EADB;font-weight:700;color:var(--sub);border-top:1.5px solid #D8CDB4}
table.mat tr.subtot td.tl{text-align:right;padding-right:10px}
table.mat tr.subtot td.p{background:#E4EDD9;color:var(--open)}
table.mat a{color:var(--point);text-decoration:none}
.ship-note{font-size:9.5px;color:var(--offline);font-weight:600}
.matsum{margin-top:10px;text-align:right;font-family:"DM Mono",monospace;font-size:14px}.matsum b{font-size:17px;color:var(--online)}
.ordsec{margin-top:22px;padding-top:18px;border-top:2px solid var(--line)}
.ordh{font-family:Hahmlet,serif;font-size:16px;font-weight:700;margin-bottom:12px}.ordh .s{font-family:Pretendard;font-size:11.5px;color:var(--wait);font-weight:400;margin-left:8px}
.ordwrap{overflow-x:auto}
table.ord{width:100%;border-collapse:collapse;font-size:11px;table-layout:auto}
table.ord th{background:#F1EADB;padding:6px 7px;border:1px solid var(--line);font-weight:600;color:var(--sub);font-size:10.5px}
table.ord td{padding:5px 7px;border:1px solid var(--line);white-space:nowrap}
table.ord th{white-space:nowrap}
.nw{white-space:nowrap}.mut{color:#CFC5B2}
.gu-mini{font-size:8.5px;background:#EFEADD;color:var(--wait);padding:1px 5px;border-radius:6px;font-weight:600}
table.ord .tl{text-align:left}table.ord .rt{text-align:right;font-family:"DM Mono",monospace}
table.ord .ck{text-align:center;font-size:13px}table.ord .sm{font-size:10.5px;color:var(--wait)}
table.ord a{text-decoration:none}
.kp{font-size:10px;font-weight:700;padding:2px 7px;border-radius:9px;font-family:"DM Mono",monospace}
.kp-frz{background:#E4EEF6;color:var(--online)}.kp-cool{background:#E6F4E6;color:var(--open)}.kp-room{background:#EEE9DD;color:var(--wait)}.kp-none{background:transparent;color:#CFC5B2}
.stk{font-size:9.5px;background:#EFEADD;color:var(--point);padding:1px 6px;border-radius:8px;font-weight:700}
.ordnote{font-size:11.5px;color:var(--offline);margin-top:9px;font-weight:600}
.ordnote2{font-size:11px;color:var(--wait);margin-top:10px;line-height:1.6}
.cmpwrap{margin-top:14px}
table.cmp{border-collapse:collapse;font-size:12.5px;min-width:340px}
table.cmp th{background:#F1EADB;padding:8px 14px;border:1px solid var(--line);font-weight:700}
table.cmp td{padding:8px 14px;border:1px solid var(--line)}
table.cmp .tl{text-align:left}table.cmp .rt{text-align:right;font-family:"DM Mono",monospace}
table.cmp .cmp-net td{font-weight:700;background:#FBF8EF}table.cmp .cmp-net .rt{color:var(--open)}
.pend{font-size:11px;color:var(--wait);font-style:italic}
.cmpsub{font-size:9.5px;color:var(--wait);font-weight:400;display:block}
.plan2{display:block;font-size:9px;color:var(--wait);font-weight:400}
.ost{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:9px;white-space:nowrap}
.ost-ok{background:#E6F4E6;color:var(--open)}
.ost-short{background:#FFE7DF;color:var(--coral)}
.ost-stock{background:#E4EEF6;color:var(--online)}
.ost-none{background:transparent;color:#CFC5B2;font-weight:400}
.sm2{color:var(--wait);font-size:10.5px}
.dfsave{color:var(--open);font-weight:700}
.dfover{color:var(--coral);font-weight:700}
.df0{color:var(--wait)}
.matwrap{overflow-x:auto}
.flow{border:1px solid var(--line);border-radius:12px;overflow:hidden;margin-top:4px}
.flow-sec{border-bottom:1px solid var(--line)}.flow-sec:last-child{border-bottom:none}
.flow-h{background:#F4EFE3;padding:8px 14px;font-size:12px;font-weight:700;color:var(--sub)}
.flow-h2{display:flex;justify-content:space-between;align-items:center;gap:8px}
.fh-badge{font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;font-family:"DM Mono",monospace}
.fhb-done{background:#E6F4E6;color:var(--open)}.fhb-wait{background:#EEE9DD;color:var(--wait)}
.d-photo{margin-bottom:14px}
.d-photo img{width:100%;height:150px;object-fit:cover;border-radius:12px;display:block}
.flow-r{display:flex;justify-content:space-between;padding:8px 14px;font-size:13.5px;border-top:1px dashed var(--line)}
.flow-r:first-of-type{border-top:none}.flow-r .fv{font-family:"DM Mono",monospace}
.flow-r.tot{font-weight:700;background:#FBFAF6}.flow-r.tot .fv{font-size:15px}
.flow-r.pos .fv{color:var(--open)}.flow-r.neg .fv{color:var(--short)}
.bigpay{display:flex;justify-content:space-between;align-items:center;background:#EEF6EE;border:1px solid #CBE5CB;border-radius:10px;padding:13px 16px;margin:4px 0 12px}
.bigpay .bk{font-size:13px;color:var(--open);font-weight:600}.bigpay .bv{font-family:"DM Mono",monospace;font-size:20px;font-weight:700;color:var(--open)}
"""

JS = """
function tsN(base,ext){var n=new Date(),p=function(x){return(x<10?'0':'')+x};return base+'_'+n.getFullYear()+p(n.getMonth()+1)+p(n.getDate())+'.'+ext;}
function dl(bl,f){var u=URL.createObjectURL(bl),a=document.createElement('a');a.href=u;a.download=f;document.body.appendChild(a);a.click();setTimeout(function(){URL.revokeObjectURL(u);a.remove();},100);}
function saveHTML(){dl(new Blob(['<!DOCTYPE html>\\n'+document.documentElement.outerHTML],{type:'text/html;charset=utf-8'}),'class.html');}
function toCSV(rows){if(!rows.length)return'';var k=Object.keys(rows[0]);var e=function(v){v=(v==null?'':''+v);return /[",\\n]/.test(v)?'"'+v.replace(/"/g,'""')+'"':v;};var o=k.join(',')+'\\n';rows.forEach(function(r){o+=k.map(function(x){return e(r[x]);}).join(',')+'\\n';});return '\\ufeff'+o;}
function saveXLSX(){var rows=window.CLASS_ROWS||[];if(!rows.length){alert('내보낼 데이터가 없습니다.');return;}
if(typeof XLSX!=='undefined'){try{var ws=XLSX.utils.json_to_sheet(rows);var wb=XLSX.utils.book_new();XLSX.utils.book_append_sheet(wb,ws,'회차');XLSX.writeFile(wb,tsN('class-master-open','xlsx'));return;}catch(e){}}
dl(new Blob([toCSV(rows)],{type:'text/csv;charset=utf-8'}),tsN('class-master-open','csv'));}
(function(){var u=document.querySelector('meta[name=updated]'),v=document.querySelector('meta[name=version]');var el=document.getElementById('upd');if(el)el.textContent='업데이트 '+((u&&u.content)||'')+' · '+((v&&v.content)||'');})();

var D=window.CLASS_DETAIL||{};var CUR=null;
function won(n){if(n==null||n==='')return '-';return Math.round(Number(n)).toLocaleString()+'원';}
function pct(n){if(n==null||n==='')return '-';return (Math.round(n*1000)/10)+'%';}
function num(v){v=parseFloat((''+v).replace(/,/g,''));return isNaN(v)?0:v;}
function fmtN(n){if(n==null||n==='')return '';return Math.round(n).toLocaleString();}
function payChip(v,kind){var done=(v===(kind+'완료'));return '<span class="paychip '+(done?'pc-done':'pc-wait')+'">'+(done?'✓ 완료':'대기')+'</span>';}
function esc2(v){return v==null?'':String(v).replace(/</g,'&lt;');}
function row(k,v,mono){if(v==null||v===''||v==='-')return '';return '<div class="drow"><span class="k">'+k+'</span><span class="v'+(mono?' mono':'')+'">'+esc2(v)+'</span></div>';}
function blk(label,text){return '<div class="dsub">'+label+'</div>'+(text?('<div class="dtext">'+esc2(text)+'</div>'):'<div class="dempty">미입력</div>');}
function drowFlow(k,v){if(v==null||v==='')return '';return '<div class="flow-r"><span>'+k+'</span><span class="fv">'+esc2(v)+'</span></div>';}
function payBadge(v,kind){if(!v)return '';var done=(v===kind+'완료');return '<span class="fh-badge '+(done?'fhb-done':'fhb-wait')+'">'+esc2(v)+'</span>';}
function parseQ(code){if(!code)return null;var m=(''+code).match(/QDATA:([A-Za-z0-9+\\/=]+)/);if(!m)return null;try{return JSON.parse(decodeURIComponent(escape(atob(m[1]))));}catch(e){return null;}}
function matTable(code,paxFallback){
  var d=parseQ(code);if(!d||!d.rows||!d.rows.length)return '<div class="dempty">재료 정보가 없습니다. 계산기에서 등록하면 채워집니다.</div>';
  var pax=num(d.pax)||paxFallback||1;var body='',tot=0,cur=null,sub=0;
  var subRow=function(g,st){return '<tr class="subtot"><td colspan="13" class="tl">'+esc2(g||'기타')+' 소계 (1인)</td><td class="rt p">'+fmtN(st)+'</td></tr>';};
  d.rows.forEach(function(r){
    var g=r.gu||'';
    if(cur!==null&&g!==cur){body+=subRow(cur,sub);sub=0;}
    cur=g;
    var v1=num(r.v1),pv=num(r.pv),pm=num(r.pm),ou=num(r.ou);
    var sh=num(r.sh);var need=v1*pax,ord=ou*pv,total=pm*ou+sh;var per=(pax>0&&ord>0)?total/pax*(need/ord):0;tot+=per;sub+=per;
    body+='<tr><td>'+esc2(r.gu)+'</td><td class="tl">'+esc2(r.nm)+(sh>0?' <span class="ship-note">+배송 '+fmtN(sh)+'</span>':'')+'</td><td>'+esc2(r.ba)+'</td>'
      +'<td>'+(r.lk?'<a href="'+esc2(r.lk)+'" target="_blank">링크</a>':'')+'</td>'
      +'<td class="rt">'+esc2(r.v1)+'</td><td>'+esc2(r.u1)+'</td><td class="rt">'+esc2(r.pv)+'</td><td>'+esc2(r.pu)+'</td><td class="rt">'+fmtN(pm)+'</td>'
      +'<td class="rt a">'+fmtN(need)+'</td><td class="rt">'+esc2(r.ou)+'</td><td class="rt a">'+fmtN(ord)+'</td><td class="rt a">'+fmtN(total)+'</td><td class="rt p">'+fmtN(per)+'</td></tr>';
  });
  if(cur!==null)body+=subRow(cur,sub);
  return '<div class="matwrap"><table class="mat"><thead><tr>'
    +'<th rowspan="2">구분</th><th rowspan="2">구성품</th><th rowspan="2">레시피<br>기준량</th><th rowspan="2">구매처</th>'
    +'<th class="grp" colspan="2">1인 기준</th><th class="grp" colspan="3">판매 기준단위</th>'
    +'<th rowspan="2">총<br>필요량</th><th rowspan="2">주문<br>단위</th><th rowspan="2">주문<br>총량</th><th rowspan="2">총금액</th><th rowspan="2">1인<br>금액</th></tr>'
    +'<tr><th>용량</th><th>단위</th><th>용량</th><th>단위</th><th>금액</th></tr></thead><tbody>'+body+'</tbody></table></div>'
    +'<div class="matsum">재료비 합계(1인) <b>'+fmtN(tot)+'</b> 원 · 기준 인원 '+pax+'명</div>';
}
function keepBadge(k){var m={'냉동':'kp-frz','냉장':'kp-cool','실온':'kp-room'};if(!k||!m[k])return '<span class="kp kp-none">—</span>';return '<span class="kp '+m[k]+'">'+k+'</span>';}
function diffCell(diff){if(diff===0)return '<span class="df0">0</span>';if(diff<0)return '<span class="dfsave">▼'+fmtN(-diff)+'</span>';return '<span class="dfover">▲'+fmtN(diff)+'</span>';}
function orderTable(code,od){
  var d=parseQ(code);if(!d||!d.rows||!d.rows.length)return '';
  var pax=num(d.pax)||1;
  var body='',planTot=0,realTot=0,hasReal=false,merge={};
  d.rows.forEach(function(r){
    var v1=num(r.v1),ou=num(r.ou),pv=num(r.pv),pm=num(r.pm),sh=num(r.sh);
    var total=pm*ou+sh;planTot+=total;
    var need=v1*pax;
    var plan=fmtN(ou)+'×'+fmtN(pv)+(r.pu||'');
    var base=(r.nm||'').replace(/\s*\(.*$/,'').trim();merge[base]=(merge[base]||0)+ou;
    var o=(od&&od[r.nm])||{};
    var qty=(o.qty!=null&&o.qty!=='')?num(o.qty):null;
    var real=(o.realamt!=null&&o.realamt!=='')?num(o.realamt):null;
    if(real!=null){realTot+=real;hasReal=true;}
    var stat;
    if(o.stock)stat='<span class="ost ost-stock">🔵 재고</span>';
    else if(qty==null)stat='<span class="ost ost-none">—</span>';
    else if(qty>=need)stat='<span class="ost ost-ok">✅ 충족</span>';
    else{var lack=need-qty;var more=pv>0?Math.ceil(lack/pv):0;stat='<span class="ost ost-short">⚠️ '+fmtN(lack)+(r.pu||'')+' 부족'+(more>0?' ('+more+'개↑)':'')+'</span>';}
    var stockTag=o.stock?' <span class="stk">재고</span>':'';
    body+='<tr>'
      +'<td class="tl"><span class="gu-mini">'+esc2(r.gu)+'</span> '+esc2(r.nm)+(r.lk?' <a href="'+esc2(r.lk)+'" target="_blank">🔗</a>':'')+'</td>'
      +'<td class="rt nw"><b>'+fmtN(need)+(r.pu||'')+'</b> / '+(qty!=null?fmtN(qty)+(r.pu||''):'<span class="mut">—</span>')+'</td>'
      +'<td>'+stat+'</td>'
      +'<td>'+keepBadge(o.keep)+'</td>'
      +'<td class="sm">'+(o.store?esc2(o.store):'—')+'</td>'
      +'<td class="sm nw">'+((o.odate||o.idate)?(esc2(o.odate||'?')+' → '+esc2(o.idate||'?')):'—')+'</td>'
      +'<td class="ck nw">'+(o.ordered?'✅':'⬜')+(o.arrived?'📦':'⬜')+'</td>'
      +'<td class="rt nw"><span class="sm2">'+fmtN(total)+'</span> / '+(real!=null?fmtN(real)+stockTag:(o.stock?'<span class="stk">재고</span>':'<span class="mut">—</span>'))+'</td>'
      +'<td class="rt">'+(real!=null?diffCell(real-total):(o.stock?'<span class="stk">재고</span>':'<span class="mut">—</span>'))+'</td>'
      +'</tr>';
  });
  // 합산 코멘트 (2회 이상 나온 재료)
  var cnt={};d.rows.forEach(function(r){var b=(r.nm||'').replace(/\s*\(.*$/,'').trim();cnt[b]=(cnt[b]||0)+1;});
  var notes=[];for(var b in cnt){if(cnt[b]>1)notes.push(b+' 총 '+fmtN(merge[b])+'개 발주');}
  var mergeNote=notes.length?'<div class="ordnote">※ 합산 발주: '+notes.join(' · ')+' (같은 상품은 한 번에 주문)</div>':'';
  // 손익 비교
  var supply=num(d.price);if((d.vat||'')==='incl')supply=Math.round(supply/1.1);
  var pack=num(d.pack),ship2=num(d.ship),teach=num(d.teach);
  var useTot=0;
  d.rows.forEach(function(r){var v1=num(r.v1),pv2=num(r.pv),pm2=num(r.pm),ou2=num(r.ou),sh2=num(r.sh);var need=v1*pax,ord=ou2*pv2,tt=pm2*ou2+sh2;if(pax>0&&ord>0)useTot+=tt/pax*(need/ord);});
  var planMat=Math.round(useTot);var planNet=supply-planMat-pack-ship2-teach;
  var realPer=hasReal?Math.round(realTot/pax):null;
  var realNet=hasReal?(supply-realPer-pack-ship2-teach):null;
  var cmp='<div class="cmpwrap"><table class="cmp"><thead><tr><th>구분</th><th>계획(원가)</th><th>실제(발주)</th></tr></thead><tbody>'
    +'<tr><td class="tl">발주 총액 <span class="cmpsub">전체·구매기준</span></td><td class="rt">'+fmtN(planTot)+'</td><td class="rt">'+(hasReal?fmtN(realTot):'<span class="pend">발주 후 입력</span>')+'</td></tr>'
    +'<tr><td class="tl">재료비(1인) <span class="cmpsub">계획=사용 / 실제=지출</span></td><td class="rt">'+fmtN(planMat)+'</td><td class="rt">'+(hasReal?fmtN(realPer):'—')+'</td></tr>'
    +'<tr class="cmp-net"><td class="tl">순이익(1인)</td><td class="rt">'+fmtN(planNet)+'</td><td class="rt">'+(hasReal?fmtN(realNet):'—')+'</td></tr>'
    +'</tbody></table></div>';
  return '<div class="ordsec"><div class="ordh">📦 발주 · 입고 <span class="s">필요량 대비 주문량 · 부족분 · 실제 비용</span></div>'
    +'<div class="ordwrap"><table class="ord"><thead><tr>'
    +'<th>재료</th><th>필요 / 주문</th><th>상태</th><th>보관</th><th>발주처</th><th>발주 → 입고</th><th>주문/입고</th><th>계획 / 실제</th><th>차액</th>'
    +'</tr></thead><tbody>'+body+'</tbody></table></div>'+mergeNote
    +cmp
    +'<div class="ordnote2">필요량 = 1인 사용량 × 인원 · 주문량이 필요량 이상이면 ✅충족, 미만이면 ⚠️부족(더 살 개수 표시). 발주 정보는 실제 발주하며 채팅으로 알려주시면 채워집니다.</div></div>';
}
function openDetail(cid){
  var d=D[cid];if(!d)return;CUR=cid;
  document.getElementById('mOrderer').textContent=(d.orderer||'')+' · '+(d.kind||'');
  document.getElementById('mName').textContent=d.name||'';
  var c=d.content||{};
  // ① 정보
  document.getElementById('d1').innerHTML=
    '<div class="d-photo"><img src="class-images/'+CUR+'.jpg" alt="" onerror="this.parentElement.style.display=&quot;none&quot;"></div>'
    +row('수업명',d.name)+row('강의명',c.강의명)+row('발주처',d.orderer)+row('구분',d.kind)
    +row('수업일',d.date)+row('시각',d.time)+row('장소/배송',d.place)
    +row('대상',c.대상)+row('선발방식',c.선발방식)+row('소요시간',c.소요시간)+row('줌 링크',c.줌링크)+row('상태',d.state);
  // ② 과정개요서
  document.getElementById('d2').innerHTML=
    '<button class="dcopy" onclick="copyOverview()">📋 과정개요서 전체 복사</button>'
    +blk('강의명',c.강의명)+blk('소개문',c.소개문)+blk('강의내용',c.강의내용)
    +blk('키트 구성품',c.키트구성품)+blk('개별 준비물',c.개별준비물)
    +(d.kind==='온라인'?(blk('배송 안내',c.배송안내)+blk('참여 안내',c.참여안내)):(blk('장소',c.장소상세)+blk('동반인 규정',c.동반인규정)))
    +blk('강사 이력',c.강사이력)+blk('해시태그',c.해시태그)
    +'<div class="dnote">수강신청 확인사항·GSEEK 참여안내 등 고정 문구는 채널로 낼 때 자동으로 붙습니다.</div>';
  // ③ 레시피 (조리법 · 레시피 카드용)
  document.getElementById('d3').innerHTML=
    (c.레시피?'<button class="dcopy" onclick="copyRecipe()">📋 레시피 전체 복사</button>':'')
    +blk('조리법',c.레시피)
    +'<div class="dnote">레시피 카드 제작용 조리법입니다. 들어가는 재료·단가는 옆 [구성품(원가)] 탭에서 봅니다.</div>';
  // ④ 구성품(원가) — 원가표만
  document.getElementById('d4').innerHTML=matTable(c.견적코드,d.applied||d.mn);
  // 발주·입고 (독립 탭)
  document.getElementById('d8').innerHTML=orderTable(c.견적코드,d.order);
  // ④ 가격·정산 (단가 + 부가세금액 + 강사료 분리 + 공급가 기준 손익)
  var vatAmt=null,supply=null;
  if(d.price!=null){if(d.vat==='포함'){supply=d.price/1.1;vatAmt=d.price-supply;}else if(d.vat==='별도'){supply=d.price;vatAmt=d.price*0.1;}else{supply=d.price;}}
  var costEx=(d.mat||0)+(d.pack||0)+(d.ship||0);   // 원가(강사 제외)
  var ap4=d.applied||d.mn||0;
  var recv4=(supply!=null&&ap4)?supply*ap4:null;    // 받을 것 = 공급가 × 인원
  var take4=(recv4!=null&&ap4)?(recv4-costEx*ap4-(d.settle||0)):null;
  document.getElementById('d5').innerHTML='<div class="flow">'
    +'<div class="flow-sec"><div class="flow-h">단가 (인당)</div>'
    +'<div class="flow-r"><span>판매가</span><span class="fv">'+won(d.price)+'</span></div>'
    +(vatAmt!=null?('<div class="flow-r"><span>부가세 ('+(d.vat||'')+')</span><span class="fv">'+won(vatAmt)+'</span></div>'
      +'<div class="flow-r"><span>'+(d.vat==='포함'?'공급가 (부가세 제외)':'합계 (부가세 포함)')+'</span><span class="fv">'+won(d.vat==='포함'?supply:(d.price+vatAmt))+'</span></div>'):'')
    +'</div>'
    +'<div class="flow-sec"><div class="flow-h">원가 (인당) · 쿠킹박스</div>'
    +'<div class="flow-r"><span>재료비</span><span class="fv">'+won(d.mat)+'</span></div>'
    +(d.kind==='온라인'?('<div class="flow-r"><span>패킹비</span><span class="fv">'+won(d.pack)+'</span></div>'
      +'<div class="flow-r"><span>배송비</span><span class="fv">'+won(d.ship)+'</span></div>'):'')
    +'<div class="flow-r tot"><span>원가 합</span><span class="fv">'+won(costEx)+'</span></div></div>'
    +((d.teach)?('<div class="flow-sec"><div class="flow-h">강사료 (인당) · 별도</div>'
      +'<div class="flow-r"><span>강사료 (선생님 인건비)</span><span class="fv">'+won(d.teach)+'</span></div></div>'):'')
    +'<div class="flow-sec"><div class="flow-r tot pos"><span>수익 (인당)</span><span class="fv">'+won(d.profit)+'</span></div>'
    +'<div class="flow-r"><span>수익률</span><span class="fv">'+pct(d.rate)+'</span></div></div>'
    +(take4!=null?('<div class="flow-sec"><div class="flow-h">이 클래스 전체 손익 (신청 '+ap4+'명 기준)</div>'
      +'<div class="flow-r"><span>받을 것 ('+(d.vat==='포함'?'공급가':'판매가')+' × '+ap4+'명)</span><span class="fv">'+won(recv4)+'</span></div>'
      +'<div class="flow-r neg"><span>− 원가 ('+won(costEx)+' × '+ap4+'명)</span><span class="fv">-'+fmtN(costEx*ap4)+'원</span></div>'
      +'<div class="flow-r neg"><span>− 선생님 정산액 (강사료)</span><span class="fv">-'+fmtN(d.settle||0)+'원</span></div>'
      +'<div class="flow-r tot pos"><span>순이익</span><span class="fv">'+won(take4)+'</span></div></div>'):'')
    +'</div>'
    +'<div class="dnote">부가세 제외(공급가) 기준 손익입니다. 강사료는 원가와 분리해 별도 차감합니다. 정산구조: '+(d.struct==='A'?'A · 주최수금':(d.struct==='B'?'B · 우리수금':'-'))+'</div>';
  // ⑦ 모집 현황
  var mtype=d.mtype==='A'?'A · 사전확정':(d.mtype==='B'?'B · 모객':'-');
  var openLabel={'사전확정':'사전확정','신청중':'신청중','확정':'확정','미달':'미달'}[d.openst]||d.openst;
  var h7=row('모집유형',mtype);
  if(d.mtype==='A'){h7+=row('확정 인원',(d.applied||d.mn)+'명 (사전확정)');}
  else{h7+=row('신청 현황','신청 '+(d.applied||0)+'명 · 최소 '+d.mn+' · 최대 '+d.mx)+row('개폐강',openLabel);}
  if(d.signupdue&&d.mtype!=='A'){h7+='<div class="dsub">개최 문자 알림</div><div class="drow"><span class="k">📩 발송 예정</span><span class="v mono">'+esc(d.signupdue)+' 09:00 · 선생님·수강생에게 개최 여부 통보</span></div>';}
  h7+=row('확정기준일',d.confirmday)+row('환불마감일',d.refundday)+row('환불규칙',d.refundsrc);
  if(d.kind==='온라인'){h7+='<div class="dsub">배송 현황</div><div class="dempty">배송 집계는 완전판 마스터(수강생 시트)에서 관리합니다.</div>';}
  h7+='<div class="dnote">수강생 명단(연락처·주소·송장)은 개인정보라 표시하지 않습니다. 완전판 마스터(채팅창)에서 확인하세요.</div>';
  document.getElementById('d7').innerHTML=h7;
  // ⑥ 정산 (선생님 지급에 집중 · 우리 몫은 ④에)
  var applied=d.applied||d.mn||0;
  var kitTotal=(d.price!=null&&applied)?d.price*applied:null;
  var settle=d.settle;var payReal=null,payNote='';
  if(settle!=null&&settle!==''){
    if(d.ttype==='개인'){var tax=Math.round(settle*0.033);payReal=settle-tax;payNote='정산액 '+won(settle)+' − 원천세 3.3% '+won(tax);}
    else if(d.ttype==='사업자'){payReal=settle;payNote='세금계산서 발행 · 원천징수 없음';}
    else payReal=settle;
  }
  var p6='';
  if(payReal!=null)p6+='<div class="bigpay"><span class="bk">👩‍🍳 선생님께 입금할 금액</span><span class="bv">'+won(payReal)+'</span></div>'+(payNote?'<div class="dnote" style="margin-top:-6px;margin-bottom:12px">'+payNote+'</div>':'');
  p6+='<div class="flow">'
    +'<div class="flow-sec"><div class="flow-h flow-h2"><span>우리가 받을 것 (발주처 → 우리)</span>'+payBadge(d.paidin,'입금')+'</div>'
    +'<div class="flow-r"><span>판매가 × 신청 '+(applied||0)+'명</span><span class="fv">'+won(kitTotal)+'</span></div>'
    +drowFlow('세금계산서 발행처',d.taxto)+drowFlow('지급주체',d.payer)+drowFlow('지급시기',d.paywhen)+drowFlow('정산예정일',d.paydate)+'</div>'
    +'<div class="flow-sec"><div class="flow-h flow-h2"><span>우리가 줄 것 (우리 → 선생님)</span>'+payBadge(d.paidout,'지급')+'</div>'
    +'<div class="flow-r"><span>정산방식</span><span class="fv">'+(d.feeway==='인당1만'?'인당 1만원':(d.feeway==='수수료%'?('수수료 '+pct(d.feerate)):'-'))+'</span></div>'
    +'<div class="flow-r"><span>선생님 정산액</span><span class="fv">'+won(settle)+'</span></div>'
    +'<div class="flow-r"><span>선생님 유형</span><span class="fv">'+(d.ttype==='개인'?'개인(원천세 3.3%)':(d.ttype==='사업자'?'사업자(계산서)':'-'))+'</span></div>'
    +'<div class="flow-r neg tot"><span>실지급액</span><span class="fv">'+won(payReal)+'</span></div></div>'
    +'</div>'
    +'<div class="dnote">💳 선생님 계좌·은행은 개인정보라 표시하지 않습니다. <b>완전판 마스터(채팅창)</b>에서 확인하세요.</div>';
  document.getElementById('d6').innerHTML=p6;
  document.querySelectorAll('.mtab').forEach(function(t,i){t.classList.toggle('on',i===0);});
  document.querySelectorAll('.mpanel').forEach(function(p,i){p.classList.toggle('on',i===0);});
  var mp=document.getElementById('mPack');
  if(mp){ if(d.packfile){mp.href=d.packfile;mp.style.display='';} else {mp.style.display='none';} }
  var edit=document.getElementById('mEdit');
  edit.onclick=function(ev){
    var sentence=(d.name||'')+'('+cid+') 계산기 불러오기 코드 만들어줘';
    if(navigator.clipboard)navigator.clipboard.writeText(sentence).catch(function(){});
    edit.textContent='✓ 문장 복사됨 → 대화창에 붙여넣기';
    setTimeout(function(){edit.textContent='✏ 클래스 수정';},2200);
  };
  document.getElementById('modal').classList.add('on');
  document.body.style.overflow='hidden';
}
function copyOverview(){
  var d=D[CUR];if(!d)return;var c=d.content||{};var NL=String.fromCharCode(10);
  var L=['[과정개요서] '+(d.name||'')];
  var add=function(k,v){if(v)L.push(NL+'◆ '+k+NL+v);};
  add('강의명',c.강의명);add('소개문',c.소개문);add('강의내용',c.강의내용);
  add('키트 구성품',c.키트구성품);add('개별 준비물',c.개별준비물);
  if(d.kind==='온라인'){add('배송 안내',c.배송안내);add('참여 안내',c.참여안내);}
  else{add('장소',c.장소상세);add('동반인 규정',c.동반인규정);}
  add('강사 이력',c.강사이력);add('해시태그',c.해시태그);
  var txt=L.join(NL);
  var done=function(){var btn=document.querySelector('#d2 .dcopy');if(btn){var o=btn.textContent;btn.textContent='✓ 복사됨';setTimeout(function(){btn.textContent=o;},1500);}};
  if(navigator.clipboard)navigator.clipboard.writeText(txt).then(done,function(){var ta=document.createElement('textarea');ta.value=txt;document.body.appendChild(ta);ta.select();document.execCommand('copy');ta.remove();done();});
}
function copyRecipe(){var d=D[CUR];if(!d)return;var c=d.content||{};if(!c.레시피)return;var txt='[레시피] '+(d.name||'')+String.fromCharCode(10)+String.fromCharCode(10)+c.레시피;
  var done=function(){var btn=document.querySelector('#d3 .dcopy');if(btn){var o=btn.textContent;btn.textContent='✓ 복사됨';setTimeout(function(){btn.textContent=o;},1500);}};
  if(navigator.clipboard)navigator.clipboard.writeText(txt).then(done,function(){var ta=document.createElement('textarea');ta.value=txt;document.body.appendChild(ta);ta.select();document.execCommand('copy');ta.remove();done();});
}
function closeDetail(){document.getElementById('modal').classList.remove('on');document.body.style.overflow='';}
document.querySelectorAll('.mtab').forEach(function(t){t.addEventListener('click',function(){
  document.querySelectorAll('.mtab').forEach(function(x){x.classList.remove('on')});
  document.querySelectorAll('.mpanel').forEach(function(p){p.classList.remove('on')});
  t.classList.add('on');document.getElementById(t.dataset.t).classList.add('on');
});});
document.addEventListener('keydown',function(e){if(e.key==='Escape')closeDetail();});
"""

# 내보내기용(개인정보 제외 회차 요약)
export=[]
for s in sessions:
    export.append({"회차ID":s["cid"],"수업명":s["name"],"발주처":s["orderer"],"구분":s["kind"],
     "수업일":s["date"].strftime("%Y-%m-%d") if s["date"] else "","모집유형":s["mtype"],
     "최소":s["mn"],"최대":s["mx"],"신청":s["applied"],"개폐강":s["openst"],
     "판매가":s["price"] or "","재료비":s["mat"] or "","수익":s["profit"] or "","수익률":s["rate"] or "",
     "선생님":s["teacher"],"상태":s["state"]})
EXPORT=json.dumps(export,ensure_ascii=False)

def dv2(x):
    if isinstance(x,(datetime.date,datetime.datetime)): return fmt_date(to_date(x))
    return "" if x is None else str(x)
detail={}
for s in sessions:
    detail[s["cid"]]=dict(name=s["name"],orderer=s["orderer"],kind=s["kind"],mtype=s["mtype"],
        date=ymd_full(s["date"]),time=s["time"] or "",place=s["place"] or "",state=s["state"],
        mn=s["mn"],mx=s["mx"],applied=s["applied"],openst=s["openst"],
        signupdue=(md_short(s["signup_deadline"]) if s["signup_deadline"] else ""),
        confirmday=fmt_date(s["confirmday"]),refundday=fmt_date(s["refundday"]),refundsrc=s["refundsrc"] or "",
        price=s["price"],vat=s["vat"],struct=s["struct"],mat=s["mat"],pack=s["pack"],ship=s["ship"],teach=s["teach"],
        profit=s["profit"],rate=s["rate"],teacher=s["teacher"],ttype=s["ttype"],feeway=s["feeway"],
        feerate=s["feerate"],settle=s["settle"],taxto=s["taxto"] or "",payer=s["payer"] or "",
        paywhen=s["paywhen"] or "",paydate=dv2(s["paydate"]),
        paidin=s["paidin"],paidout=s["paidout"],
        packfile=s.get("packfile",""),
        order=order_map.get(s["cid"],{}),
        content={k:(v or "") for k,v in (s["content"] or {}).items()})
DETAIL=json.dumps(detail,ensure_ascii=False)

# ── 이번 주 패킹 배너 (진행중 + 패킹파일 있음 + 수업일 D-7 이내) ──
_pb=[]
for s in sessions:
    if not s.get("packfile"): continue
    if s["state"] not in ("기획중","모집중","확정"): continue
    if not s["date"]: continue
    dleft=(s["date"]-TODAY).days
    if dleft<0 or dleft>7: continue
    _n=s["applied"] or s["mn"] or 0
    _dl=("오늘" if dleft==0 else f"D-{dleft}")
    _pb.append(f'<div class="packbar"><span class="pb-ic">📦</span><div class="pb-l">'
        f'<span class="pb-k">이번 주 패킹 · {_dl}</span>'
        f'<div class="pb-n">{esc(s["name"])}</div>'
        f'<div class="pb-d">{ymd_full(s["date"])} {esc(s["time"] or "")} · {esc(s["orderer"])} · {_n}명</div>'
        f'</div><a class="pb-b" href="{esc(s["packfile"])}">패킹 운영 열기 →</a></div>')
PACKBAR="".join(_pb)

doc=f'''<!DOCTYPE html><html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="version" content="{VERSION}"><meta name="updated" content="{UPDATED}">
<title>클래스 홈</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Hahmlet:wght@500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>{CSS}</style></head><body><div class="wrap">
<div class="topbar"><div class="tb-left"><a class="pill" href="index.html">← 공공의주방 홈</a><a class="pill" href="{CHAT_URL}" target="_blank" rel="noopener">💬 작업 대화</a></div>
<div class="tb-right"><span class="crumb">AI 워크스페이스 › 클래스</span><span class="tb-sep">|</span><span class="updated" id="upd"></span></div></div>
<header class="hero"><h1>클래스 홈</h1><p class="lead">온라인·오프라인 수업 일정과 정산을 관리합니다.</p></header>
<div class="summary">
<div class="s-item"><div class="s-num">{len(upcoming)}</div><div class="s-lab">다가오는 수업</div></div>
<div class="s-item s-recruit"><div class="s-num">{n_recruit}</div><div class="s-lab">모집 중</div></div>
<div class="s-item s-open"><div class="s-num">{n_open}</div><div class="s-lab">확정</div></div>
<div class="s-item"><div class="s-num">{n_stu}</div><div class="s-lab">전체 수강생</div></div>
</div>
{PACKBAR}

<section class="layer"><div class="layer-h"><h2>진행 중인 클래스</h2><span class="ls">기획·모집·확정·정산</span><span class="bar"></span></div>
{layer1}
</section>

<section class="layer"><div class="layer-h"><h2>주요 기능</h2><span class="ls">목록·리포트</span><span class="bar"></span></div>
<div class="feats">{feat_cards}</div></section>

<section class="layer"><div class="layer-h"><h2>도구</h2><span class="ls">계산·가이드</span><span class="bar"></span></div>
<div class="tools">{tool_cards}</div></section>

<section class="layer"><div class="layer-h"><h2>📦 원본 파일</h2><span class="ls">백업·동기화</span><span class="bar"></span></div>
<div class="backup-row">
<button class="bk-btn" onclick="saveHTML()">📄 이 화면 HTML 받기</button>
<button class="bk-btn" onclick="saveXLSX()">📊 엑셀 마스터 파일 받기 <span class="bk-tag">개인정보 뺀 공개용</span></button>
</div>
<p class="backup-note">위 엑셀은 개인정보(연락처·주소·수강생 이름)를 뺀 <b>회차·인원 정보(공개용)</b>입니다. → <code>class-master-open_날짜.xlsx</code><br>※ <b>완전판</b>(연락처·주소 포함 원본)은 <a href="{CHAT_URL}" target="_blank" rel="noopener">채팅창</a>에서 받으세요. → <code>class-master-all-v2.xlsx</code></p></section>

<div class="ops-note">※ <b>홈은 보기·이동만 합니다.</b><br>· 데이터를 고칠 땐 → 마스터(엑셀/대화창)에서<br>· 볼 땐 → 이 홈에서<br>· 모든 파일은 같은 폴더(github)에 두어야 링크가 작동합니다.</div>
<div class="foot">클래스 홈 · {VERSION} · 마스터에서 자동 생성 (gen_class.py)</div>
</div>

<div class="modal" id="modal"><div class="modal-bg" onclick="closeDetail()"></div>
<div class="modal-panel">
<div class="modal-head"><div><span class="m-orderer" id="mOrderer"></span><h2 id="mName"></h2></div>
<div class="m-actions"><a class="m-pack" id="mPack" href="#" style="display:none">📦 패킹 운영</a><a class="m-edit" id="mEdit" href="{CHAT_URL}" target="_blank" rel="noopener">✏ 클래스 수정</a><button class="m-close" onclick="closeDetail()">✕</button></div></div>
<div class="modal-tabs">
<button class="mtab on" data-t="d1">정보</button><button class="mtab" data-t="d2">과정개요서</button>
<button class="mtab" data-t="d3">레시피</button><button class="mtab" data-t="d4">구성품(원가)</button>
<button class="mtab" data-t="d8">발주·입고</button><button class="mtab" data-t="d5">수익 구조</button>
<button class="mtab" data-t="d6">선생님 정산</button><button class="mtab" data-t="d7">모집 현황</button></div>
<div class="modal-body">
<div class="mpanel on" id="d1"></div><div class="mpanel" id="d2"></div><div class="mpanel" id="d3"></div>
<div class="mpanel" id="d4"></div><div class="mpanel" id="d8"></div><div class="mpanel" id="d5"></div><div class="mpanel" id="d6"></div><div class="mpanel" id="d7"></div>
</div></div></div>

<script>window.CLASS_ROWS={EXPORT};window.CLASS_DETAIL={DETAIL};</script>
<script>{JS}</script>
</body></html>'''

open("class.html","w",encoding="utf-8").write(doc)
print(f"class.html {VERSION} — 진행중 온{len([s for s in sessions if s['kind']=='온라인' and s['state'] in('기획중','모집중','확정')])} / 전체 {len(sessions)}건")

# ── 전체 목록 (필터·검색) 별도 도구 ──
listdata=[]
for s in sessions:
    _df=(ymd_full(s["date"])+((" "+str(s["time"])) if s["time"] else "")) if s["date"] else "미정"
    listdata.append(dict(cid=s["cid"],name=s["name"],orderer=s["orderer"],kind=s["kind"],
        date=s["date"].strftime("%Y-%m-%d") if s["date"] else "",datef=_df,
        mtype=s["mtype"],mn=s["mn"],mx=s["mx"],applied=s["applied"],openst=s["openst"],
        state=s["state"],teacher=s["teacher"],profit=s["profit"],
        paidin=s["paidin"],paidout=s["paidout"],settled=is_settled(s)))
LISTDATA=json.dumps(listdata,ensure_ascii=False)
orderers=sorted(set(s["orderer"] for s in sessions if s["orderer"] and s["orderer"]!="-"))
ORDS=json.dumps(orderers,ensure_ascii=False)

LISTTMPL = r"""<!DOCTYPE html><html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="version" content="__VER__"><meta name="updated" content="__UPD__">
<title>전체 클래스 목록</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css">
<link href="https://fonts.googleapis.com/css2?family=Hahmlet:wght@500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{--bg:#FAF5EC;--ink:#241F1B;--sub:#6B6256;--line:#E7DCC8;--card:#FFFDF9;--point:#B4A032;--online:#2C7BB6;--offline:#C2570E;--open:#2E7D32;--short:#E2611B;--wait:#9A8F7C;--coral:#FF5019;}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--ink);font-family:Pretendard,-apple-system,BlinkMacSystemFont,"Apple SD Gothic Neo","Malgun Gothic",sans-serif;line-height:1.55;padding-bottom:60px}
.wrap{max-width:1160px;margin:0 auto;padding:0 20px}
.topbar{display:flex;justify-content:space-between;align-items:center;padding:18px 0;border-bottom:1px solid var(--line);gap:12px;flex-wrap:wrap}
.tb-left{display:flex;gap:8px}
.pill{display:inline-flex;align-items:center;gap:4px;font-size:13px;text-decoration:none;color:var(--ink);background:var(--card);border:1px solid var(--line);border-radius:20px;padding:5px 13px}
.pill:hover{border-color:var(--point);color:var(--point)}
.tb-right{font-size:12px;color:var(--wait);text-align:right}.crumb{color:var(--wait)}.tb-sep{color:#CFC5B2}.updated{font-family:"DM Mono",monospace;font-size:12px;color:var(--wait)}
.hero{padding:28px 0 6px}
.hero h1{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-weight:700;font-size:38px;letter-spacing:-.02em}
.hero .lead{margin-top:7px;color:var(--sub);font-size:15px}
.filters{display:flex;flex-wrap:wrap;gap:10px;margin:22px 0 8px;align-items:center}
.filters input,.filters select{font-family:inherit;font-size:13.5px;padding:9px 12px;border:1px solid var(--line);border-radius:9px;background:#fff;color:var(--ink)}
.filters input{flex:1;min-width:180px}
.filters input:focus,.filters select:focus{outline:none;border-color:var(--point)}
.seg{display:inline-flex;border:1px solid var(--line);border-radius:9px;overflow:hidden;background:#fff}
.seg button{font-family:inherit;font-size:13px;padding:9px 14px;border:none;background:none;cursor:pointer;color:var(--sub)}
.seg button.on{background:var(--point);color:#fff;font-weight:600}
.count{font-size:12.5px;color:var(--wait);margin:6px 2px 0}
.listwrap{margin-top:12px;overflow-x:auto;border:1px solid var(--line);border-radius:12px;background:var(--card)}
table.lt{width:100%;border-collapse:collapse;font-size:13px;min-width:720px}
table.lt>thead>tr>th,table.lt>tbody>tr.mainrow>td{padding:11px 13px;border-bottom:1px solid var(--line);text-align:left}
table.lt>thead>tr>th{background:#F4EFE3;font-size:12px;color:var(--sub);font-weight:700;cursor:pointer;user-select:none;white-space:nowrap}
table.lt>thead>tr>th:hover{color:var(--point)}
tr.mainrow{cursor:pointer}
tr.mainrow:hover{background:#FBF8EF}
tr.mainrow.open{background:#F4EFE3}
.ln{font-weight:600}.rt{text-align:center;font-family:"DM Mono",monospace}
.lt-thumb-wrap{display:inline-block;vertical-align:middle;margin-right:9px}
.lt-thumb{width:34px;height:34px;border-radius:7px;object-fit:cover;vertical-align:middle;background:#EEE9DD}
.lt-ph{align-items:center;justify-content:center;font-size:15px;color:#B8AE98}
.arrow{display:inline-block;transition:.2s;color:var(--point);font-size:11px;margin-right:6px}
tr.mainrow.open .arrow{transform:rotate(90deg)}
.badge{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:10px;font-family:"DM Mono",monospace;white-space:nowrap}
.st-기획중{background:#F3ECD7;color:#A08A4E}.st-모집중{background:#FFE7DF;color:var(--coral)}
.st-확정{background:#FFE7DF;color:var(--coral)}.st-미달{background:#EAE6DE;color:#BFB6A6;text-decoration:line-through}
.st-수업완료{background:#E4EEF6;color:var(--online)}.st-정산완료{background:#EAE6DE;color:#B0A794}.st-보류{background:#EAE6DE;color:#BFB6A6}
.kind{font-size:11px;padding:2px 7px;border-radius:8px}
.paychip{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:10px;font-family:"DM Mono",monospace;white-space:nowrap}
.pc-done{background:#E6F4E6;color:var(--open)}.pc-wait{background:#EEE9DD;color:var(--wait)}
.rt b{font-family:"DM Mono",monospace}
.pax-mini{font-size:9.5px;font-weight:700;padding:1px 6px;border-radius:8px;font-family:"DM Mono",monospace;margin-left:2px}
.pm-fix{background:#EEE9DD;color:var(--wait)}.pm-ing{background:#E8F1F8;color:var(--online)}
.pm-open{background:#E6F4E6;color:var(--open)}.pm-short{background:#F6DDD2;color:#B44A1E}
.legend{margin-top:16px;padding:15px 18px;background:var(--card);border:1px solid var(--line);border-radius:12px}
.lg-t{font-size:12.5px;font-weight:700;color:var(--ink);margin-bottom:9px}
.lg-r{font-size:12px;color:var(--sub);line-height:1.9}
.lg-r b{color:var(--ink)}.lg-x{color:var(--wait)}
.k-온라인,.k-오프라인{background:transparent;color:var(--sub);padding:0;font-weight:500}
.detailrow>td{padding:0;background:#FBF8F0;border-bottom:1px solid var(--line)}
.dwrap{padding:4px 20px 22px}
.dhead{display:flex;justify-content:space-between;align-items:center;padding:14px 0 10px}
.dhead h3{font-family:Hahmlet,"Gowun Batang","Noto Serif KR",serif;font-size:20px;font-weight:700}
.d-edit{font-size:12.5px;text-decoration:none;color:#fff;background:var(--point);border-radius:8px;padding:7px 13px;white-space:nowrap;border:none;cursor:pointer;font-family:inherit}
.d-pack{font-size:12.5px;text-decoration:none;color:var(--point);background:#FFFDF9;border:1px solid var(--point);border-radius:8px;padding:6px 12px;white-space:nowrap;margin-left:auto;margin-right:6px}
.d-pack:hover{background:var(--point);color:#fff}
.dtabs{display:flex;gap:2px;border-bottom:1px solid var(--line);overflow-x:auto}
.dtab{font-family:inherit;font-size:13px;color:var(--sub);background:none;border:none;border-bottom:2px solid transparent;padding:10px 13px;cursor:pointer;white-space:nowrap;margin-bottom:-1px}
.dtab:hover{color:var(--ink)}.dtab.on{color:var(--point);border-bottom-color:var(--point);font-weight:700}
.dbody{padding:16px 2px}
.dpane{display:none}.dpane.on{display:block}
.drow{display:flex;gap:12px;padding:8px 0;border-bottom:1px dashed var(--line);font-size:13.5px}
.drow:last-child{border-bottom:none}.drow .k{color:var(--sub);min-width:110px;flex-shrink:0}.drow .v{color:var(--ink);font-weight:500;white-space:pre-wrap}.drow .v.mono{font-family:"DM Mono",monospace}
.dsub{font-size:12px;color:var(--point);font-weight:700;margin:16px 0 5px}.dsub:first-child{margin-top:0}
.dtext{white-space:pre-wrap;font-size:13.5px;line-height:1.65;color:var(--ink);margin-top:2px}
.dempty{color:var(--wait);font-size:13px;padding:8px 0}
.dnote{margin-top:14px;font-size:12px;color:var(--wait);background:var(--card);border:1px solid var(--line);border-radius:8px;padding:10px 12px}
.dcopy{display:inline-flex;align-items:center;gap:5px;font-family:inherit;font-size:12.5px;color:#fff;background:var(--point);border:none;border-radius:8px;padding:8px 14px;cursor:pointer;margin-bottom:12px}
.dcopy:hover{filter:brightness(1.06)}
table.mat{border-collapse:collapse;width:100%;min-width:820px;font-size:12px;margin-top:4px}
table.mat th,table.mat td{border:1px solid var(--line);padding:5px 7px;text-align:center;white-space:nowrap}
table.mat th{background:#F4EFE3;font-size:11px;color:var(--sub);font-weight:700}
table.mat th.grp{background:#EFE7D4}
table.mat td.tl{text-align:left;font-weight:600}
table.mat td.rt{text-align:right;font-family:"DM Mono",monospace}
table.mat td.a{background:#F3F7FA;color:var(--online)}
table.mat td.p{background:#EAF3FA;font-weight:700}
table.mat tr.subtot td{background:#F0EADB;font-weight:700;color:var(--sub);border-top:1.5px solid #D8CDB4}
table.mat tr.subtot td.tl{text-align:right;padding-right:10px}
table.mat tr.subtot td.p{background:#E4EDD9;color:var(--open)}
table.mat a{color:var(--point);text-decoration:none}
.ship-note{font-size:9.5px;color:var(--offline);font-weight:600}
.matsum{margin-top:10px;text-align:right;font-family:"DM Mono",monospace;font-size:14px}.matsum b{font-size:17px;color:var(--online)}
.ordsec{margin-top:22px;padding-top:18px;border-top:2px solid var(--line)}
.ordh{font-family:Hahmlet,serif;font-size:16px;font-weight:700;margin-bottom:12px}.ordh .s{font-family:Pretendard;font-size:11.5px;color:var(--wait);font-weight:400;margin-left:8px}
.ordwrap{overflow-x:auto}
table.ord{width:100%;border-collapse:collapse;font-size:11px;table-layout:auto}
table.ord th{background:#F1EADB;padding:6px 7px;border:1px solid var(--line);font-weight:600;color:var(--sub);font-size:10.5px}
table.ord td{padding:5px 7px;border:1px solid var(--line);white-space:nowrap}
table.ord th{white-space:nowrap}
.nw{white-space:nowrap}.mut{color:#CFC5B2}
.gu-mini{font-size:8.5px;background:#EFEADD;color:var(--wait);padding:1px 5px;border-radius:6px;font-weight:600}
table.ord .tl{text-align:left}table.ord .rt{text-align:right;font-family:"DM Mono",monospace}
table.ord .ck{text-align:center;font-size:13px}table.ord .sm{font-size:10.5px;color:var(--wait)}
table.ord a{text-decoration:none}
.kp{font-size:10px;font-weight:700;padding:2px 7px;border-radius:9px;font-family:"DM Mono",monospace}
.kp-frz{background:#E4EEF6;color:var(--online)}.kp-cool{background:#E6F4E6;color:var(--open)}.kp-room{background:#EEE9DD;color:var(--wait)}.kp-none{background:transparent;color:#CFC5B2}
.stk{font-size:9.5px;background:#EFEADD;color:var(--point);padding:1px 6px;border-radius:8px;font-weight:700}
.ordnote{font-size:11.5px;color:var(--offline);margin-top:9px;font-weight:600}
.ordnote2{font-size:11px;color:var(--wait);margin-top:10px;line-height:1.6}
.cmpwrap{margin-top:14px}
table.cmp{border-collapse:collapse;font-size:12.5px;min-width:340px}
table.cmp th{background:#F1EADB;padding:8px 14px;border:1px solid var(--line);font-weight:700}
table.cmp td{padding:8px 14px;border:1px solid var(--line)}
table.cmp .tl{text-align:left}table.cmp .rt{text-align:right;font-family:"DM Mono",monospace}
table.cmp .cmp-net td{font-weight:700;background:#FBF8EF}table.cmp .cmp-net .rt{color:var(--open)}
.pend{font-size:11px;color:var(--wait);font-style:italic}
.cmpsub{font-size:9.5px;color:var(--wait);font-weight:400;display:block}
.plan2{display:block;font-size:9px;color:var(--wait);font-weight:400}
.ost{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:9px;white-space:nowrap}
.ost-ok{background:#E6F4E6;color:var(--open)}
.ost-short{background:#FFE7DF;color:var(--coral)}
.ost-stock{background:#E4EEF6;color:var(--online)}
.ost-none{background:transparent;color:#CFC5B2;font-weight:400}
.sm2{color:var(--wait);font-size:10.5px}
.dfsave{color:var(--open);font-weight:700}
.dfover{color:var(--coral);font-weight:700}
.df0{color:var(--wait)}
.matwrap{overflow-x:auto}
.flow{border:1px solid var(--line);border-radius:12px;overflow:hidden;margin-top:4px}
.flow-sec{border-bottom:1px solid var(--line)}.flow-sec:last-child{border-bottom:none}
.flow-h{background:#F4EFE3;padding:8px 14px;font-size:12px;font-weight:700;color:var(--sub)}
.flow-h2{display:flex;justify-content:space-between;align-items:center;gap:8px}
.fh-badge{font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;font-family:"DM Mono",monospace}
.fhb-done{background:#E6F4E6;color:var(--open)}.fhb-wait{background:#EEE9DD;color:var(--wait)}
.d-photo{margin-bottom:14px}
.d-photo img{width:100%;height:150px;object-fit:cover;border-radius:12px;display:block}
.flow-r{display:flex;justify-content:space-between;padding:8px 14px;font-size:13.5px;border-top:1px dashed var(--line)}
.flow-r:first-of-type{border-top:none}.flow-r .fv{font-family:"DM Mono",monospace}
.flow-r.tot{font-weight:700;background:#FBFAF6}.flow-r.tot .fv{font-size:15px}
.flow-r.pos .fv{color:var(--open)}.flow-r.neg .fv{color:var(--short)}
.bigpay{display:flex;justify-content:space-between;align-items:center;background:#EEF6EE;border:1px solid #CBE5CB;border-radius:10px;padding:13px 16px;margin:4px 0 12px}
.bigpay .bk{font-size:13px;color:var(--open);font-weight:600}.bigpay .bv{font-family:"DM Mono",monospace;font-size:20px;font-weight:700;color:var(--open)}
.empty{padding:40px;text-align:center;color:var(--wait)}
.foot{margin-top:24px;padding-top:16px;border-top:1px solid var(--line);font-family:"DM Mono",monospace;font-size:11px;color:var(--wait);text-align:center}
@media(max-width:520px){.hero h1{font-size:29px}.topbar{align-items:flex-start}}
</style></head><body><div class="wrap">
<div class="topbar"><div class="tb-left"><a class="pill" href="class.html">← 클래스 홈</a><a class="pill" href="__CHAT__" target="_blank" rel="noopener">💬 작업 대화</a></div>
<div class="tb-right"><span class="crumb">AI 워크스페이스 › 클래스 › 전체 목록</span><span class="tb-sep"> | </span><span class="updated" id="upd"></span></div></div>
<header class="hero"><h1>전체 클래스 목록</h1><p class="lead">행을 클릭하면 아래로 6탭 상세가 열립니다. 상태·발주처·형태로 걸러 보고 검색하세요.</p></header>
<div class="filters">
<input id="q" placeholder="🔍 수업명·발주처·선생님 검색" oninput="render()">
<select id="fState" onchange="render()"><option value="">상태 전체</option><option value="기획중">기획중</option><option value="모집중">모집중</option><option value="확정">개강확정</option><option value="미달">폐강</option><option value="수업완료">수업완료</option><option value="정산완료">정산완료</option><option value="보류">보류</option></select>
<select id="fOrd" onchange="render()"><option value="">발주처 전체</option></select>
<div class="seg" id="fKind"><button class="on" data-k="">전체</button><button data-k="온라인">온라인</button><button data-k="오프라인">오프라인</button></div>
</div>
<div class="count" id="count"></div>
<div class="listwrap"><table class="lt"><thead><tr>
<th onclick="sortBy('date')">수업일</th><th onclick="sortBy('orderer')">발주처</th><th onclick="sortBy('name')">수업명</th>
<th onclick="sortBy('kind')">형태</th><th onclick="sortBy('state')">수업상태</th><th onclick="sortBy('applied')">인원</th><th onclick="sortBy('teacher')">선생님</th><th onclick="sortBy('paidin')">입금</th><th onclick="sortBy('paidout')">지급</th>
</tr></thead><tbody id="tb"></tbody></table></div>
<div class="legend">
  <div class="lg-t">📖 열 설명</div>
  <div class="lg-r"><b>수업상태</b> — 기획중 → 모집중 → 개강확정 → 수업완료 <span class="lg-x">(모객 클래스가 최소 인원 미달이면 <b>폐강</b>)</span></div>
  <div class="lg-r"><b>인원</b> — <b>사전확정</b>: 경기도지식 등 정해진 인원(모집 안 함) / <b>모객</b>: 신청/최대 인원 + 개폐강(신청중·확정·폐강)</div>
  <div class="lg-r"><b>개폐강 기준</b> — 모객 클래스는 수업일 4일 전(D-4) 자정 마감 → 신청이 최소 인원 이상이면 <b>확정</b>, 미만이면 <b>폐강</b></div>
  <div class="lg-r"><b>입금</b> — 발주처 → 우리 (받을 돈)　·　<b>지급</b> — 우리 → 선생님 (줄 돈)</div>
</div>
<div class="foot">전체 클래스 목록 · __VER__ · 클래스 홈 도구</div>
</div>
<script>
var DATA=__LISTDATA__;var ORDS=__ORDS__;var DETAIL=__DETAIL__;var CHAT="__CHAT__";
var KIND="";var SORT="date";var DIR=1;var OPEN=null;
(function(){var u=document.querySelector('meta[name=updated]'),v=document.querySelector('meta[name=version]');document.getElementById('upd').textContent='업데이트 '+((u&&u.content)||'')+' · '+((v&&v.content)||'');
var sel=document.getElementById('fOrd');ORDS.forEach(function(o){var op=document.createElement('option');op.textContent=o;sel.appendChild(op);});})();
document.getElementById('fKind').addEventListener('click',function(e){if(e.target.tagName!=='BUTTON')return;KIND=e.target.dataset.k;this.querySelectorAll('button').forEach(function(b){b.classList.toggle('on',b.dataset.k===KIND);});render();});
function sortBy(k){if(SORT===k)DIR=-DIR;else{SORT=k;DIR=1;}render();}
function esc(v){return v==null?'':String(v).replace(/</g,'&lt;').replace(/"/g,'&quot;');}
function payChip(v,kind){var done=(v===(kind+'완료'));return '<span class="paychip '+(done?'pc-done':'pc-wait')+'">'+(done?'✓ 완료':'대기')+'</span>';}
function stLabel(s){return {'확정':'개강확정','미달':'폐강'}[s]||s;}
function paxCell(r){
  if(r.mtype==='A'){return '<b>'+(r.mx||r.mn||0)+'</b>명 <span class="pax-mini pm-fix">사전확정</span>';}
  var op={'신청중':'pm-ing','확정':'pm-open','미달':'pm-short'}[r.openst]||'pm-ing';
  return '<b>'+(r.applied||0)+'</b>/'+(r.mx||0)+' <span class="pax-mini '+op+'">'+(r.openst||'')+'</span>';
}
function fmtN(n){if(n==null||n==='')return '';return Math.round(n).toLocaleString();}
function won(n){if(n==null||n==='')return '-';return Math.round(Number(n)).toLocaleString()+'원';}
function pct(n){if(n==null||n==='')return '-';return (Math.round(n*1000)/10)+'%';}
function num(v){v=parseFloat((''+v).replace(/,/g,''));return isNaN(v)?0:v;}
function parseQ(code){if(!code)return null;var m=(''+code).match(/QDATA:([A-Za-z0-9+\/=]+)/);if(!m)return null;try{var d=JSON.parse(decodeURIComponent(escape(atob(m[1]))));return d;}catch(e){return null;}}
function render(){
  var q=(document.getElementById('q').value||'').toLowerCase();
  var st=document.getElementById('fState').value;var od=document.getElementById('fOrd').value;
  var rows=DATA.filter(function(r){
    if(st&&r.state!==st)return false;if(od&&r.orderer!==od)return false;if(KIND&&r.kind!==KIND)return false;
    if(q){var hay=((r.name||'')+(r.orderer||'')+(r.teacher||'')).toLowerCase();if(hay.indexOf(q)<0)return false;}
    return true;});
  rows.sort(function(a,b){var x=a[SORT],y=b[SORT];if(x==null)x='';if(y==null)y='';if(typeof x==='number'||typeof y==='number'){return (Number(x)-Number(y))*DIR;}return String(x).localeCompare(String(y),'ko')*DIR;});
  document.getElementById('count').textContent=rows.length+'건'+(rows.length!==DATA.length?(' (전체 '+DATA.length+')'):'');
  var tb=document.getElementById('tb');
  if(!rows.length){tb.innerHTML='<tr><td colspan="9" class="empty">조건에 맞는 클래스가 없습니다.</td></tr>';return;}
  var html='';
  rows.forEach(function(r){
    html+='<tr class="mainrow'+(OPEN===r.cid?' open':'')+'" onclick="toggle(\''+r.cid+'\')">'
      +'<td><span class="arrow">▶</span>'+(r.datef||'-')+'</td><td>'+esc(r.orderer)+'</td>'
      +'<td class="ln"><span class="lt-thumb-wrap"><img class="lt-thumb" src="class-images/'+esc(r.cid)+'.jpg" alt="" loading="lazy" onerror="this.style.display=&quot;none&quot;;this.nextElementSibling.style.display=&quot;inline-flex&quot;"><span class="lt-thumb lt-ph" style="display:none">🍽</span></span>'+esc(r.name)+'</td>'
      +'<td><span class="kind k-'+esc(r.kind)+'">'+esc(r.kind)+'</span></td>'
      +'<td><span class="badge st-'+esc(r.state)+'">'+stLabel(r.state)+'</span></td>'
      +'<td class="rt">'+paxCell(r)+'</td><td>'+esc(r.teacher)+'</td>'
      +'<td>'+payChip(r.paidin,'입금')+'</td><td>'+payChip(r.paidout,'지급')+'</td></tr>';
    if(OPEN===r.cid){html+='<tr class="detailrow"><td colspan="9">'+buildDetail(r.cid)+'</td></tr>';}
  });
  tb.innerHTML=html;
}
function toggle(cid){OPEN=(OPEN===cid?null:cid);render();
  if(OPEN){setTimeout(function(){var el=document.querySelector('tr.mainrow.open');if(el)el.scrollIntoView({block:'nearest',behavior:'smooth'});},50);}}
function tabTo(cid,i,btn){var wrap=btn.closest('.dwrap');wrap.querySelectorAll('.dtab').forEach(function(t,j){t.classList.toggle('on',j===i);});wrap.querySelectorAll('.dpane').forEach(function(p,j){p.classList.toggle('on',j===i);});}
function drow(k,v,mono){if(v==null||v===''||v==='-')return '';return '<div class="drow"><span class="k">'+k+'</span><span class="v'+(mono?' mono':'')+'">'+esc(v)+'</span></div>';}
function blk(label,text){return '<div class="dsub">'+label+'</div>'+(text?('<div class="dtext">'+esc(text)+'</div>'):'<div class="dempty">미입력</div>');}
function matTable(code,paxFallback){
  var d=parseQ(code);if(!d||!d.rows||!d.rows.length)return '<div class="dempty">재료 정보가 없습니다. 계산기에서 등록하면 채워집니다.</div>';
  var pax=num(d.pax)||paxFallback||1;var body='',tot=0,cur=null,sub=0;
  var subRow=function(g,st){return '<tr class="subtot"><td colspan="13" class="tl">'+esc(g||'기타')+' 소계 (1인)</td><td class="rt p">'+fmtN(st)+'</td></tr>';};
  d.rows.forEach(function(r){
    var g=r.gu||'';
    if(cur!==null&&g!==cur){body+=subRow(cur,sub);sub=0;}
    cur=g;
    var v1=num(r.v1),pv=num(r.pv),pm=num(r.pm),ou=num(r.ou);
    var sh=num(r.sh);var need=v1*pax,ord=ou*pv,total=pm*ou+sh;var per=(pax>0&&ord>0)?total/pax*(need/ord):0;tot+=per;sub+=per;
    body+='<tr><td>'+esc(r.gu)+'</td><td class="tl">'+esc(r.nm)+(sh>0?' <span class="ship-note">+배송 '+fmtN(sh)+'</span>':'')+'</td><td>'+esc(r.ba)+'</td>'
      +'<td>'+(r.lk?'<a href="'+esc(r.lk)+'" target="_blank">링크</a>':'')+'</td>'
      +'<td class="rt">'+esc(r.v1)+'</td><td>'+esc(r.u1)+'</td><td class="rt">'+esc(r.pv)+'</td><td>'+esc(r.pu)+'</td><td class="rt">'+fmtN(pm)+'</td>'
      +'<td class="rt a">'+fmtN(need)+'</td><td class="rt">'+esc(r.ou)+'</td><td class="rt a">'+fmtN(ord)+'</td><td class="rt a">'+fmtN(total)+'</td><td class="rt p">'+fmtN(per)+'</td></tr>';
  });
  if(cur!==null)body+=subRow(cur,sub);
  return '<div class="matwrap"><table class="mat"><thead><tr>'
    +'<th rowspan="2">구분</th><th rowspan="2">구성품</th><th rowspan="2">레시피<br>기준량</th><th rowspan="2">구매처</th>'
    +'<th class="grp" colspan="2">1인 기준</th><th class="grp" colspan="3">판매 기준단위</th>'
    +'<th rowspan="2">총<br>필요량</th><th rowspan="2">주문<br>단위</th><th rowspan="2">주문<br>총량</th><th rowspan="2">총금액</th><th rowspan="2">1인<br>금액</th></tr>'
    +'<tr><th>용량</th><th>단위</th><th>용량</th><th>단위</th><th>금액</th></tr></thead><tbody>'+body+'</tbody></table></div>'
    +'<div class="matsum">재료비 합계(1인) <b>'+fmtN(tot)+'</b> 원 · 기준 인원 '+pax+'명</div>';
}
function keepBadge(k){var m={'냉동':'kp-frz','냉장':'kp-cool','실온':'kp-room'};if(!k||!m[k])return '<span class="kp kp-none">—</span>';return '<span class="kp '+m[k]+'">'+k+'</span>';}
function diffCell(diff){if(diff===0)return '<span class="df0">0</span>';if(diff<0)return '<span class="dfsave">▼'+fmtN(-diff)+'</span>';return '<span class="dfover">▲'+fmtN(diff)+'</span>';}
function orderTable(code,od){
  var d=parseQ(code);if(!d||!d.rows||!d.rows.length)return '';
  var pax=num(d.pax)||1;
  var body='',planTot=0,realTot=0,hasReal=false,merge={};
  d.rows.forEach(function(r){
    var v1=num(r.v1),ou=num(r.ou),pv=num(r.pv),pm=num(r.pm),sh=num(r.sh);
    var total=pm*ou+sh;planTot+=total;
    var need=v1*pax;
    var plan=fmtN(ou)+'×'+fmtN(pv)+(r.pu||'');
    var base=(r.nm||'').replace(/\s*\(.*$/,'').trim();merge[base]=(merge[base]||0)+ou;
    var o=(od&&od[r.nm])||{};
    var qty=(o.qty!=null&&o.qty!=='')?num(o.qty):null;
    var real=(o.realamt!=null&&o.realamt!=='')?num(o.realamt):null;
    if(real!=null){realTot+=real;hasReal=true;}
    var stat;
    if(o.stock)stat='<span class="ost ost-stock">🔵 재고</span>';
    else if(qty==null)stat='<span class="ost ost-none">—</span>';
    else if(qty>=need)stat='<span class="ost ost-ok">✅ 충족</span>';
    else{var lack=need-qty;var more=pv>0?Math.ceil(lack/pv):0;stat='<span class="ost ost-short">⚠️ '+fmtN(lack)+(r.pu||'')+' 부족'+(more>0?' ('+more+'개↑)':'')+'</span>';}
    var stockTag=o.stock?' <span class="stk">재고</span>':'';
    body+='<tr>'
      +'<td class="tl"><span class="gu-mini">'+esc(r.gu)+'</span> '+esc(r.nm)+(r.lk?' <a href="'+esc(r.lk)+'" target="_blank">🔗</a>':'')+'</td>'
      +'<td class="rt nw"><b>'+fmtN(need)+(r.pu||'')+'</b> / '+(qty!=null?fmtN(qty)+(r.pu||''):'<span class="mut">—</span>')+'</td>'
      +'<td>'+stat+'</td>'
      +'<td>'+keepBadge(o.keep)+'</td>'
      +'<td class="sm">'+(o.store?esc(o.store):'—')+'</td>'
      +'<td class="sm nw">'+((o.odate||o.idate)?(esc(o.odate||'?')+' → '+esc(o.idate||'?')):'—')+'</td>'
      +'<td class="ck nw">'+(o.ordered?'✅':'⬜')+(o.arrived?'📦':'⬜')+'</td>'
      +'<td class="rt nw"><span class="sm2">'+fmtN(total)+'</span> / '+(real!=null?fmtN(real)+stockTag:(o.stock?'<span class="stk">재고</span>':'<span class="mut">—</span>'))+'</td>'
      +'<td class="rt">'+(real!=null?diffCell(real-total):(o.stock?'<span class="stk">재고</span>':'<span class="mut">—</span>'))+'</td>'
      +'</tr>';
  });
  var cnt={};d.rows.forEach(function(r){var b=(r.nm||'').replace(/\s*\(.*$/,'').trim();cnt[b]=(cnt[b]||0)+1;});
  var notes=[];for(var b in cnt){if(cnt[b]>1)notes.push(b+' 총 '+fmtN(merge[b])+'개 발주');}
  var mergeNote=notes.length?'<div class="ordnote">※ 합산 발주: '+notes.join(' · ')+' (같은 상품은 한 번에 주문)</div>':'';
  var supply=num(d.price);if((d.vat||'')==='incl')supply=Math.round(supply/1.1);
  var pack=num(d.pack),ship2=num(d.ship),teach=num(d.teach);
  // 계획 원가(1인) = 사용량 기준 = 원가표와 동일
  var useTot=0;
  d.rows.forEach(function(r){var v1=num(r.v1),pv2=num(r.pv),pm2=num(r.pm),ou2=num(r.ou),sh2=num(r.sh);var need=v1*pax,ord=ou2*pv2,tt=pm2*ou2+sh2;if(pax>0&&ord>0)useTot+=tt/pax*(need/ord);});
  var planMat=Math.round(useTot);var planNet=supply-planMat-pack-ship2-teach;
  var realPer=hasReal?Math.round(realTot/pax):null;
  var realNet=hasReal?(supply-realPer-pack-ship2-teach):null;
  var cmp='<div class="cmpwrap"><table class="cmp"><thead><tr><th>구분</th><th>계획(원가)</th><th>실제(발주)</th></tr></thead><tbody>'
    +'<tr><td class="tl">발주 총액 <span class="cmpsub">전체·구매기준</span></td><td class="rt">'+fmtN(planTot)+'</td><td class="rt">'+(hasReal?fmtN(realTot):'<span class="pend">발주 후 입력</span>')+'</td></tr>'
    +'<tr><td class="tl">재료비(1인) <span class="cmpsub">계획=사용 / 실제=지출</span></td><td class="rt">'+fmtN(planMat)+'</td><td class="rt">'+(hasReal?fmtN(realPer):'—')+'</td></tr>'
    +'<tr class="cmp-net"><td class="tl">순이익(1인)</td><td class="rt">'+fmtN(planNet)+'</td><td class="rt">'+(hasReal?fmtN(realNet):'—')+'</td></tr>'
    +'</tbody></table></div>';
  return '<div class="ordsec"><div class="ordh">📦 발주 · 입고 <span class="s">필요량 대비 주문량 · 부족분 · 실제 비용</span></div>'
    +'<div class="ordwrap"><table class="ord"><thead><tr>'
    +'<th>재료</th><th>필요 / 주문</th><th>상태</th><th>보관</th><th>발주처</th><th>발주 → 입고</th><th>주문/입고</th><th>계획 / 실제</th><th>차액</th>'
    +'</tr></thead><tbody>'+body+'</tbody></table></div>'+mergeNote+cmp
    +'<div class="ordnote2">필요량 = 1인 사용량 × 인원 · 주문량이 필요량 이상이면 ✅충족, 미만이면 ⚠️부족(더 살 개수 표시). 발주 정보는 실제 발주하며 채팅으로 알려주시면 채워집니다.</div></div>';
}
function buildDetail(cid){
  var d=DETAIL[cid];if(!d)return '<div class="dempty">상세 정보 없음</div>';
  var c=d.content||{};
  // ① 정보
  var photo='<div class="d-photo"><img src="class-images/'+esc(cid)+'.jpg" alt="" '
    +'onerror="this.parentElement.style.display=&quot;none&quot;"></div>';
  var p1=photo+drow('수업명',d.name)+drow('강의명',c['강의명'])+drow('발주처',d.orderer)+drow('구분',d.kind)
    +drow('수업일',d.date)+drow('시각',d.time)+drow('장소/배송',d.place)
    +drow('대상',c['대상'])+drow('선발방식',c['선발방식'])+drow('소요시간',c['소요시간'])+drow('줌 링크',c['줌링크'])+drow('상태',d.state);
  // ② 과정개요서
  var p2='<button class="dcopy" onclick="copyOverview(\''+cid+'\',this)">📋 과정개요서 전체 복사</button>'
    +blk('강의명',c['강의명'])+blk('소개문',c['소개문'])+blk('강의내용',c['강의내용'])
    +blk('키트 구성품',c['키트구성품'])+blk('개별 준비물',c['개별준비물'])
    +(d.kind==='온라인'?(blk('배송 안내',c['배송안내'])+blk('참여 안내',c['참여안내'])):(blk('장소',c['장소상세'])+blk('동반인 규정',c['동반인규정'])))
    +blk('강사 이력',c['강사이력'])+blk('해시태그',c['해시태그'])
    +'<div class="dnote">수강신청 확인사항·GSEEK 참여안내 등 고정 문구는 채널로 낼 때 자동으로 붙습니다.</div>';
  // ③ 레시피 (조리법 · 레시피 카드용)
  var pRec=(c['레시피']?'<button class="dcopy" onclick="copyRecipe(\''+cid+'\',this)">📋 레시피 전체 복사</button>':'')
    +blk('조리법',c['레시피'])
    +'<div class="dnote">레시피 카드 제작용 조리법입니다. 들어가는 재료·단가는 [구성품(원가)] 탭에서 봅니다.</div>';
  // ④ 구성품(원가) (견적 표)
  var p3=matTable(c['견적코드'],d.applied||d.mn);
  var p3o=orderTable(c['견적코드'],d.order);
  // ④ 가격·정산 (단가 + 부가세금액 + 강사료 분리 + 공급가 기준 손익)
  var vatAmt=null,supply=null;
  if(d.price!=null){if(d.vat==='포함'){supply=d.price/1.1;vatAmt=d.price-supply;}else if(d.vat==='별도'){supply=d.price;vatAmt=d.price*0.1;}else{supply=d.price;}}
  var costEx=(d.mat||0)+(d.pack||0)+(d.ship||0);
  var ap4=d.applied||d.mn||0;
  var recv4=(supply!=null&&ap4)?supply*ap4:null;
  var take4=(recv4!=null&&ap4)?(recv4-costEx*ap4-(d.settle||0)):null;
  var p4='<div class="flow">'
    +'<div class="flow-sec"><div class="flow-h">단가 (인당)</div>'
    +'<div class="flow-r"><span>판매가</span><span class="fv">'+won(d.price)+'</span></div>'
    +(vatAmt!=null?('<div class="flow-r"><span>부가세 ('+(d.vat||'')+')</span><span class="fv">'+won(vatAmt)+'</span></div>'
      +'<div class="flow-r"><span>'+(d.vat==='포함'?'공급가 (부가세 제외)':'합계 (부가세 포함)')+'</span><span class="fv">'+won(d.vat==='포함'?supply:(d.price+vatAmt))+'</span></div>'):'')
    +'</div>'
    +'<div class="flow-sec"><div class="flow-h">원가 (인당) · 쿠킹박스</div>'
    +'<div class="flow-r"><span>재료비</span><span class="fv">'+won(d.mat)+'</span></div>'
    +(d.kind==='온라인'?('<div class="flow-r"><span>패킹비</span><span class="fv">'+won(d.pack)+'</span></div>'
      +'<div class="flow-r"><span>배송비</span><span class="fv">'+won(d.ship)+'</span></div>'):'')
    +'<div class="flow-r tot"><span>원가 합</span><span class="fv">'+won(costEx)+'</span></div></div>'
    +((d.teach)?('<div class="flow-sec"><div class="flow-h">강사료 (인당) · 별도</div>'
      +'<div class="flow-r"><span>강사료 (선생님 인건비)</span><span class="fv">'+won(d.teach)+'</span></div></div>'):'')
    +'<div class="flow-sec"><div class="flow-r tot pos"><span>수익 (인당)</span><span class="fv">'+won(d.profit)+'</span></div>'
    +'<div class="flow-r"><span>수익률</span><span class="fv">'+pct(d.rate)+'</span></div></div>'
    +(take4!=null?('<div class="flow-sec"><div class="flow-h">이 클래스 전체 손익 (신청 '+ap4+'명 기준)</div>'
      +'<div class="flow-r"><span>받을 것 ('+(d.vat==='포함'?'공급가':'판매가')+' × '+ap4+'명)</span><span class="fv">'+won(recv4)+'</span></div>'
      +'<div class="flow-r neg"><span>− 원가 ('+won(costEx)+' × '+ap4+'명)</span><span class="fv">-'+fmtN(costEx*ap4)+'원</span></div>'
      +'<div class="flow-r neg"><span>− 선생님 정산액 (강사료)</span><span class="fv">-'+fmtN(d.settle||0)+'원</span></div>'
      +'<div class="flow-r tot pos"><span>순이익</span><span class="fv">'+won(take4)+'</span></div></div>'):'')
    +'</div>'
    +'<div class="dnote">부가세 제외(공급가) 기준 손익입니다. 강사료는 원가와 분리해 별도 차감합니다. 정산구조: '+(d.struct==='A'?'A · 주최수금':(d.struct==='B'?'B · 우리수금':'-'))+'</div>';
  // ⑦ 모집 현황
  var mtype=d.mtype==='A'?'A · 사전확정':(d.mtype==='B'?'B · 모객':'-');
  var p5=drow('모집유형',mtype);
  if(d.mtype==='A'){p5+=drow('확정 인원',(d.applied||d.mn)+'명 (사전확정)');}
  else{p5+=drow('신청 현황','신청 '+(d.applied||0)+'명 · 최소 '+d.mn+' · 최대 '+d.mx)+drow('개폐강',d.openst);}
  if(d.signupdue&&d.mtype!=='A'){p5+='<div class="dsub">개최 문자 알림</div><div class="drow"><span class="k">📩 발송 예정</span><span class="v mono">'+esc(d.signupdue)+' 09:00 · 선생님·수강생에게 개최 여부 통보</span></div>';}
  p5+=drow('확정기준일',d.confirmday)+drow('환불마감일',d.refundday)+drow('환불규칙',d.refundsrc);
  if(d.kind==='온라인'){p5+='<div class="dsub">배송 현황</div><div class="dempty">배송 집계는 완전판 마스터(수강생 시트)에서 관리합니다.</div>';}
  p5+='<div class="dnote">수강생 명단(연락처·주소·송장)은 개인정보라 표시하지 않습니다. 완전판 마스터(채팅창)에서 확인하세요.</div>';
  // ⑥ 정산 (선생님 지급에 집중 · 우리 몫은 ④에)
  var applied=d.applied||d.mn||0;
  var kitTotal=(d.price!=null&&applied)?d.price*applied:null;
  var settle=d.settle;var payReal=null,payNote='';
  if(settle!=null&&settle!==''){
    if(d.ttype==='개인'){var tax=Math.round(settle*0.033);payReal=settle-tax;payNote='정산액 '+won(settle)+' − 원천세 3.3% '+won(tax);}
    else if(d.ttype==='사업자'){payReal=settle;payNote='세금계산서 발행 · 원천징수 없음';}
    else payReal=settle;
  }
  var p6='';
  if(payReal!=null)p6+='<div class="bigpay"><span class="bk">👩‍🍳 선생님께 입금할 금액</span><span class="bv">'+won(payReal)+'</span></div>'+(payNote?'<div class="dnote" style="margin-top:-6px;margin-bottom:12px">'+payNote+'</div>':'');
  p6+='<div class="flow">'
    +'<div class="flow-sec"><div class="flow-h flow-h2"><span>우리가 받을 것 (발주처 → 우리)</span>'+payBadge(d.paidin,'입금')+'</div>'
    +'<div class="flow-r"><span>판매가 × 신청 '+(applied||0)+'명</span><span class="fv">'+won(kitTotal)+'</span></div>'
    +drowFlow('세금계산서 발행처',d.taxto)+drowFlow('지급주체',d.payer)+drowFlow('지급시기',d.paywhen)+drowFlow('정산예정일',d.paydate)+'</div>'
    +'<div class="flow-sec"><div class="flow-h flow-h2"><span>우리가 줄 것 (우리 → 선생님)</span>'+payBadge(d.paidout,'지급')+'</div>'
    +'<div class="flow-r"><span>정산방식</span><span class="fv">'+(d.feeway==='인당1만'?'인당 1만원':(d.feeway==='수수료%'?('수수료 '+pct(d.feerate)):'-'))+'</span></div>'
    +'<div class="flow-r"><span>선생님 정산액</span><span class="fv">'+won(settle)+'</span></div>'
    +'<div class="flow-r"><span>선생님 유형</span><span class="fv">'+(d.ttype==='개인'?'개인(원천세 3.3%)':(d.ttype==='사업자'?'사업자(계산서)':'-'))+'</span></div>'
    +'<div class="flow-r neg tot"><span>실지급액</span><span class="fv">'+won(payReal)+'</span></div></div>'
    +'</div>'
    +'<div class="dnote">💳 선생님 계좌·은행은 개인정보라 표시하지 않습니다. <b>완전판 마스터(채팅창)</b>에서 확인하세요.</div>';

  return '<div class="dwrap">'
    +'<div class="dhead"><h3>'+esc(d.name)+'</h3>'
    +(d.packfile?('<a class="d-pack" href="'+esc(d.packfile)+'">📦 패킹 운영</a>'):'')
    +'<button class="d-edit" onclick="editClass(\''+cid+'\',this)">✏ 클래스 수정</button></div>'
    +'<div class="dtabs">'
    +'<button class="dtab on" onclick="tabTo(\''+cid+'\',0,this)">정보</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',1,this)">과정개요서</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',2,this)">레시피</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',3,this)">구성품(원가)</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',4,this)">발주·입고</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',5,this)">수익 구조</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',6,this)">선생님 정산</button>'
    +'<button class="dtab" onclick="tabTo(\''+cid+'\',7,this)">모집 현황</button></div>'
    +'<div class="dbody">'
    +'<div class="dpane on">'+p1+'</div><div class="dpane">'+p2+'</div><div class="dpane">'+pRec+'</div>'
    +'<div class="dpane">'+p3+'</div><div class="dpane">'+p3o+'</div><div class="dpane">'+p4+'</div><div class="dpane">'+p6+'</div><div class="dpane">'+p5+'</div></div></div>';
}
function drowFlow(k,v){if(v==null||v==='')return '';return '<div class="flow-r"><span>'+k+'</span><span class="fv">'+esc(v)+'</span></div>';}
function payBadge(v,kind){if(!v)return '';var done=(v===kind+'완료');return '<span class="fh-badge '+(done?'fhb-done':'fhb-wait')+'">'+esc(v)+'</span>';}
function copyRecipe(cid,btn){var d=DETAIL[cid];if(!d)return;var c=d.content||{};if(!c['레시피'])return;var NL=String.fromCharCode(10);var txt='[레시피] '+(d.name||'')+NL+NL+c['레시피'];
  var ok=function(){var o=btn.textContent;btn.textContent='✓ 복사됨';setTimeout(function(){btn.textContent=o;},1500);};
  if(navigator.clipboard)navigator.clipboard.writeText(txt).then(ok,function(){var t=document.createElement('textarea');t.value=txt;document.body.appendChild(t);t.select();document.execCommand('copy');t.remove();ok();});}
function copyOverview(cid,btn){var d=DETAIL[cid];if(!d)return;var c=d.content||{};var NL=String.fromCharCode(10);var L=['[과정개요서] '+(d.name||'')];
  var add=function(k,v){if(v)L.push(NL+'◆ '+k+NL+v);};
  add('강의명',c['강의명']);add('소개문',c['소개문']);add('강의내용',c['강의내용']);add('키트 구성품',c['키트구성품']);add('개별 준비물',c['개별준비물']);
  if(d.kind==='온라인'){add('배송 안내',c['배송안내']);add('참여 안내',c['참여안내']);}else{add('장소',c['장소상세']);add('동반인 규정',c['동반인규정']);}
  add('강사 이력',c['강사이력']);add('해시태그',c['해시태그']);
  var txt=L.join(NL);var ok=function(){var o=btn.textContent;btn.textContent='✓ 복사됨';setTimeout(function(){btn.textContent=o;},1500);};
  if(navigator.clipboard)navigator.clipboard.writeText(txt).then(ok,function(){var t=document.createElement('textarea');t.value=txt;document.body.appendChild(t);t.select();document.execCommand('copy');t.remove();ok();});}
function editClass(cid,btn){var d=DETAIL[cid];var s=(d?d.name:'')+'('+cid+') 계산기 불러오기 코드 만들어줘';
  if(navigator.clipboard)navigator.clipboard.writeText(s).catch(function(){});
  var o=btn.textContent;btn.textContent='✓ 문장 복사됨 → 대화창';setTimeout(function(){btn.textContent=o;},2200);window.open(CHAT,'_blank');}
render();
</script></body></html>"""
listdoc = (LISTTMPL.replace("__VER__",VERSION).replace("__UPD__",UPDATED).replace("__CHAT__",CHAT_URL)
    .replace("__LISTDATA__",LISTDATA).replace("__ORDS__",ORDS).replace("__DETAIL__",DETAIL))
open("class-list.html","w",encoding="utf-8").write(listdoc)
print(f"class-list.html {VERSION} — 아코디언 7탭 목록 {len(sessions)}건")
